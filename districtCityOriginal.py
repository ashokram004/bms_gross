import json
import os
import random
import time
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
from utils.generateImageReport import generate_city_image_report

# Example URL provided by you
url = "https://www.district.in/movies/orange-2010-movie-tickets-in-hyderabad-MV160920"
MOVIE_DATE = "2026-02-09"


url += "?fromdate=" + MOVIE_DATE

def get_driver():
    options = Options()
    options.add_argument("--headless")
    options.add_argument("start-maximized")
    return webdriver.Chrome(options=options)

def get_district_seat_layout(driver, cinema_id, session_id):
    api_url = "https://www.district.in/gw/consumer/movies/v1/select-seat?version=3&site_id=1&channel=mweb&child_site_id=1&platform=district"
    payload = json.dumps({"cinemaId": int(cinema_id), "sessionId": str(session_id)})
    
    guest_token = str(random.randint(1, 9999999999))
    
    js = f"""
    var cb = arguments[0];
    var xhr = new XMLHttpRequest();
    xhr.open("POST", "{api_url}", true);
    xhr.setRequestHeader("Content-Type", "application/json");
    xhr.setRequestHeader("x-guest-token", "{guest_token}");
    xhr.onload = function() {{ cb(xhr.responseText); }};
    xhr.onerror = function() {{ cb(null); }};
    xhr.send('{payload}');
    """
    try:
        resp = driver.execute_async_script(js)
        if resp:
            return json.loads(resp)
    except Exception as e:
        print("⚠️  Error fetching seat layout via API." + str(e))
        pass
    return None

def extract_district_data(url):
    driver = get_driver()
    try:
        driver.get(url)
        html = driver.page_source

        # Locate the Next.js data script
        marker = 'id="__NEXT_DATA__"'
        start_idx = html.find(marker)
        if start_idx == -1:
            raise ValueError("Could not find __NEXT_DATA__ in page.")
        
        start_json = html.find('>', start_idx) + 1
        end_json = html.find('</script>', start_json)
        raw_json = html[start_json:end_json]
        
        full_data = json.loads(raw_json)
        
        # Path based on your shared structure
        # props -> pageProps -> data -> serverState -> movieSessions -> [key] -> pageData -> nearbyCinemas
        sessions_data = full_data['props']['pageProps']['data']['serverState']['movieSessions']
        
        results = []
        grand_total = 0

        # 1. Access the dynamic event key data
        dynamic_key = list(sessions_data.keys())[0]
        event_data = sessions_data[dynamic_key]

        # 2. Iterate through arrangedSessions instead of nearbyCinemas
        arranged_sessions = event_data.get('arrangedSessions', [])

        for cinema_data in arranged_sessions:
            venue_name = cinema_data.get('entityName') # e.g., "KR Complex, Chilakaluripet"
            
            for session in cinema_data.get('sessions', []):
                sessionId = session['sid']
                cid = session['cid']
                show_time = session['showTime']
                
                # Build Price Map from cached data
                price_map = {}
                for area in session.get('areas', []):
                    price_map[area['code']] = float(area['price'])

                session_total_seats = 0
                session_booked_seats = 0
                session_booked_gross = 0
                session_potential_gross = 0
                
                # Try fetching accurate seat layout
                layout_res = None
                if cid:
                    layout_res = get_district_seat_layout(driver, cid, sessionId)
                
                if layout_res and 'seatLayout' in layout_res:
                    col_areas = layout_res['seatLayout'].get('colAreas', {})
                    obj_areas = col_areas.get('objArea', [])
                    
                    for area in obj_areas:
                        area_code = area.get('AreaCode')
                        price = area.get('AreaPrice', price_map.get(area_code, 0)) # Fallback to cached price
                        
                        for row in area.get('objRow', []):
                            for seat in row.get('objSeat', []):
                                status = seat.get('SeatStatus')
                                session_total_seats += 1
                                session_potential_gross += price
                                
                                if status != '0' and status != 0: # Booked (0 is available)
                                    session_booked_seats += 1
                                    session_booked_gross += price
                else:
                    # Fallback to cached data
                    print("⚠️  Could not fetch seat layout, using cached data for gross calculation.")
                    for area in session.get('areas', []):
                        total = area['sTotal']
                        avail = area['sAvail']
                        price = area['price']
                        
                        booked = total - avail
                        session_total_seats += total
                        session_booked_seats += booked
                        session_booked_gross += (booked * price)
                        session_potential_gross += (total * price)
                
                occ = round((session_booked_seats / session_total_seats) * 100, 2) if session_total_seats > 0 else 0
                
                show_data = {
                    "venue": venue_name,
                    "showTime": show_time,
                    "total_tickets": session_total_seats,
                    "booked_tickets": session_booked_seats,
                    "occupancy": occ,
                    "total_gross": int(session_potential_gross),
                    "booked_gross": int(session_booked_gross)
                }
                results.append(show_data)
                grand_total += int(session_booked_gross)
                
                print(f"🎬 {venue_name} | {show_time} | Gross: ₹{session_booked_gross}")

    finally:
        driver.quit()

    return results, grand_total


def generate_excel(results, total):
    reports_dir = "reports"
    os.makedirs(reports_dir, exist_ok=True)
    wb = Workbook()

    # ================= SHEET 1 : THEATRE WISE COLLECTIONS =================
    theatre_sheet = wb.active
    theatre_sheet.title = "Theatre Wise Collections"
    theatre_sheet.append([
        "Venue", "Show count", "Total Seats", 
        "Booked Seats", "Occupancy %", 
        "Total Gross", "Booked Gross"
    ])

    theatre_data = {}
    for r in results:
        v = r["venue"]
        if v not in theatre_data:
            theatre_data[v] = {
                "num_shows": 0, "total_tickets": 0, "booked_tickets": 0, 
                "occupancies": [], "total_gross": 0, "booked_gross": 0
            }
        theatre_data[v]["num_shows"] += 1
        theatre_data[v]["total_tickets"] += r["total_tickets"]
        theatre_data[v]["booked_tickets"] += r["booked_tickets"]
        theatre_data[v]["occupancies"].append(r["occupancy"])
        theatre_data[v]["total_gross"] += r["total_gross"]
        theatre_data[v]["booked_gross"] += r["booked_gross"]

    for venue, data in theatre_data.items():
        avg_occ = round(sum(data["occupancies"]) / data["num_shows"], 2)
        theatre_sheet.append([
            venue, data["num_shows"], data["total_tickets"], 
            data["booked_tickets"], avg_occ, data["total_gross"], data["booked_gross"]
        ])

    # ================= SHEET 2 : SHOW WISE COLLECTIONS =================
    sheet2 = wb.create_sheet(title="Show Wise Collections")
    sheet2.append([
        "Venue", "Show Time", "Total Seats", 
        "Booked Seats", "Occupancy %", 
        "Total Gross", "Booked Gross"
    ])
    for r in results:
        sheet2.append([
            r["venue"], r["showTime"], r["total_tickets"], 
            r["booked_tickets"], r["occupancy"], r["total_gross"], r["booked_gross"]
        ])

    # ================= SHEET 3 : SUMMARY (The Missing Piece!) =================
    summary = wb.create_sheet(title="Summary")
    
    # Calculate Aggregates
    total_theatres = len(theatre_data)
    total_shows = len(results)
    agg_total_seats = sum(r["total_tickets"] for r in results)
    agg_booked_seats = sum(r["booked_tickets"] for r in results)
    agg_potential_gross = sum(r["total_gross"] for r in results)
    agg_booked_gross = total # Using the grand_total passed to the function
    overall_occ = round((agg_booked_seats / agg_total_seats) * 100, 2) if agg_total_seats > 0 else 0
    
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Add Data to Summary
    summary.append(["Metric", "Value"])
    summary.append(["Total Theatres", total_theatres])
    summary.append(["Total Shows", total_shows])
    summary.append(["Overall Occupancy (%)", overall_occ])
    summary.append(["Total Potential Gross (₹)", agg_potential_gross])
    summary.append(["Total Booked Gross (₹)", agg_booked_gross])
    summary.append(["Report Generated At", timestamp])

    # Save File
    file_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = os.path.join(reports_dir, f"district_city_report_{file_ts}.xlsx")
    wb.save(filename)
    print(f"📊 Full Report with Summary Saved: {filename}")


# EXECUTION
results, total = extract_district_data(url)
generate_excel(results, total)

img_path = f"reports/district_city_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
generate_city_image_report(results, url, img_path, "district")