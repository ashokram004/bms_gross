import json
import os
import time
import random
from datetime import datetime
from base64 import b64decode
from Crypto.Cipher import AES
from Crypto.Util.Padding import unpad
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from fake_useragent import UserAgent

from utils.generateHybridCityImageReport import generate_hybrid_city_image_report

processed_sids = set()

# ================= CONFIGURATION =================
# User Inputs
DISTRICT_URL = "https://www.district.in/movies/mana-shankara-varaprasad-garu-movie-tickets-in-vizag-MV203929"
BMS_URL = "https://in.bookmyshow.com/movies/vizag-visakhapatnam/mana-shankara-vara-prasad-garu/buytickets/ET00457184/20260127"

SHOW_DATE = "2026-01-27"  # Ensure this matches the date in BMS URL

# Appended URL for District API
DISTRICT_FULL_URL = f"{DISTRICT_URL}?fromdate={SHOW_DATE}"

# BMS Settings
BMS_KEY = "kYp3s6v9y$B&E)H+MbQeThWmZq4t7w!z"
BOOKED_CODES = {"2"}
SLEEP_TIME = 1.0

# ================= SHARED DRIVER =================
def get_driver():
    ua = UserAgent()
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("start-maximized")
    options.add_argument(f"user-agent={ua.random}")
    options.add_argument("--disable-web-security")
    return webdriver.Chrome(options=options)

# ================= DISTRICT LOGIC =================
def fetch_district_data(driver):
    print("\n🚀 STARTING DISTRICT FETCH...")
    results = []
    
    try:
        driver.get(DISTRICT_FULL_URL)
        time.sleep(1.5)
        html = driver.page_source
        
        marker = 'id="__NEXT_DATA__"'
        start = html.find(marker)
        if start == -1: return []
        
        start = html.find('>', start) + 1
        end = html.find('</script>', start)
        data = json.loads(html[start:end])
        
        sessions = data['props']['pageProps']['data']['serverState']['movieSessions']
        key = list(sessions.keys())[0]
        # Use 'arrangedSessions' for single city page
        cinemas = sessions[key].get('arrangedSessions', [])

        for cin in cinemas:
            venue = cin['entityName']
            for s in cin.get('sessions', []):
                sid = str(s.get('sid', ''))

                if sid in processed_sids: continue
                processed_sids.add(sid)

                b_gross, p_gross, b_tkts, t_tkts = 0, 0, 0, 0
                for a in s.get('areas', []):
                    tot, av, pr = a['sTotal'], a['sAvail'], a['price']
                    bk = tot - av
                    b_tkts += bk; t_tkts += tot
                    b_gross += (bk*pr); p_gross += (tot*pr)
                
                occ = round((b_tkts/t_tkts)*100, 2) if t_tkts else 0
                
                # LOGGING
                print(f"   🎬 {venue[:20]:<20} | {s['showTime']} | Occ: {occ:>5}% | Gross: ₹{b_gross:<8,}")

                results.append({
                    "source": "district", "sid": sid,
                    "venue": venue, "showTime": s['showTime'],
                    "total_tickets": t_tkts, "booked_tickets": b_tkts,
                    "total_gross": p_gross, "booked_gross": b_gross,
                    "occupancy": occ
                })
        
        print(f"✅ District: Found {len(results)} shows.")
    except Exception as e:
        print(f"❌ District Error: {e}")
        
    return results

# ================= BMS LOGIC =================
def decrypt_data(enc):
    try:
        decoded = b64decode(enc)
        cipher = AES.new(BMS_KEY.encode(), AES.MODE_CBC, iv=bytes(16))
        return unpad(cipher.decrypt(decoded), AES.block_size).decode()
    except: return None

def calculate_bms_collection(decrypted, price_map):
    header, rows_part = decrypted.split("||")
    rows = rows_part.split("|")
    
    # Category Map
    cat_map = {}
    for p in header.split("|"):
        parts = p.split(":")
        if len(parts) >= 3: cat_map[parts[1]] = parts[2]

    seats, booked = {}, {}
    for row in rows:
        if not row: continue
        parts = row.split(":")
        if len(parts) < 3: continue
        block = parts[3][0] if len(parts) > 3 else parts[2][0]
        area = cat_map.get(block)
        if not area: continue

        for seat in parts:
            if len(seat) < 2: continue
            status = seat[1]
            if seat[0] == block and status in ("1", "2"):
                seats[area] = seats.get(area, 0) + 1
            if status in BOOKED_CODES:
                booked[area] = booked.get(area, 0) + 1

    t_tkts, b_tkts, t_gross, b_gross = 0, 0, 0, 0
    for area, total in seats.items():
        bk = booked.get(area, 0)
        pr = price_map.get(area, 0)
        t_tkts += total; b_tkts += bk
        t_gross += total * pr; b_gross += bk * pr

    occ = round((b_tkts / t_tkts) * 100, 2) if t_tkts else 0
    return t_tkts, b_tkts, int(t_gross), int(b_gross), occ

def get_seat_layout(driver, venue_code, session_id):
    """ Returns tuple: (EncryptedDataString, ErrorMessage) """
    api = "https://services-in.bookmyshow.com/doTrans.aspx"
    js = f"""
    var cb = arguments[0]; var x = new XMLHttpRequest();
    x.open("POST", "{api}", true);
    x.setRequestHeader("Content-Type", "application/x-www-form-urlencoded");
    x.onload = function() {{ cb(x.responseText); }};
    x.send("strCommand=GETSEATLAYOUT&strAppCode=WEB&strVenueCode={venue_code}&strParam1={session_id}&strParam2=WEB&strParam5=Y&strFormat=json");
    """
    
    max_retries = 2
    for attempt in range(max_retries + 1):
        try:
            resp = driver.execute_async_script(js)
            j_resp = json.loads(resp).get("BookMyShow", {})
            
            if j_resp.get("blnSuccess") == "true":
                return j_resp.get("strData"), None
            
            error_msg = j_resp.get("strException", "")
            if "Rate limit" in error_msg:
                if attempt < max_retries:
                    time.sleep(60)
                    continue
                else:
                    return None, "Rate limit exceeded"
            
            return None, error_msg

        except Exception as e:
            return None, str(e)
            
    return None, "Unknown Error"

def fetch_bms_data():
    print("\n🚀 STARTING BMS FETCH...")
    results = []
    driver = get_driver()
    
    try:
        driver.get(BMS_URL)
        time.sleep(2.5)
        html = driver.page_source
        
        marker = "window.__INITIAL_STATE__"
        start = html.find(marker)
        if start == -1: return []
        
        start = html.find("{", start)
        brace, end = 0, start
        while end < len(html):
            if html[end] == "{": brace += 1
            elif html[end] == "}": brace -= 1
            if brace == 0: break
            end += 1
        
        state_data = json.loads(html[start:end+1])
        
        venues = []
        try:
            sbe = state_data.get("showtimesByEvent")
            dc = sbe.get("currentDateCode")
            widgets = sbe["showDates"][dc]["dynamic"]["data"]["showtimeWidgets"]
            for w in widgets:
                if w.get("type") == "groupList":
                    for g in w["data"]:
                        if g.get("type") == "venueGroup": venues = g["data"]
        except: venues = []

        for v in venues:
            v_name = v["additionalData"]["venueName"]
            v_code = v["additionalData"]["venueCode"]
            
            # 1. Init Capacity Map for THIS Venue
            screen_capacity_map = {}

            # 2. Sort Shows: Available (3,2,1) First, Sold Out (0) Last
            # This helps us learn capacity from available shows before hitting sold out ones
            shows = v.get("showtimes", [])
            shows.sort(key=lambda s: s["additionalData"].get("availStatus", "0"), reverse=True)

            for show in shows:
                sid = show["additionalData"]["sessionId"]
                show_time = show["title"]

                if sid in processed_sids: continue
                processed_sids.add(sid)

                # 3. Determine Screen Name (or default to Main Screen)
                raw_screen = show.get("screenAttr", "")
                screenName = raw_screen if raw_screen else "Main Screen"

                cats = show["additionalData"].get("categories", [])
                price_map = {c["areaCatCode"]: float(c["curPrice"]) for c in cats}
                
                try:
                    # Get Data & Error
                    enc, error_msg = get_seat_layout(driver, v_code, sid)
                    
                    data = None
                    soldOut = False
                    
                    if not enc:
                        # FAILURE HANDLING
                        if not price_map: continue
                        max_price = max(price_map.values())

                        # SMART FALLBACK LOGIC
                        if error_msg and "sold out" in error_msg.lower():
                            # Case 1: Sold Out -> 100%
                            
                            # Check if we know the capacity for this screen
                            if screenName in screen_capacity_map:
                                FALLBACK_SEATS = screen_capacity_map[screenName]
                                print(f"   ⚡ Smart Fallback: Using {FALLBACK_SEATS} seats for {screenName}")
                            else:
                                FALLBACK_SEATS = 400
                                print(f"   ⚠️ No history for {screenName}, using default {FALLBACK_SEATS}")

                            t_tkts = FALLBACK_SEATS; b_tkts = FALLBACK_SEATS
                            t_gross = int(t_tkts * max_price)
                            b_gross = t_gross
                            occ = 100.0
                            soldOut = True
                            
                            data = {
                                "total_tickets": t_tkts, "booked_tickets": b_tkts,
                                "total_gross": t_gross, "booked_gross": b_gross, "occupancy": occ
                            }
                        
                        elif error_msg and "Rate limit" in error_msg:
                            # Case 2: Rate Limit -> Skip
                            print(f"   ⚠️ Skipping {v_name[:15]} (Rate Limit)")
                            continue
                        
                        else:
                            # Case 3: Other -> 50%
                            t_tkts = 400; b_tkts = 200
                            t_gross = int(t_tkts * max_price)
                            b_gross = int(b_tkts * max_price)
                            occ = 50.0
                            data = {
                                "total_tickets": t_tkts, "booked_tickets": b_tkts,
                                "total_gross": t_gross, "booked_gross": b_gross, "occupancy": occ
                            }

                    else:
                        # SUCCESS HANDLING
                        decrypted = decrypt_data(enc)
                        res = calculate_bms_collection(decrypted, price_map)
                        data = {
                            "total_tickets": res[0], "booked_tickets": res[1],
                            "total_gross": res[2], "booked_gross": res[3], "occupancy": res[4]
                        }
                        
                        # Cache Capacity for this Screen
                        if data["total_tickets"] > 0:
                            screen_capacity_map[screenName] = data["total_tickets"]

                    if data:
                        tag = "(SOLD OUT)" if soldOut else ""
                        print(f"   🎬 {v_name[:15]:<15} | {show_time} | Occ: {data['occupancy']:>5}% | Gross: ₹{data['booked_gross']:<8,} {tag}")

                        results.append({
                            "source": "bms", "sid": str(sid),
                            "venue": v_name, "showTime": show_time,
                            "total_tickets": abs(data["total_tickets"]), 
                            "booked_tickets": min(abs(data["booked_tickets"]), abs(data["total_tickets"])),
                            "total_gross": abs(data["total_gross"]), 
                            "booked_gross": min(abs(data["booked_gross"]), abs(data["total_gross"])),
                            "occupancy": min(100, abs(data["occupancy"]))
                        })
                except Exception: pass
                time.sleep(SLEEP_TIME)
        
        print(f"✅ BMS: Found {len(results)} shows.")

    except Exception as e:
        print(f"❌ BMS Error: {e}")
    finally:
        driver.quit()
        
    return results

# ================= EXCEL GENERATOR =================
def generate_excel(data, filename):
    wb = Workbook()
    reports_dir = "reports"
    os.makedirs(reports_dir, exist_ok=True)

    # 1. THEATRE WISE
    ws_th = wb.active
    ws_th.title = "Theatre Wise"
    ws_th.append(["Source", "Venue", "Shows", "Total Seats", "Booked Seats", "Total Gross ₹", "Booked Gross ₹", "Occ %"])
    
    th_map = {}
    for r in data:
        k = r["venue"]
        if k not in th_map: th_map[k] = {"shows":0, "t_seats":0, "b_seats":0, "p_gross":0, "b_gross":0, "src": r["source"]}
        d = th_map[k]
        d["shows"] += 1; d["t_seats"] += r["total_tickets"]; d["b_seats"] += r["booked_tickets"]
        d["p_gross"] += r["total_gross"]; d["b_gross"] += r["booked_gross"]

    for v, d in th_map.items():
        occ = round((d["b_seats"] / d["t_seats"]) * 100, 2) if d["t_seats"] else 0
        ws_th.append([d["src"], v, d["shows"], d["t_seats"], d["b_seats"], d["p_gross"], d["b_gross"], occ])

    # 2. SHOW WISE
    ws_show = wb.create_sheet(title="Show Wise")
    ws_show.append(["Source", "Venue", "Time", "SID", "Total Seats", "Booked Seats", "Total Gross ₹", "Booked Gross ₹", "Occ %"])
    for r in data:
        ws_show.append([r["source"], r["venue"], r["showTime"], r["sid"], r["total_tickets"], r["booked_tickets"], r["total_gross"], r["booked_gross"], r["occupancy"]])

    # 3. SUMMARY
    ws_sum = wb.create_sheet(title="Summary")
    agg_p_gross = sum(r["total_gross"] for r in data)
    agg_b_gross = sum(r["booked_gross"] for r in data)
    agg_t_seats = sum(r["total_tickets"] for r in data)
    agg_b_seats = sum(r["booked_tickets"] for r in data)
    occ = round((agg_b_seats / agg_t_seats) * 100, 2) if agg_t_seats else 0
    
    ws_sum.append(["Metric", "Value"])
    ws_sum.append(["Total Theatres", len(th_map)])
    ws_sum.append(["Total Shows", len(data)])
    ws_sum.append(["Total Seats", agg_t_seats])
    ws_sum.append(["Booked Tickets", agg_b_seats])
    ws_sum.append(["Total Potential Gross", agg_p_gross])
    ws_sum.append(["Total Booked Gross", agg_b_gross])
    ws_sum.append(["Overall Occupancy %", occ])
    ws_sum.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    path = os.path.join(reports_dir, filename)
    wb.save(path)
    print(f"📊 Report Saved: {path}")

# ================= EXECUTION =================
if __name__ == "__main__":
    d_driver = get_driver()
    dist_data = fetch_district_data(d_driver)
    d_driver.quit()

    bms_data = fetch_bms_data()

    print("\n🔄 Merging Data (Higher Revenue Wins)...")
    merged_map = {}
    
    # 1. District Data Base
    for r in dist_data:
        key = r["sid"] if r["sid"] else f"{r['venue']}_{r['showTime']}"
        merged_map[key] = r

    # 2. BMS Override
    for r in bms_data:
        key = r["sid"] if r["sid"] else f"{r['venue']}_{r['showTime']}"
        if key in merged_map:
            if r["booked_gross"] > merged_map[key]["booked_gross"]:
                merged_map[key] = r
        else:
            merged_map[key] = r

    final_data = list(merged_map.values())
    
    if final_data:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        generate_excel(final_data, f"Total_City_Report_{ts}.xlsx")
        generate_hybrid_city_image_report(final_data, DISTRICT_FULL_URL, f"reports/Total_City_Report_{ts}.png")
    else:
        print("❌ No data collected.")