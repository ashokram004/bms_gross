import json
import os
import sys
import time
import random
import shutil
import threading
import requests
from datetime import datetime, timedelta
from base64 import b64decode
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()
from Crypto.Cipher import AES
from Crypto.Util.Padding import unpad
from fake_useragent import UserAgent
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
from collections import defaultdict, deque
from concurrent.futures import ThreadPoolExecutor, as_completed

from utils.generatePremiumCityImageReport import generate_premium_city_image_report
from utils.generateHybridCityHTMLReport import generate_hybrid_city_html_report
from utils.sendReportEmail import send_collection_report

# =============================================================================
# ── 1. CONFIGURATION ─────────────────────────────────────────────────────────
# =============================================================================

SHOW_DATE = "2026-03-19"

DISTRICT_URL_TEMPLATE = (
    "https://www.district.in/movies/ustaad-bhagat-singh-movie-tickets-in-{city}-MV161614"
    "?frmtid=tvqjmjqme&fromdate=" + SHOW_DATE
)
BMS_URL_TEMPLATE = (
    "https://in.bookmyshow.com/movies/{city}/ustaad-bhagat-singh/buytickets/ET00339939/20260319"
)

DISTRICT_CITIES = [
    "hyderabad"
]
BMS_CITIES = [
    "hyderabad"
]

BMS_KEY      = "kYp3s6v9y$B&E)H+MbQeThWmZq4t7w!z"
BOOKED_CODES = {"2"}

# Performance tuning
DISTRICT_CITY_WORKERS = 12    # parallel city workers for District (pure HTTP)
BMS_DRIVER_POOL_SIZE  = 3     # cities processed in parallel (each gets a fresh Chrome)
DISTRICT_RATE         = 5     # max requests/second to district.in

# =============================================================================
# ── 2. GLOBAL STATE & LOCKS ──────────────────────────────────────────────────
# =============================================================================

# Sets to keep track of processed session IDs (SIDs) across workers
_global_bms_sids = set()
_global_bms_sids_lock = threading.Lock()

_global_district_sids = set()
_global_district_sids_lock = threading.Lock()

VENUE_MAP = {}  # BMS VenueCode -> District cinema_id mapping


# =============================================================================
# ── 3. UTILITIES & HELPERS ───────────────────────────────────────────────────
# =============================================================================

class RateLimiter:
    """Thread-safe, non-blocking rate limiter."""
    def __init__(self, rate):
        self.min_interval = 1.0 / rate
        self._lock = threading.Lock()
        self._next_slot = 0.0

    def acquire(self):
        with self._lock:
            now = time.monotonic()
            if self._next_slot <= now:
                self._next_slot = now + self.min_interval
                return
            wait_until = self._next_slot
            self._next_slot = wait_until + self.min_interval
        sleep_time = wait_until - time.monotonic()
        if sleep_time > 0:
            time.sleep(sleep_time)

def extract_movie_name_from_url(url):
    """Extracts a readable movie name from the given URL."""
    try:
        if '/movies/' in url and '/buytickets/' in url:
            parts = url.split('/movies/')[1].split('/buytickets/')[0].split('/')
            movie_slug = parts[-1] if len(parts) > 1 else parts[0]
            return movie_slug.replace('-', ' ').title()
        if '/movies/' in url and '-movie-tickets-in-' in url:
            movie_slug = url.split('/movies/')[1].split('-movie-tickets-in-')[0]
            return movie_slug.replace('-', ' ').title()
    except Exception as e:
        print(f"Could not extract movie name from URL: {e}")
    return "Movie Collection"

def district_gmt_to_ist(dt_str):
    """Converts a District GMT time string to IST."""
    gmt = datetime.fromisoformat(dt_str)
    ist = gmt + timedelta(hours=5, minutes=30)
    return ist.strftime("%Y-%m-%d %H:%M")

def normalize_bms_time(show_date, show_time):
    """Formats BMS show date and time to a standard format."""
    dt = datetime.strptime(f"{show_date} {show_time}", "%Y-%m-%d %I:%M %p")
    return dt.strftime("%Y-%m-%d %H:%M")

def build_seat_signature(seat_map):
    """Creates a unique signature string based on the seat map counts."""
    return "|".join(str(c) for c in sorted(seat_map.values()))


# =============================================================================
# ── 4. DISTRICT DATA EXTRACTION ──────────────────────────────────────────────
# =============================================================================

district_limiter = RateLimiter(DISTRICT_RATE)
_thread_local = threading.local()

def get_http_session():
    """Returns a thread-local configured requests session for District."""
    if not hasattr(_thread_local, 'session'):
        s = requests.Session()
        ua = UserAgent()
        s.headers.update({
            'User-Agent': ua.random,
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Cache-Control': 'no-cache',
        })
        adapter = requests.adapters.HTTPAdapter(
            pool_connections=20,
            pool_maxsize=20,
            max_retries=requests.adapters.Retry(total=2, backoff_factor=0.5,
                                                 status_forcelist=[502, 503, 504]),
        )
        s.mount('https://', adapter)
        s.mount('http://', adapter)
        _thread_local.session = s
    return _thread_local.session

def get_district_seat_layout_http(cinema_id, session_id):
    """Direct HTTP POST for District seat layout API — no Selenium needed."""
    api_url = "https://www.district.in/gw/consumer/movies/v1/select-seat"
    params  = {
        "version": "3", "site_id": "1", "channel": "mweb",
        "child_site_id": "1", "platform": "district",
    }
    payload = {"cinemaId": int(cinema_id), "sessionId": str(session_id)}
    headers = {
        "Content-Type": "application/json",
        "x-guest-token": str(random.randint(1, 9999999999)),
        "Origin": "https://www.district.in",
        "Referer": "https://www.district.in/",
    }
    try:
        session = get_http_session()
        resp = session.post(api_url, params=params, json=payload,
                            headers=headers, timeout=10)
        if resp.status_code == 200:
            return resp.json()
    except Exception:
        pass
    return None

def district_process_venue_http(cin, city_name):
    """Processes a single District venue via HTTP — no Selenium driver needed.
    Uses global _global_district_sids set to skip already-processed SIDs across all workers.
    """
    results = []
    try:
        venue = cin['entityName']
        for s in cin.get('sessions', []):
            sid = str(s.get('sid', ''))
            cid = s.get('cid')

            # Check if SID already processed (thread-safe global check)
            with _global_district_sids_lock:
                if sid in _global_district_sids:
                    continue
                _global_district_sids.add(sid)

            code_to_label  = {}
            default_prices = {}
            for a in s.get('areas', []):
                code_to_label[a['code']]  = a['label']
                default_prices[a['code']] = float(a['price'])

            b_gross, p_gross, b_tkts, t_tkts = 0, 0, 0, 0
            seat_map        = defaultdict(int)
            label_price_map = {}
            layout_res      = None

            if cid:
                layout_res = get_district_seat_layout_http(cid, sid)

            if layout_res and 'seatLayout' in layout_res:
                for area in layout_res['seatLayout'].get('colAreas', {}).get('objArea', []):
                    area_code = area.get('AreaCode')
                    label     = code_to_label.get(area_code, area_code)
                    price     = float(area.get('AreaPrice') or default_prices.get(area_code, 0))
                    label_price_map[label] = price
                    for row in area.get('objRow', []):
                        for seat in row.get('objSeat', []):
                            status = seat.get('SeatStatus')
                            t_tkts += 1; p_gross += price
                            seat_map[label] += 1
                            if status != '0' and status != 0:
                                b_tkts += 1; b_gross += price
            else:
                for a in s.get('areas', []):
                    tot, av, pr = a['sTotal'], a['sAvail'], a['price']
                    bk = tot - av
                    seat_map[a['label']]       = tot
                    label_price_map[a['label']] = float(pr)
                    b_tkts += bk; t_tkts += tot
                    b_gross += bk * pr; p_gross += tot * pr

            price_seat_map  = defaultdict(int)
            price_seat_list = []
            for label, count in seat_map.items():
                pr = label_price_map.get(label, 0.0)
                price_seat_map[float(pr)] += count
                price_seat_list.append((float(pr), count))

            occ             = round((b_tkts / t_tkts) * 100, 2) if t_tkts else 0
            normalized_time = district_gmt_to_ist(s['showTime'])

            results.append({
                "source":               "district",
                "sid":                  sid,
                "city":                 city_name,
                "venue":                venue,
                "cinema_id":            str(cid) if cid else "",
                "showTime":             s['showTime'],
                "normalized_show_time": normalized_time,
                "seat_category_map":    dict(seat_map),
                "price_seat_map":       dict(price_seat_map),
                "price_seat_signature": sorted(price_seat_list),
                "seat_signature":       build_seat_signature(seat_map),
                "total_tickets":        abs(t_tkts),
                "booked_tickets":       min(abs(b_tkts), abs(t_tkts)),
                "total_gross":          abs(p_gross),
                "booked_gross":         min(abs(int(b_gross)), abs(int(p_gross))),
                "occupancy":            min(100, abs(occ)),
                "is_fallback":          False,
            })
    except Exception as e:
        print(f"   ❌ [District][{city_name}] Venue worker error: {str(e).splitlines()[0]}")
    return results

def fetch_district_city(city_name, city_slug, city_counter_str):
    """
    Fetches all District data for one city via HTTP — no Selenium needed.
    Step 1: HTTP GET page to extract __NEXT_DATA__ JSON.
    Step 2: Process each venue's shows (seat layout via HTTP POST).
    """
    url = DISTRICT_URL_TEMPLATE.format(city=city_slug)

    cinemas = []
    for attempt in range(2):
        try:
            district_limiter.acquire()
            session = get_http_session()
            resp    = session.get(url, timeout=15)

            if resp.status_code == 403 and attempt == 0:
                time.sleep(15)
                if hasattr(_thread_local, 'session'):
                    delattr(_thread_local, 'session')
                continue

            if resp.status_code != 200:
                print(f"   ⚠️  [District] {city_counter_str} {city_name:<15} — HTTP {resp.status_code}")
                return []
            html = resp.text

            marker = 'id="__NEXT_DATA__"'
            idx    = html.find(marker)
            if idx == -1:
                return []

            start    = html.find('>', idx) + 1
            end      = html.find('</script>', start)
            data     = json.loads(html[start:end])
            sessions = data['props']['pageProps']['data']['serverState']['movieSessions']
            if not sessions:
                return []
            key     = list(sessions.keys())[0]
            cinemas = sessions[key].get('arrangedSessions', [])
            break
        except Exception as e:
            print(f"   ❌ [District] {city_counter_str} {city_name:<15} — Error: {str(e).splitlines()[0]}")
            return []

    if not cinemas:
        return []

    city_results = []
    for cin in cinemas:
        results = district_process_venue_http(cin, city_name)
        city_results.extend(results)

    gross = sum(r['booked_gross'] for r in city_results)
    if city_results:
        print(f"   ✅ [District] {city_counter_str} {city_name:<15} | Shows: {len(city_results):<3} | Gross: ₹{gross:<10,}")
    return city_results

def run_district(city_pairs):
    """Executes District scraping for all given cities in parallel."""
    all_results = []
    total       = len(city_pairs)
    completed   = [0]
    lock        = threading.Lock()
    print(f"\n🚀 [District] Starting — {total} cities, {DISTRICT_CITY_WORKERS} workers\n")

    def _wrapped(idx, city_name, city_slug):
        counter_str = f"[{idx}/{total}]"
        results = fetch_district_city(city_name, city_slug, counter_str)
        with lock:
            completed[0] += 1
        return results

    with ThreadPoolExecutor(max_workers=DISTRICT_CITY_WORKERS) as executor:
        futures = {
            executor.submit(_wrapped, idx, city_name, city_slug): city_name
            for idx, (city_name, city_slug) in enumerate(city_pairs, 1)
        }
        for future in as_completed(futures):
            try:
                all_results.extend(future.result())
            except Exception as e:
                print(f"   ❌ [District] City worker error: {str(e).splitlines()[0]}")

    return all_results

# =============================================================================
# ── 5. BMS DATA EXTRACTION ───────────────────────────────────────────────────
# =============================================================================

def _create_chrome_driver():
    """Creates a configured headless Chrome driver for BMS."""
    ua = UserAgent()
    options = Options()
    options.add_argument(f"user-agent={ua.random}")
    options.add_argument("--headless=new")
    options.add_argument("start-maximized")
    options.add_argument("--disable-web-security")
    options.add_argument("--disable-site-isolation-trials")
    options.add_argument("disable-csp")
    options.add_argument("--disable-blink-features=AutomationControlled")
    # Optimize loading by disabling images and notifications
    prefs = {
        "profile.managed_default_content_settings.images": 2,
        "profile.default_content_setting_values.notifications": 2,
    }
    options.add_experimental_option("prefs", prefs)
    return webdriver.Chrome(options=options)

def extract_initial_state_from_page(driver, url):
    """Loads BMS page and extracts the embedded JSON state."""
    try:
        driver.get(url)
        driver.set_script_timeout(12)
        try:
            result = driver.execute_async_script("""
                var cb = arguments[0];
                var attempts = 0;
                function check() {
                    attempts++;
                    try {
                        var s = window.__INITIAL_STATE__;
                        if (s) {
                            if (s.showtimesByEvent && s.showtimesByEvent.currentDateCode) {
                                cb(JSON.stringify(s));
                                return;
                            }
                            if (s.appConfig) {
                                cb(null);
                                return;
                            }
                        }
                    } catch(e) {}
                    if (attempts > 50) { cb(null); return; }
                    setTimeout(check, 200);
                }
                check();
            """)
            if result:
                return json.loads(result)
        except Exception:
            pass
            
        # Fallback to parsing page source
        html = driver.page_source
        marker = "window.__INITIAL_STATE__"
        start = html.find(marker)
        if start == -1: return None
        start = html.find("{", start)
        brace_count = 0; end = start
        while end < len(html):
            if html[end] == "{": brace_count += 1
            elif html[end] == "}": brace_count -= 1
            if brace_count == 0: break
            end += 1
        return json.loads(html[start:end + 1])
    except Exception:
        return None

def extract_venues(state):
    """Extracts venue data from the BMS initial state JSON."""
    if not state: return []
    try:
        sbe = state.get("showtimesByEvent")
        if not sbe: return []
        date_code = sbe.get("currentDateCode")
        if not date_code: return []
        widgets = sbe["showDates"][date_code]["dynamic"]["data"]["showtimeWidgets"]
        for w in widgets:
            if w.get("type") == "groupList":
                for g in w["data"]:
                    if g.get("type") == "venueGroup":
                        return g["data"]
    except Exception:
        pass
    return []

def get_single_seat_layout(driver, venue_code, session_id):
    """Executes an XHR in the browser to get seat layout data for a BMS show."""
    api_url = "https://services-in.bookmyshow.com/doTrans.aspx"
    js = (
        "var cb = arguments[0];"
        "var x = new XMLHttpRequest();"
        "x.open('POST', '%s', true);"
        "x.setRequestHeader('Content-Type', 'application/x-www-form-urlencoded');"
        "x.timeout = 15000;"
        "x.onload = function() { cb(x.responseText); };"
        "x.onerror = function() { cb(null); };"
        "x.ontimeout = function() { cb(null); };"
        "x.send('strCommand=GETSEATLAYOUT&strAppCode=WEB&strVenueCode=%s&lngTransactionIdentifier=0&strParam1=%s&strParam2=WEB&strParam5=Y&strFormat=json');"
    ) % (api_url, venue_code, session_id)

    try:
        driver.set_script_timeout(20)
        resp = driver.execute_async_script(js)
    except Exception as e:
        return None, str(e).split('\n')[0]

    if not resp:
        return None, "Empty response"

    data = json.loads(resp).get("BookMyShow", {})
    if data.get("blnSuccess") == "true":
        return data.get("strData"), None
    return None, data.get("strException", "")

def decrypt_data(enc):
    """Decrypts AES encrypted seat layout data from BMS."""
    decoded = b64decode(enc)
    cipher  = AES.new(BMS_KEY.encode(), AES.MODE_CBC, iv=bytes(16))
    return unpad(cipher.decrypt(decoded), AES.block_size).decode()

def calculate_bms_collection(decrypted, price_map):
    """Calculates tickets and gross from decrypted BMS seat layout string."""
    header, rows_part = decrypted.split("||")
    rows = rows_part.split("|")

    cat_map         = {}
    local_price_map = price_map.copy()
    last_price      = 0.0

    for p in header.split("|"):
        parts = p.split(":")
        if len(parts) >= 3:
            cat_map[parts[1]] = parts[2]
            current_price = local_price_map.get(parts[2], 0.0)
            if current_price > 0:   last_price = current_price
            elif last_price > 0:    local_price_map[parts[2]] = last_price

    seats, booked = {}, {}
    for row in rows:
        if not row: continue
        parts = row.split(":")
        if len(parts) < 3: continue
        block = parts[3][0] if len(parts) > 3 else parts[2][0]
        area  = cat_map.get(block)
        if not area: continue
        for seat in parts:
            if len(seat) < 2: continue
            status = seat[1]
            if seat[0] == block and status in ("1", "2"):
                seats[area] = seats.get(area, 0) + 1
            if seat[0] == block and status in BOOKED_CODES:
                booked[area] = booked.get(area, 0) + 1

    t_tkts, b_tkts, t_gross, b_gross = 0, 0, 0, 0
    for area, total in seats.items():
        bk = booked.get(area, 0); pr = local_price_map.get(area, 0)
        t_tkts += total; b_tkts += bk
        t_gross += total * pr; b_gross += bk * pr

    occ = round((b_tkts / t_tkts) * 100, 2) if t_tkts else 0
    return t_tkts, b_tkts, int(t_gross), int(b_gross), occ, seats, local_price_map

def fetch_bms_city(city_name, city_slug, city_counter_str):
    """Processes all venues and shows for a specific BMS city."""
    url = BMS_URL_TEMPLATE.format(city=city_slug)
    driver = None
    results_all = []
    
    try:
        driver = _create_chrome_driver()
        state_data = extract_initial_state_from_page(driver, url)
        if not state_data:
            print(f"   ⚠️  [BMS] {city_counter_str} {city_name:<15} — skipped (no state data)")
            return []
            
        venues = extract_venues(state_data)
        if not venues:
            print(f"   ⚠️  [BMS] {city_counter_str} {city_name:<15} — skipped (no venues)")
            return []

        for venue in venues:
            v_name = venue["additionalData"]["venueName"]
            v_code = venue["additionalData"]["venueCode"]
            screen_capacity_map = {}
            shows = venue.get("showtimes", [])
            shows.sort(key=lambda s: s["additionalData"].get("availStatus", "0"), reverse=True)

            for show in shows:
                sid = str(show["additionalData"]["sessionId"])
                show_time = show.get("title")
                raw_screen = show.get("screenAttr", "")
                screenName = raw_screen if raw_screen else "Main Screen"

                with _global_bms_sids_lock:
                    if sid in _global_bms_sids: continue
                    _global_bms_sids.add(sid)

                seat_map = {}
                is_fallback = False
                price_seat_map = {}

                try:
                    cats = show["additionalData"].get("categories", [])
                    price_map = {c["areaCatCode"]: float(c["curPrice"]) for c in cats}
                    enc, error_msg = get_single_seat_layout(driver, v_code, sid)
                    data = None

                    if not enc:
                        if not price_map: continue
                        max_price = max(price_map.values())
                        is_fallback = True
                        for p in price_map.values(): price_seat_map[float(p)] = 0

                        if error_msg and "sold out" in error_msg.lower():
                            recovered_capacity = None
                            recovered_seat_map = None
                            if screenName in screen_capacity_map:
                                recovered_seat_map = screen_capacity_map[screenName]
                                recovered_capacity = sum(recovered_seat_map.values()) if isinstance(recovered_seat_map, dict) else recovered_seat_map

                            if not recovered_capacity:
                                try:
                                    base_sid = int(sid)
                                    for offset in range(7, 0, -1):
                                        target_sid = str(base_sid + offset)
                                        time.sleep(1)
                                        n_enc, _ = get_single_seat_layout(driver, v_code, target_sid)
                                        if n_enc:
                                            n_dec = decrypt_data(n_enc)
                                            n_res = calculate_bms_collection(n_dec, {})
                                            if n_res[0] > 0:
                                                recovered_capacity = n_res[0]
                                                recovered_seat_map = n_res[5]
                                                break
                                except Exception:
                                    pass

                            if recovered_capacity:
                                calc_gross = sum(count * price_map.get(ac, 0) for ac, count in recovered_seat_map.items())
                                if calc_gross > 0:
                                    t_tkts = b_tkts = recovered_capacity
                                    t_gross = b_gross = calc_gross
                                    screen_capacity_map[screenName] = recovered_seat_map and sum(recovered_seat_map.values()) or recovered_capacity
                                    seat_map = recovered_seat_map
                                    ps_map = defaultdict(int)
                                    for ac, count in seat_map.items():
                                        ps_map[float(price_map.get(ac, 0))] += count
                                    price_seat_map = dict(ps_map)
                                else:
                                    recovered_capacity = None

                            if not recovered_capacity:
                                FALLBACK_SEATS = screen_capacity_map.get(screenName, 400)
                                t_tkts = b_tkts = FALLBACK_SEATS
                                t_gross = b_gross = int(FALLBACK_SEATS * max_price)

                            occ = 100.0
                            data = {"total_tickets": t_tkts, "booked_tickets": b_tkts,
                                    "total_gross": t_gross, "booked_gross": b_gross, "occupancy": occ}
                        else:
                            if error_msg and "rate limit" in (error_msg or "").lower(): continue
                            t_tkts = 400; b_tkts = 200
                            t_gross = int(t_tkts * max_price); b_gross = int(b_tkts * max_price)
                            data = {"total_tickets": t_tkts, "booked_tickets": b_tkts,
                                    "total_gross": t_gross, "booked_gross": b_gross, "occupancy": 50.0}
                    else:
                        decrypted = decrypt_data(enc)
                        res = calculate_bms_collection(decrypted, price_map)
                        data = {"total_tickets": abs(res[0]), "booked_tickets": min(abs(res[1]), abs(res[0])),
                                "total_gross": abs(res[2]), "booked_gross": min(abs(res[3]), abs(res[2])),
                                "occupancy": min(100, abs(res[4]))}
                        seat_map = res[5]
                        final_price_map = res[6] if len(res) > 5 else {}

                        if data["total_tickets"] > 0:
                            ps_map = defaultdict(int); ps_list = []
                            for ac, count in seat_map.items():
                                pr = float(final_price_map.get(ac, 0)) if final_price_map else float(price_map.get(ac, 0))
                                ps_map[pr] += count; ps_list.append((pr, count))
                            price_seat_map = dict(ps_map)
                            data["price_seat_signature"] = sorted(ps_list)
                            screen_capacity_map[screenName] = data["total_tickets"]

                    if data and data.get('total_tickets', 0) > 0:
                        normalized_time = normalize_bms_time(SHOW_DATE, show_time)
                        data.update({
                            "source": "bms", "sid": sid,
                            "city": city_name,
                            "venue": v_name, "venue_code": v_code, "showTime": show_time,
                            "normalized_show_time": normalized_time,
                            "seat_category_map": seat_map, "price_seat_map": price_seat_map,
                            "price_seat_signature": data.get("price_seat_signature", []),
                            "seat_signature": build_seat_signature(seat_map),
                            "is_fallback": is_fallback,
                        })
                        results_all.append(data)
                except Exception:
                    pass
                try: time.sleep(1)
                except Exception: pass
    except Exception as e:
        print(f"   ❌ [BMS] {city_counter_str} {city_name:<15} — Error: {str(e).splitlines()[0]}")
    finally:
        if driver:
            try: driver.quit()
            except Exception: pass

    if results_all:
        gross = sum(r.get('booked_gross', 0) for r in results_all)
        print(f"   ✅ [BMS] {city_counter_str} {city_name:<15} | Shows: {len(results_all):<3} | Gross: ₹{gross:<10,}")

    return results_all

def run_bms(city_pairs):
    """Executes BMS scraping for all given cities using parallel browser workers."""
    all_results = []
    total = len(city_pairs)
    workers = min(BMS_DRIVER_POOL_SIZE, total)
    print(f"\n🚀 [BMS] Starting — {total} cities, {workers} parallel browser workers\n")

    def _process_city(args):
        idx, (city_name, city_slug) = args
        counter_str = f"[{idx}/{total}]"
        return fetch_bms_city(city_name, city_slug, counter_str)

    with ThreadPoolExecutor(max_workers=workers) as city_pool:
        futures = [
            city_pool.submit(_process_city, (idx, city_info))
            for idx, city_info in enumerate(city_pairs, 1)
        ]
        for f in as_completed(futures):
            try:
                all_results.extend(f.result())
            except Exception as e:
                print(f"   ❌ [BMS] Worker error: {str(e).splitlines()[0]}")

    return all_results

# =============================================================================
# ── 6. DATA MERGING & DEDUPLICATION ──────────────────────────────────────────
# =============================================================================

def load_venue_mapping():
    """Loads mapping to correlate BMS venues with District venues."""
    global VENUE_MAP
    mapping_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'utils', 'venue_mapping.json')
    if os.path.exists(mapping_path):
        with open(mapping_path, encoding='utf-8') as f:
            data = json.load(f)
        VENUE_MAP = data.get('bms_to_district', {})
        print(f"📍 Loaded venue mapping: {len(VENUE_MAP)} BMS→District pairs")
    else:
        print("⚠️  Venue mapping not found. Venue-based matching disabled.")

def _is_same_venue(bms_show, dist_show):
    """Checks if a BMS show and District show map to the same physical venue."""
    bms_code = bms_show.get('venue_code', '')
    dist_cid = dist_show.get('cinema_id', '')
    if not bms_code or not dist_cid:
        return 'unmapped'
    mapped_dist = VENUE_MAP.get(bms_code)
    if mapped_dist is None:
        return 'unmapped'
    return mapped_dist == dist_cid

def dedup_same_platform(records, source_label):
    """Removes duplicate shows within the same platform based on SID."""
    seen = {}
    best = []
    dropped = 0
    for r in records:
        sid = r.get('sid', '')
        if not sid:
            best.append(r); continue
        if sid not in seen:
            seen[sid] = len(best)
            best.append(r)
        else:
            dropped += 1
            existing = best[seen[sid]]
            if r.get('booked_gross', 0) > existing.get('booked_gross', 0):
                best[seen[sid]] = r
    if dropped:
        print(f"   ♻️  [{source_label}] Dedup removed {dropped} duplicate(s).")
    return best

def merge_data(all_dist_data, all_bms_data):
    """Merges District and BMS datasets, preventing duplicate shows across platforms."""
    all_dist_data = dedup_same_platform(all_dist_data, "District")
    all_bms_data  = dedup_same_platform(all_bms_data,  "BMS")

    print(f"\n🔄 Merging {len(all_dist_data)} District + {len(all_bms_data)} BMS shows...")

    final_data = []
    SEAT_TOLERANCE = 5
    district_index = defaultdict(list)
    
    for r in all_dist_data:
        district_index[(r['city_order'], r['normalized_show_time'])].append(r)

    for bms in all_bms_data:
        key = (bms['city_order'], bms['normalized_show_time'])
        candidates = district_index.get(key, [])
        match = None

        for c in candidates:
            # 1. Exact Match: Session ID matching between BMS and District. Highest confidence.
            if c['sid'] == bms['sid']:
                match = c
                print(f"   🔗 SID Match: {bms['sid']}")
                break

        if not match and not bms.get('is_fallback', False):
            # 2. Strong Deduplication: Compare exact prices and the count of seats per price category, combined with venue validation.
            b_sig = bms.get('price_seat_signature', [])
            for c in candidates:
                d_sig = c.get('price_seat_signature', [])
                if b_sig and d_sig and len(b_sig) == len(d_sig):
                    if all(bp == dp and abs(bs - ds) <= SEAT_TOLERANCE for (bp, bs), (dp, ds) in zip(b_sig, d_sig)):
                        if _is_same_venue(bms, c) is True:
                            match = c
                            print(f"   🔗 Price/Seat Sig + Venue Map: {bms['venue']} == {c['venue']}")
                            break

        if not match and not bms.get('is_fallback', False):
            # 3. Moderate Deduplication: Compare total seat counts per category, without price, combined with venue validation.
            b_seats = sorted(bms.get('seat_category_map', {}).values())
            for c in candidates:
                d_seats = sorted(c.get('seat_category_map', {}).values())
                if b_seats and d_seats and len(b_seats) == len(d_seats):
                    if all(abs(bs - ds) <= SEAT_TOLERANCE for bs, ds in zip(b_seats, d_seats)):
                        if _is_same_venue(bms, c) is True:
                            match = c
                            print(f"   🔗 Seat Sig + Venue Map: {bms['venue']} == {c['venue']}")
                            break

        if not match and candidates:
            # 4. Weak Deduplication: Fall back to validating if both shows share the exact same ticket price points and venue.
            b_prices = {p for p in bms.get('price_seat_map', {}).keys() if p > 0}
            for c in candidates:
                d_prices = {p for p in c.get('price_seat_map', {}).keys() if p > 0}
                if b_prices == d_prices and _is_same_venue(bms, c) is True:
                    match = c
                    print(f"   🔗 Venue Map + Price: {bms['venue']} == {c['venue']}")
                    break

        if not match and candidates:
            # 5. Venue-Only Deduplication: Last resort, purely based on time and mapped venue.
            for c in candidates:
                if _is_same_venue(bms, c) is True:
                    match = c
                    print(f"   🔗 Venue Map Only: {bms['venue']} == {c['venue']}")
                    break

        if match:
            candidates.remove(match)
            # If BMS data is genuinely scraped (no fallbacks), it is the most accurate source of truth.
            # If BMS used a fallback, District is preferred as it might have a better recent cache.
            if not bms.get('is_fallback', False):
                match.update({
                    'total_tickets': bms['total_tickets'],
                    'booked_tickets': bms['booked_tickets'],
                    'total_gross': bms['total_gross'],
                    'booked_gross': bms['booked_gross'],
                    'occupancy': bms['occupancy'],
                    'seat_category_map': bms['seat_category_map'],
                    'price_seat_map': bms['price_seat_map'],
                    'seat_signature': bms['seat_signature'],
                })
            match['bms_sid'] = bms['sid']
            match['district_sid'] = match['sid']
            final_data.append(match)
        else:
            bms['bms_sid'] = bms['sid']
            bms['district_sid'] = None
            final_data.append(bms)

    for sublist in district_index.values():
        for show in sublist:
            show['bms_sid'] = None
            show['district_sid'] = show['sid']
            final_data.append(show)

    print(f"✅ Merge complete — {len(final_data)} final shows.")
    return final_data


# =============================================================================
# ── 7. REPORT EXPORT ─────────────────────────────────────────────────────────
# =============================================================================

def generate_excel(data, filename):
    wb          = Workbook()
    reports_dir = "reports"
    os.makedirs(reports_dir, exist_ok=True)

    # Theatre Wise
    ws_th = wb.active
    ws_th.title = "Theatre Wise"
    ws_th.append(["Venue","Shows","Total Seats","Booked Seats","Total Gross","Booked Gross","Occ %"])
    th_map = {}
    for r in data:
        k = r["venue"]
        if k not in th_map:
            th_map[k] = {"shows":0,"t_seats":0,"b_seats":0,"p_gross":0,"b_gross":0}
        d = th_map[k]
        d["shows"]+=1; d["t_seats"]+=r["total_tickets"]; d["b_seats"]+=r["booked_tickets"]
        d["p_gross"]+=r["total_gross"]; d["b_gross"]+=r["booked_gross"]
    for v, d in th_map.items():
        occ = round((d["b_seats"]/d["t_seats"])*100,2) if d["t_seats"] else 0
        ws_th.append([v,d["shows"],d["t_seats"],d["b_seats"],d["p_gross"],d["b_gross"],occ])

    # City Wise
    ws_city = wb.create_sheet(title="City Wise")
    ws_city.append(["City","Theatres","Shows","Total Seats","Booked Seats","Total Gross","Booked Gross","Occ %"])
    city_map = {}
    for r in data:
        k = r["city"]
        if k not in city_map:
            city_map[k] = {"shows":0,"t_seats":0,"b_seats":0,"p_gross":0,"b_gross":0,"venues":set()}
        d = city_map[k]
        d["shows"]+=1; d["t_seats"]+=r["total_tickets"]; d["b_seats"]+=r["booked_tickets"]
        d["p_gross"]+=r["total_gross"]; d["b_gross"]+=r["booked_gross"]; d["venues"].add(r["venue"])
    for city, d in city_map.items():
        occ = round((d["b_seats"]/d["t_seats"])*100,2) if d["t_seats"] else 0
        ws_city.append([city,len(d["venues"]),d["shows"],d["t_seats"],d["b_seats"],d["p_gross"],d["b_gross"],occ])

    # Show Wise
    ws_show = wb.create_sheet(title="Show Wise")
    ws_show.append(["Source","City","Venue","Time","SID","Total Seats","Booked Seats","Total Gross","Booked Gross","Occ %"])
    for r in data:
        ws_show.append([r["source"],r["city"],r["venue"],r["normalized_show_time"],r["sid"],
                        r["total_tickets"],r["booked_tickets"],r["total_gross"],r["booked_gross"],r["occupancy"]])

    # Summary
    ws_sum = wb.create_sheet(title="Summary")
    agg_t  = sum(r["total_tickets"]  for r in data)
    agg_b  = sum(r["booked_tickets"] for r in data)
    agg_bg = sum(r["booked_gross"]   for r in data)
    occ    = round((agg_b/agg_t)*100,2) if agg_t else 0
    ws_sum.append(["Metric","Value"])
    for row in [("Total Cities",len(city_map)),("Total Theatres",len(th_map)),
                ("Total Shows",len(data)),("Booked Gross",agg_bg),
                ("Occupancy %",occ),("Generated At",datetime.now().strftime("%Y-%m-%d %H:%M:%S"))]:
        ws_sum.append(list(row))

    path = os.path.join(reports_dir, filename)
    wb.save(path)


def get_report_base_name(movie_name, show_date, report_type):
    """Generates a consistent base filename for reports."""
    slug = movie_name.replace(" ", "_")
    date_str = datetime.strptime(show_date, "%Y-%m-%d").strftime("%d%b")
    return f"{slug}_{date_str}_{report_type}Report"

def load_previous_report_data(base_name, reports_dir="reports"):
    """Loads past show data to merge newly extracted shows."""
    path = os.path.join(reports_dir, f"{base_name}_data.json")
    if os.path.exists(path):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            return data
        except Exception:
            pass
    return None

def save_report_data(final_data, base_name, reports_dir="reports"):
    """Saves the final merged show data into a JSON file."""
    path = os.path.join(reports_dir, f"{base_name}_data.json")
    serializable = []
    for show in final_data:
        s = {}
        for k, v in show.items():
            s[k] = list(v) if isinstance(v, set) else v
        serializable.append(s)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(serializable, f, ensure_ascii=False, indent=2)

def merge_with_previous_data(new_data, old_data):
    """Combines current scraping session data with previously cached show collections."""
    if not old_data:
        return new_data

    new_sids = set()
    for show in new_data:
        if show.get('bms_sid'): new_sids.add(show['bms_sid'])
        if show.get('district_sid'): new_sids.add(show['district_sid'])

    merged = list(new_data)
    preserved = 0
    for show in old_data:
        old_show_sids = {s for s in (show.get('bms_sid'), show.get('district_sid')) if s}
        if not old_show_sids and show.get('sid'):
            old_show_sids = {show['sid']}
        if old_show_sids & new_sids:
            continue
        merged.append(show)
        preserved += 1

    if preserved:
        print(f"   📌 Preserved {preserved} shows from previous run")

    return merged

def archive_previous_reports(base_name, reports_dir="reports", old_reports_dir="old_reports"):
    """Moves older report files into an old_reports directory to avoid overwriting them."""
    os.makedirs(old_reports_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    for ext in ['.xlsx', '.png', '.html', '_data.json']:
        src = os.path.join(reports_dir, f"{base_name}{ext}")
        if os.path.exists(src):
            dest = os.path.join(old_reports_dir, f"{base_name}_{ts}{ext}")
            shutil.move(src, dest)


# =============================================================================
# ── 8. MAIN EXECUTION ────────────────────────────────────────────────────────
# =============================================================================

if __name__ == "__main__":
    # Build city pairs — city name derived from District slug
    district_pairs = [
        (slug.replace("-", " ").title(), slug)
        for slug in DISTRICT_CITIES
    ]
    bms_pairs = [
        (d_slug.replace("-", " ").title(), b_slug)
        for d_slug, b_slug in zip(DISTRICT_CITIES, BMS_CITIES)
    ]

    with _global_bms_sids_lock: _global_bms_sids.clear()
    with _global_district_sids_lock: _global_district_sids.clear()

    total = len(district_pairs)
    print(f"🎬 Initializing run: {len(bms_pairs)} BMS Cities, {len(district_pairs)} District Cities")

    start_time = time.monotonic()

    with ThreadPoolExecutor(max_workers=2) as platform_pool:
        bms_future  = platform_pool.submit(run_bms, bms_pairs)
        dist_future = platform_pool.submit(run_district, district_pairs)
        all_bms_data  = bms_future.result()
        all_dist_data = dist_future.result()

    elapsed = time.monotonic() - start_time
    print(f"\n📋 Platforms finished in {elapsed/60:.1f} minutes.")
    print(f"   BMS: {len(all_bms_data)} shows | District: {len(all_dist_data)} shows")

    # Stamp city_order on records — BMS and District cities are provided in
    # the same order, so index-based matching avoids city-name spelling issues.
    dist_city_order = {name: i for i, (name, _) in enumerate(district_pairs)}
    bms_city_order  = {name: i for i, (name, _) in enumerate(bms_pairs)}
    for r in all_dist_data:
        r['city_order'] = dist_city_order.get(r['city'], -1)
    for r in all_bms_data:
        r['city_order'] = bms_city_order.get(r['city'], -1)

    load_venue_mapping()
    final_data = merge_data(all_dist_data, all_bms_data)

    if final_data:
        movie_name = extract_movie_name_from_url(DISTRICT_URL_TEMPLATE.replace("{city}", DISTRICT_CITIES[0]))
        show_date_fmt = datetime.strptime(SHOW_DATE, "%Y-%m-%d").strftime("%d %b %Y")
        base_name = get_report_base_name(movie_name, SHOW_DATE, "Cities")
        is_show_day = SHOW_DATE == datetime.now().strftime("%Y-%m-%d")
        current_run_data = list(final_data)

        old_data = load_previous_report_data(base_name)
        if old_data:
            final_data = merge_with_previous_data(final_data, old_data)

        archive_previous_reports(base_name)

        # Generate aggregated reports
        generate_excel(final_data, f"{base_name}.xlsx")
        generate_premium_city_image_report(
            final_data, f"reports/{base_name}.png",
            movie_name=movie_name, show_date=show_date_fmt,
        )
        generate_hybrid_city_html_report(
            final_data,
            DISTRICT_URL_TEMPLATE.replace("{city}", DISTRICT_CITIES[0]),
            f"reports/{base_name}.html",
        )
        save_report_data(final_data, base_name)

        # Generate snapshot reports
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        snapshot_name = f"{base_name}_{ts}"
        os.makedirs("old_reports", exist_ok=True)
        generate_excel(current_run_data, f"old_reports/{snapshot_name}.xlsx")
        generate_premium_city_image_report(
            current_run_data, f"old_reports/{snapshot_name}.png",
            movie_name=movie_name, show_date=show_date_fmt,
        )
        generate_hybrid_city_html_report(
            current_run_data,
            DISTRICT_URL_TEMPLATE.replace("{city}", DISTRICT_CITIES[0]),
            f"old_reports/{snapshot_name}.html",
        )

        aggregated_files = [
            f"reports/{base_name}.xlsx",
            f"reports/{base_name}.png",
            f"reports/{base_name}.html",
        ]
        snapshot_files = [
            f"old_reports/{snapshot_name}.xlsx",
            f"old_reports/{snapshot_name}.png",
            f"old_reports/{snapshot_name}.html",
        ]

        if is_show_day and old_data:
            send_collection_report(
                report_type="cities",
                movie_name=movie_name,
                show_date=show_date_fmt,
                subject_label="Tracked Gross + Advance Sales",
                attachment_paths=aggregated_files + snapshot_files,
                sections=[
                    {
                        "label": "A. Tracked Gross + Advance Sales (Full Day)",
                        "note":  "Cumulative gross tracked so far, including advance sales for remaining shows today.",
                        "files": aggregated_files,
                    },
                    {
                        "label": "B. Advance Sales (Remaining Shows)",
                        "note":  "Current advance booking status for shows yet to begin.",
                        "files": snapshot_files,
                    },
                ],
            )
        else:
            send_collection_report(
                report_type="cities",
                movie_name=movie_name,
                show_date=show_date_fmt,
                subject_label="Advance Sales",
                attachment_paths=aggregated_files,
            )

        total_elapsed = time.monotonic() - start_time
        print(f"\n🏁 Complete in {total_elapsed/60:.1f} minutes. Output: {base_name}")
    else:
        print("❌ No data found.")