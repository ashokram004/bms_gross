import json
import time
import os
import random
from base64 import b64decode
from concurrent.futures import ThreadPoolExecutor, as_completed
from itertools import cycle
from Crypto.Cipher import AES
from Crypto.Util.Padding import unpad
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from fake_useragent import UserAgent
from datetime import datetime

# --- IMPORT IMAGE GENERATORS ---
from utils.generateDistrictMultiStateImageReport import generate_multi_state_image_report
from utils.generateHybridStatesImageReport import generate_hybrid_image_report

# =========================== CONFIGURATION ===========================
INPUT_STATE_LIST = ["Andhra Pradesh", "Telangana"] 
SHOW_DATE = "2026-01-26"

# Config Paths
DISTRICT_CONFIG_PATH = os.path.join("utils", "district_cities_config.json")
BMS_CONFIG_PATH = os.path.join("utils", "bms_cities_config.json")

# Mapping Paths
DISTRICT_MAP_PATH = os.path.join("utils", "district_area_city_mapping.json")
BMS_MAP_PATH = os.path.join("utils", "bms_area_city_mapping.json")

# URLs
DISTRICT_URL_BASE = "https://www.district.in/movies/mana-shankara-varaprasad-garu-movie-tickets-in-"
BMS_URL_TEMPLATE = "https://in.bookmyshow.com/movies/{city}/mana-shankara-vara-prasad-garu/buytickets/ET00457184/20260126"

# BMS Settings
ENCRYPTION_KEY = "kYp3s6v9y$B&E)H+MbQeThWmZq4t7w!z"
BOOKED_STATES = {"2"}
SLEEP_TIME = 1.0        # Sleep between shows inside a worker
MAX_WORKERS = 3         # Number of parallel browsers for BMS

# SPEED OPTIMIZATION FLAG
SKIP_DUPLICATES_IN_BMS = True 

# PROXY LIST
PROXY_LIST = [] 
proxy_pool = cycle(PROXY_LIST) if PROXY_LIST else None

# =========================== MAPPING SYSTEM ===========================

def load_mapping_dict(file_path):
    """Returns {(state, city_name): reporting_city} or empty dict if fails."""
    mapping = {}
    if os.path.exists(file_path):
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                for state, cities in data.items():
                    if isinstance(cities, list):
                        for entry in cities:
                            if "name" in entry and "reporting_city" in entry:
                                mapping[(state, entry["name"])] = entry["reporting_city"]
        except Exception as e:
            print(f"⚠️ Warning: Could not parse mapping {file_path}: {e}")
    return mapping

# Pre-load mappings globally
DISTRICT_CITY_MAP = load_mapping_dict(DISTRICT_MAP_PATH)
BMS_CITY_MAP = load_mapping_dict(BMS_MAP_PATH)

def get_normalized_city_name(state, raw_city, source):
    """
    If the state exists in mapping and city is found, returns reporting_city.
    Otherwise, returns raw_city (ensures cross-state safety).
    """
    lookup = DISTRICT_CITY_MAP if source == "district" else BMS_CITY_MAP
    return lookup.get((state, raw_city), raw_city)

# =========================== CORE FUNCTIONS ===========================

def get_driver(proxy=None):
    ua = UserAgent()
    options = Options()
    options.add_argument(f"user-agent={ua.random}")
    options.add_argument("--headless=new") 
    options.add_argument("start-maximized")
    options.add_argument("--disable-web-security")
    options.add_argument("--disable-site-isolation-trials")
    options.add_argument("disable-csp")
    options.add_argument("--disable-blink-features=AutomationControlled")
    
    prefs = {
        "profile.managed_default_content_settings.images": 2,
        "profile.default_content_setting_values.notifications": 2
    }
    options.add_experimental_option("prefs", prefs)

    if proxy:
        options.add_argument(f'--proxy-server={proxy}')

    return webdriver.Chrome(options=options)

# ================= DISTRICT LOGIC =================
def fetch_district_data(driver):
    print("\n🚀 STARTING DISTRICT APP PROCESS...")
    if not os.path.exists(DISTRICT_CONFIG_PATH):
        print(f"❌ District Config missing: {DISTRICT_CONFIG_PATH}")
        return []

    with open(DISTRICT_CONFIG_PATH, 'r', encoding='utf-8') as f:
        config = json.load(f)
    
    results = []
    processed_sids = set() 
    
    for state in INPUT_STATE_LIST:
        cities = config.get(state, [])
        if not cities: continue

        for city in cities:
            url = f"{DISTRICT_URL_BASE}{city['slug']}-MV203929?fromdate={SHOW_DATE}"
            print(f"🌐 [{state}] Fetching {city['name']}...", end="\r")
            
            try:
                driver.get(url)
                time.sleep(1)
                html = driver.page_source
                
                marker = 'id="__NEXT_DATA__"'
                idx = html.find(marker)
                if idx == -1: continue
                
                start = html.find('>', idx) + 1
                end = html.find('</script>', start)
                data = json.loads(html[start:end])
                
                sessions = data['props']['pageProps']['data']['serverState']['movieSessions']
                key = list(sessions.keys())[0]
                cinemas = sessions[key]['pageData']['nearbyCinemas']

                # --- NORMALIZE CITY ---
                reporting_city = get_normalized_city_name(state, city['name'], "district")

                city_res = []
                for cin in cinemas:
                    venue = cin['cinemaInfo']['name']
                    for s in cin.get('sessions', []):
                        sid = str(s.get('sid', ''))

                        if sid in processed_sids: continue
                        processed_sids.add(sid)
                        
                        b_gross, p_gross, b_tkts, t_tkts = 0, 0, 0, 0
                        for a in s.get('areas', []):
                            tot, av, pr = a['sTotal'], a['sAvail'], a['price']
                            bk = tot - av
                            b_tkts += bk; t_tkts += tot
                            b_gross += (bk*pr); p_gross += (tot*pr)
                        
                        occ = round((b_tkts/t_tkts)*100, 2) if t_tkts else 0
                        
                        city_res.append({
                            "source": "district", "sid": sid,
                            "state": state, "city": reporting_city, "venue": venue,
                            "showTime": s['showTime'], "total_tickets": t_tkts,
                            "booked_tickets": b_tkts, "total_gross": p_gross,
                            "booked_gross": b_gross, "occupancy": occ
                        })
                
                if city_res:
                    gross = sum(x['booked_gross'] for x in city_res)
                    print(f"✅ {city['name']:<15} -> {reporting_city:<15} | Shows: {len(city_res):<3} | Gross: ₹{gross:<10,}")
                    results.extend(city_res)
            except Exception: pass
            
    return results

# ================= BMS LOGIC =================

def extract_initial_state_from_page(driver, url):
    try:
        driver.get(url)
        time.sleep(2)
        html = driver.page_source
        marker = "window.__INITIAL_STATE__"
        start = html.find(marker)
        if start == -1: return None 
        start = html.find("{", start)
        brace_count, end = 0, start
        while end < len(html):
            if html[end] == "{": brace_count += 1
            elif html[end] == "}": brace_count -= 1
            if brace_count == 0: break
            end += 1
        return json.loads(html[start:end + 1])
    except: return None

def extract_venues(state):
    if not state: return []
    try:
        sbe = state.get("showtimesByEvent")
        if not sbe: return []
        date_code = sbe.get("currentDateCode")
        if not date_code: return []
        widgets = sbe["showDates"][date_code]["dynamic"]["data"]["showtimeWidgets"]
        for w in widgets:
            if w.get("type") == "groupList":
                for g in w["data"]:
                    if g.get("type") == "venueGroup": return g["data"]
    except: pass
    return []

def get_seat_layout(driver, venue_code, session_id):
    api_url = "https://services-in.bookmyshow.com/doTrans.aspx"
    max_retries = 2
    js = """
        var cb = arguments[0]; var x = new XMLHttpRequest();
        x.open("POST", "%s", true);
        x.setRequestHeader("Content-Type", "application/x-www-form-urlencoded");
        x.onload = function() { cb(x.responseText); };
        x.send("strCommand=GETSEATLAYOUT&strAppCode=WEB&strVenueCode=%s&lngTransactionIdentifier=0&strParam1=%s&strParam2=WEB&strParam5=Y&strFormat=json");
    """ % (api_url, venue_code, session_id)

    for i in range(max_retries + 1):
        try:
            resp = driver.execute_async_script(js)
            data = json.loads(resp).get("BookMyShow", {})
            if data.get("blnSuccess") == "true": return data.get("strData")
            if "Rate limit" in data.get("strException", "") and i < max_retries:
                time.sleep(60)
                continue
            return None
        except: return None
    return None

def decrypt_data(enc):
    decoded = b64decode(enc)
    cipher = AES.new(ENCRYPTION_KEY.encode(), AES.MODE_CBC, iv=bytes(16))
    return unpad(cipher.decrypt(decoded), AES.block_size).decode()

def calculate_show_collection(decrypted, price_map):
    header, rows_part = decrypted.split("||")
    rows = rows_part.split("|")
    cat_map = {}
    for p in header.split("|"):
        parts = p.split(":")
        if len(parts) >= 3: cat_map[parts[1]] = parts[2]

    seats, booked = {}, {}
    for row in rows:
        if not row: continue
        parts = row.split(":")
        if len(parts) < 3: continue
        block = parts[3][0] if len(parts) > 3 else parts[2][0]
        area = cat_map.get(block)
        if not area: continue

        for seat in parts:
            if len(seat) < 2: continue
            status = seat[1]
            if seat[0] == block and status in ("1", "2"):
                seats[area] = seats.get(area, 0) + 1
            if status in BOOKED_STATES:
                booked[area] = booked.get(area, 0) + 1

    t_tkts, b_tkts, t_gross, b_gross = 0, 0, 0, 0
    for area, total in seats.items():
        bk = booked.get(area, 0)
        pr = price_map.get(area, 0)
        t_tkts += total; b_tkts += bk
        t_gross += total * pr; b_gross += bk * pr

    occ = round((b_tkts / t_tkts) * 100, 2) if t_tkts else 0
    return t_tkts, b_tkts, int(t_gross), int(b_gross), occ

def process_single_city(task_data):
    """ Worker Function for BMS Parallel Execution """
    city_name, city_slug, state_name, district_sids = task_data
    current_proxy = next(proxy_pool) if proxy_pool else None
    
    # --- NORMALIZE CITY ---
    reporting_city = get_normalized_city_name(state_name, city_name, "bms")

    print(f"🚀 Starting BMS: {city_name} -> {reporting_city}...")
    url = BMS_URL_TEMPLATE.format(city=city_slug)
    driver = get_driver(current_proxy)
    city_results = []
    city_total = 0

    try:
        state_data = extract_initial_state_from_page(driver, url)
        if not state_data: return [], 0, None
        venues = extract_venues(state_data)
        
        for venue in venues:
            v_name = venue["additionalData"]["venueName"]
            v_code = venue["additionalData"]["venueCode"]

            for show in venue.get("showtimes", []):
                sid = str(show["additionalData"]["sessionId"])
                show_time = show["title"]
                
                if SKIP_DUPLICATES_IN_BMS and sid in district_sids:
                    continue 

                soldOut = False
                try:
                    cats = show["additionalData"].get("categories", [])
                    price_map = {c["areaCatCode"]: float(c["curPrice"]) for c in cats}
                    enc = get_seat_layout(driver, v_code, sid)
                    
                    data = None
                    if not enc:
                        if not price_map: continue
                        fallback_p = max(price_map.values())
                        data = {
                            "total_tickets": 200, "booked_tickets": 200, "occupancy": 100.0,
                            "total_gross": int(200*fallback_p), "booked_gross": int(200*fallback_p)
                        }
                        soldOut = True
                    else:
                        decrypted = decrypt_data(enc)
                        res = calculate_show_collection(decrypted, price_map)
                        data = {
                            "total_tickets": abs(res[0]), "booked_tickets": min(abs(res[1]), abs(res[0])),
                            "total_gross": abs(res[2]), "booked_gross": min(abs(res[3]), abs(res[2])), "occupancy": min(100, abs(res[4]))
                        }

                    if data and data['total_tickets'] > 0:
                        data.update({
                            "source": "bms", "sid": sid,
                            "state": state_name, "city": reporting_city,
                            "venue": v_name, "showTime": show_time
                        })
                        city_results.append(data)
                        city_total += data["booked_gross"]

                except Exception: continue
                time.sleep(SLEEP_TIME) 
    except Exception: pass
    finally: driver.quit()
    return city_results, city_total, url

# ================= PART 3: EXCEL GENERATOR =================
def generate_consolidated_excel(all_results, filename):
    print("\n📊 Generating Consolidated Excel Report...")
    wb = Workbook()
    reports_dir = "reports"
    os.makedirs(reports_dir, exist_ok=True)

    # 1. STATE WISE
    ws_state = wb.active
    ws_state.title = "State Wise"
    ws_state.append(["State", "Cities", "Theatres", "Shows", "Total Seats", "Booked Seats", "Total Gross ₹", "Booked Gross ₹", "Occ %"])
    state_map, city_tracker, theatre_tracker = {}, {}, {}

    for r in all_results:
        st = r["state"]
        if st not in state_map:
            state_map[st] = {"shows":0, "t_seats":0, "b_seats":0, "p_gross":0, "b_gross":0}
            city_tracker[st] = set(); theatre_tracker[st] = set()
        
        d = state_map[st]
        d["shows"] += 1; d["t_seats"] += r["total_tickets"]; d["b_seats"] += r["booked_tickets"]
        d["p_gross"] += r["total_gross"]; d["b_gross"] += r["booked_gross"]
        city_tracker[st].add(r["city"]); theatre_tracker[st].add(r["venue"])

    for st, d in state_map.items():
        avg_occ = round((d["b_seats"] / d["t_seats"]) * 100, 2) if d["t_seats"] > 0 else 0
        ws_state.append([st, len(city_tracker[st]), len(theatre_tracker[st]), d["shows"], d["t_seats"], d["b_seats"], d["p_gross"], d["b_gross"], avg_occ])

    # 2. CITY WISE (Aggregates Reporting Cities)
    ws_city = wb.create_sheet(title="City Wise")
    ws_city.append(["State", "City", "Theatres", "Shows", "Total Seats", "Booked Seats", "Total Gross ₹", "Booked Gross ₹", "Occ %"])
    city_map, city_theatre_tracker = {}, {}

    for r in all_results:
        k = (r["state"], r["city"])
        if k not in city_map:
            city_map[k] = {"shows":0, "t_seats":0, "b_seats":0, "p_gross":0, "b_gross":0}
            city_theatre_tracker[k] = set()
        d = city_map[k]
        d["shows"] += 1; d["t_seats"] += r["total_tickets"]; d["b_seats"] += r["booked_tickets"]
        d["p_gross"] += r["total_gross"]; d["b_gross"] += r["booked_gross"]
        city_theatre_tracker[k].add(r["venue"])

    for (st, ct), d in city_map.items():
        avg_occ = round((d["b_seats"] / d["t_seats"]) * 100, 2) if d["t_seats"] > 0 else 0
        ws_city.append([st, ct, len(city_theatre_tracker[(st, ct)]), d["shows"], d["t_seats"], d["b_seats"], d["p_gross"], d["b_gross"], avg_occ])

    # 3. THEATRE WISE
    ws_th = wb.create_sheet(title="Theatre Wise")
    ws_th.append(["Source", "State", "City", "Venue", "Shows", "Total Seats", "Booked Seats", "Total Gross ₹", "Booked Gross ₹", "Occ %"])
    th_map = {}
    for r in all_results:
        k = (r["source"], r["state"], r["city"], r["venue"])
        if k not in th_map: th_map[k] = {"shows":0, "t_seats":0, "b_seats":0, "p_gross":0, "b_gross":0}
        d = th_map[k]
        d["shows"] += 1; d["t_seats"] += r["total_tickets"]; d["b_seats"] += r["booked_tickets"]
        d["p_gross"] += r["total_gross"]; d["b_gross"] += r["booked_gross"]

    for (src, st, ct, vn), d in th_map.items():
        avg_occ = round((d["b_seats"] / d["t_seats"]) * 100, 2) if d["t_seats"] > 0 else 0
        ws_th.append([src, st, ct, vn, d["shows"], d["t_seats"], d["b_seats"], d["p_gross"], d["b_gross"], avg_occ])

    # 4. SHOW WISE
    ws_show = wb.create_sheet(title="Show Wise")
    ws_show.append(["Source", "State", "City", "Venue", "Time", "SID", "Total Seats", "Booked Seats", "Total Gross ₹", "Booked Gross ₹", "Occ %"])
    for r in all_results:
        ws_show.append([r["source"], r["state"], r["city"], r["venue"], r["showTime"], r["sid"], r["total_tickets"], r["booked_tickets"], r["total_gross"], r["booked_gross"], r["occupancy"]])

    # 5. SUMMARY
    ws_sum = wb.create_sheet(title="Summary")
    agg_p_gross = sum(r["total_gross"] for r in all_results)
    agg_b_gross = sum(r["booked_gross"] for r in all_results)
    agg_t_seats = sum(r["total_tickets"] for r in all_results)
    agg_b_seats = sum(r["booked_tickets"] for r in all_results)
    overall_occ = round((agg_b_seats / agg_t_seats) * 100, 2) if agg_t_seats > 0 else 0
    
    ws_sum.append(["Metric", "Value"])
    ws_sum.append(["States Processed", len(state_map)])
    ws_sum.append(["Total Cities", len(city_map)])
    ws_sum.append(["Total Shows", len(all_results)])
    ws_sum.append(["Total Booked Gross (₹)", agg_b_gross])
    ws_sum.append(["Overall Occupancy %", overall_occ])
    ws_sum.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    path = os.path.join(reports_dir, filename)
    wb.save(path)
    print(f"📊 Consolidated Excel Saved: {path}")

# ================= MAIN EXECUTION FLOW =================
if __name__ == "__main__":
    # 1. DISTRICT
    d_driver = get_driver()
    district_data = []
    district_known_sids = set()

    try:
        district_data = fetch_district_data(d_driver)
        for r in district_data:
            if r["sid"]: district_known_sids.add(r["sid"])
        
        if district_data:
            ts = datetime.now().strftime("%H%M")
            generate_consolidated_excel(district_data, f"District_Only_{ts}.xlsx")
    finally:
        d_driver.quit() 

    # 2. BMS
    if os.path.exists(BMS_CONFIG_PATH):
        with open(BMS_CONFIG_PATH, 'r', encoding='utf-8') as f:
            bms_config = json.load(f)

        bms_tasks = []
        for state in INPUT_STATE_LIST:
            for city in bms_config.get(state, []):
                bms_tasks.append((city['name'], city['slug'], state, district_known_sids))

        print(f"🔥 Launching {MAX_WORKERS} Workers for {len(bms_tasks)} BMS cities...")
        bms_data = []
        last_valid_url = ""
        executor = ThreadPoolExecutor(max_workers=MAX_WORKERS)
        futures = [executor.submit(process_single_city, task) for task in bms_tasks]

        try:
            for future in as_completed(futures):
                res, total, url = future.result()
                if res:
                    bms_data.extend(res)
                    if url: last_valid_url = url
        except KeyboardInterrupt:
            executor.shutdown(wait=False)

    # 3. MERGE LOGIC
    print("\n🔄 Merging Data Sources...")
    merged_map = {}
    for r in district_data:
        key = r["sid"] if r["sid"] else f"{r['city']}_{r['venue']}_{r['showTime']}"
        merged_map[key] = r

    for r in bms_data:
        key = r["sid"] if r["sid"] else f"{r['city']}_{r['venue']}_{r['showTime']}"
        if key in merged_map:
            if not SKIP_DUPLICATES_IN_BMS and r["booked_gross"] > merged_map[key]["booked_gross"]:
                merged_map[key] = r 
        else:
            merged_map[key] = r

    final_data = list(merged_map.values())
    if final_data:
        ts_final = datetime.now().strftime("%Y%m%d_%H%M%S")
        generate_consolidated_excel(final_data, f"Total_States_Report_{ts_final}.xlsx")
        
        ref_url_final = last_valid_url if last_valid_url else (DISTRICT_URL_BASE + "city")
        generate_hybrid_image_report(final_data, BMS_URL_TEMPLATE, f"reports/Total_States_Report_{ts_final}.png", "bms")
    else:
        print("❌ No data found.")