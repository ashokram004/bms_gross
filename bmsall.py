import json
import time
import os
from base64 import b64decode
from Crypto.Cipher import AES
from Crypto.Util.Padding import unpad
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from fake_useragent import UserAgent
from datetime import datetime

from utils.generateBMSMultiCityImageReport import generate_multi_city_image_report

# --- CONFIGURATION --- Check bms_cities_config file under utils folder to find states 
INPUT_STATE_LIST = ["Andhra Pradesh", "Telangana"]  # Add states here
BMS_CONFIG_PATH = os.path.join("utils", "bms_cities_config.json")

# URL Template
MOVIE_URL_TEMPLATE = "https://in.bookmyshow.com/movies/{city}/mana-shankara-vara-prasad-garu/buytickets/ET00457184/20260124"

ENCRYPTION_KEY = "kYp3s6v9y$B&E)H+MbQeThWmZq4t7w!z"
BOOKED_STATES = {"2"}
SLEEP_TIME = 1

# ---------------- DRIVER ----------------
def get_driver():
    ua = UserAgent()
    options = Options()
    options.add_argument(f"user-agent={ua.random}")
    options.add_argument("--headless=new")
    options.add_argument("start-maximized")
    options.add_argument("--disable-web-security")
    options.add_argument("--disable-site-isolation-trials")
    options.add_argument("disable-csp")
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)

# ---------------- INITIAL STATE ----------------
def extract_initial_state_from_page(driver, url: str):
    driver.get(url)
    time.sleep(2)
    html = driver.page_source

    marker = "window.__INITIAL_STATE__"
    start = html.find(marker)
    if start == -1:
        if "BookMyShow" in driver.title:
            return None 
        raise ValueError("INITIAL_STATE not found")

    start = html.find("{", start)
    brace_count, end = 0, start

    while end < len(html):
        if html[end] == "{":
            brace_count += 1
        elif html[end] == "}":
            brace_count -= 1
        if brace_count == 0:
            break
        end += 1

    return json.loads(html[start:end + 1])

# ---------------- VENUES ----------------
def extract_venues(state):
    if not state: return []
    try:
        sbe = state.get("showtimesByEvent")
        if not sbe: return []
        
        date_code = sbe.get("currentDateCode")
        if not date_code: return []

        widgets = sbe["showDates"][date_code]["dynamic"]["data"]["showtimeWidgets"]

        for widget in widgets:
            if widget.get("type") == "groupList":
                for group in widget["data"]:
                    if group.get("type") == "venueGroup":
                        return group["data"]
    except Exception as e:
        print(f"⚠️ Error parsing venues: {e}")
    return []

# ---------------- SEAT LAYOUT ----------------
def get_seat_layout(driver, venue_code, session_id):
    global SLEEP_TIME
    api_url = "https://services-in.bookmyshow.com/doTrans.aspx"
    max_retries = 2

    for attempt in range(max_retries+1):
        js = """
        var callback = arguments[0];
        var xhr = new XMLHttpRequest();
        xhr.open("POST", "%s", true);
        xhr.setRequestHeader("Content-Type", "application/x-www-form-urlencoded");
        xhr.onload = function() { callback(xhr.responseText); };
        xhr.send(
            "strCommand=GETSEATLAYOUT" +
            "&strAppCode=WEB" +
            "&strVenueCode=%s" +
            "&lngTransactionIdentifier=0" +
            "&strParam1=%s" +
            "&strParam2=WEB" +
            "&strParam5=Y" +
            "&strFormat=json"
        );
        """ % (api_url, venue_code, session_id)
        
        try:
            response = driver.execute_async_script(js)
            data = json.loads(response)["BookMyShow"]

            if data.get("blnSuccess") == "true":
                return data.get("strData")
            else:
                exception = data.get("strException", "")
                if "Rate limit exceeded" in exception and attempt < max_retries:
                    print(f"🐢 Rate limit hit, retrying in 60s... (attempt {attempt + 1})")
                    time.sleep(60)
                    if SLEEP_TIME < 2.0: SLEEP_TIME = 2.0
                    continue
                else:
                    return None
        except Exception as e:
            return None
    return None

# ---------------- PRICE & DECRYPT ----------------
def extract_price_map_from_show(show):
    price_map = {}
    for cat in show["additionalData"].get("categories", []):
        price_map[cat["areaCatCode"]] = float(cat["curPrice"])
    return price_map

def decrypt_data(enc):
    decoded = b64decode(enc)
    cipher = AES.new(ENCRYPTION_KEY.encode(), AES.MODE_CBC, iv=bytes(16))
    return unpad(cipher.decrypt(decoded), AES.block_size).decode()

def extract_category_map(decrypted):
    header = decrypted.split("||")[0]
    category_map = {}
    for part in header.split("|"):
        pieces = part.split(":")
        if len(pieces) >= 3:
            category_map[pieces[1]] = pieces[2]
    return category_map

# ---------------- CALCULATION ----------------
def calculate_show_collection(decrypted, price_map):
    header, rows_part = decrypted.split("||")
    rows = rows_part.split("|")
    category_map = extract_category_map(decrypted)

    seats_map = {}
    booked_map = {}

    for row in rows:
        if not row: continue
        parts = row.split(":")
        if len(parts) < 3: continue
        elif len(parts) > 3: block_letter = parts[3][0]
        else: block_letter = parts[2][0]

        area_code = category_map.get(block_letter)
        if not area_code: continue

        for seat in parts:
            if len(seat) < 2: continue
            status = seat[1]
            
            if seat[0] == block_letter and status in ("1", "2"):
                seats_map[area_code] = seats_map.get(area_code, 0) + 1
            if status in BOOKED_STATES:
                booked_map[area_code] = booked_map.get(area_code, 0) + 1

    total_tickets = booked_tickets = total_gross = booked_gross = 0

    for area_code, total in seats_map.items():
        booked = booked_map.get(area_code, 0)
        price = price_map.get(area_code, 0)
        total_tickets += total
        booked_tickets += booked
        total_gross += total * price
        booked_gross += booked * price

    occupancy = round((booked_tickets / total_tickets) * 100, 2) if total_tickets else 0

    return {
        "total_tickets": total_tickets,
        "booked_tickets": booked_tickets,
        "occupancy": occupancy,
        "total_gross": int(total_gross),
        "booked_gross": int(booked_gross)
    }

# ---------------- PROCESS SINGLE MOVIE ----------------
def process_movie(driver, url, city_name, state_name):
    try:
        state = extract_initial_state_from_page(driver, url)
        if not state:
            print(f"   ⚠️ No initial state found for {city_name}")
            return [], 0
    except Exception as e:
        print(f"   ❌ Failed to load state for {city_name}: {e}")
        return [], 0

    venues = extract_venues(state)
    results = []
    grand_total = 0
    curShowNo = 0

    if not venues:
        print(f"   ⚠️ No venues found in {city_name}")

    for venue in venues:
        venue_code = venue["additionalData"]["venueCode"]
        venue_name = venue["additionalData"]["venueName"]

        for show in venue.get("showtimes", []):
            curShowNo += 1
            session_id = show["additionalData"]["sessionId"]
            show_time = show["title"]
            soldOut = False

            try:
                price_map = extract_price_map_from_show(show)
                enc = get_seat_layout(driver, venue_code, session_id)
                
                if not enc:
                    FALLBACK_TOTAL_SEATS = 200
                    if not price_map:
                        continue
                        
                    fallback_price = max(price_map.values())
                    data = {
                        "total_tickets": FALLBACK_TOTAL_SEATS,
                        "booked_tickets": FALLBACK_TOTAL_SEATS,
                        "occupancy": 100.0,
                        "total_gross": int(FALLBACK_TOTAL_SEATS * fallback_price),
                        "booked_gross": int(FALLBACK_TOTAL_SEATS * fallback_price)
                    }
                    soldOut = True
                else:
                    decrypted = decrypt_data(enc)
                    data = calculate_show_collection(decrypted, price_map)

            except Exception as e:
                continue

            if data['total_tickets'] > 0:
                tag = "(SOLD OUT HEURISTIC)" if soldOut else ""
                print(f"   🎬 {venue_name} | {show_time} | Occ: {data['occupancy']}% | Gross: ₹{data['booked_gross']} {tag}")

                data.update({
                    "state": state_name,
                    "city": city_name, 
                    "venue": venue_name, 
                    "showTime": show_time
                })
                results.append(data)
                grand_total += data["booked_gross"]
            
            time.sleep(SLEEP_TIME)

    return results, grand_total

# ---------------- EXCEL GENERATION ----------------
def generate_excel(all_results):
    print("\n📊 Generating Excel Report...")
    reports_dir = "reports"
    os.makedirs(reports_dir, exist_ok=True)
    wb = Workbook()

    # SHEET 1: CITY WISE
    city_sheet = wb.active
    city_sheet.title = "City Wise Collections"
    city_sheet.append(["State", "City", "Show Count", "Total Seats", "Booked Seats", "Occupancy %", "Total Gross", "Booked Gross"])

    city_totals = {}
    for r in all_results:
        key = (r["state"], r["city"])
        if key not in city_totals:
            city_totals[key] = {
                "show_count": 0, "total_seats": 0, "booked_seats": 0,
                "occupancies": [], "total_gross": 0, "booked_gross": 0
            }
        city_totals[key]["show_count"] += 1
        city_totals[key]["total_seats"] += r["total_tickets"]
        city_totals[key]["booked_seats"] += r["booked_tickets"]
        city_totals[key]["occupancies"].append(r["occupancy"])
        city_totals[key]["total_gross"] += r["total_gross"]
        city_totals[key]["booked_gross"] += r["booked_gross"]

    for (st, ct), data in city_totals.items():
        avg_occ = round(sum(data["occupancies"]) / data["show_count"], 2) if data["show_count"] else 0
        city_sheet.append([
            st, ct, data["show_count"], data["total_seats"], data["booked_seats"],
            avg_occ, data["total_gross"], data["booked_gross"]
        ])

    tc_shows = sum(d["show_count"] for d in city_totals.values())
    tc_seats = sum(d["total_seats"] for d in city_totals.values())
    tc_booked = sum(d["booked_seats"] for d in city_totals.values())
    tc_gross = sum(d["total_gross"] for d in city_totals.values())
    tc_bgross = sum(d["booked_gross"] for d in city_totals.values())
    tc_occ = round((tc_booked / tc_seats) * 100, 2) if tc_seats else 0
    city_sheet.append(["TOTAL", "", tc_shows, tc_seats, tc_booked, tc_occ, tc_gross, tc_bgross])

    # SHEET 2: THEATRE WISE
    theatre_sheet = wb.create_sheet(title="Theatre Wise Collections")
    theatre_sheet.append(["State", "City", "Venue", "Show count", "Total Seats", "Booked Seats", "Occupancy %", "Total Gross", "Booked Gross"])

    theatre_data = {}
    for r in all_results:
        key = (r["state"], r["city"], r["venue"])
        if key not in theatre_data:
            theatre_data[key] = {
                "num_shows": 0, "total_tickets": 0, "booked_tickets": 0,
                "occupancies": [], "total_gross": 0, "booked_gross": 0
            }
        theatre_data[key]["num_shows"] += 1
        theatre_data[key]["total_tickets"] += r["total_tickets"]
        theatre_data[key]["booked_tickets"] += r["booked_tickets"]
        theatre_data[key]["occupancies"].append(r["occupancy"])
        theatre_data[key]["total_gross"] += r["total_gross"]
        theatre_data[key]["booked_gross"] += r["booked_gross"]

    for key, data in theatre_data.items():
        avg_occ = round(sum(data["occupancies"]) / data["num_shows"], 2) if data["num_shows"] else 0
        theatre_sheet.append([
            key[0], key[1], key[2], data["num_shows"], data["total_tickets"], data["booked_tickets"],
            avg_occ, data["total_gross"], data["booked_gross"]
        ])

    # SHEET 3: SHOW WISE
    show_sheet = wb.create_sheet(title="Show Wise Collections")
    show_sheet.append(["State", "City", "Venue", "Show Time", "Total Seats", "Booked Seats", "Occupancy %", "Total Gross", "Booked Gross"])
    for r in all_results:
        show_sheet.append([
            r["state"], r["city"], r["venue"], r["showTime"], r["total_tickets"], r["booked_tickets"],
            r["occupancy"], r["total_gross"], r["booked_gross"]
        ])

    # SHEET 4: SUMMARY
    summary = wb.create_sheet(title="Summary")
    summary.append(["Metric", "Value"])
    summary.append(["Total Cities", len(city_totals)])
    summary.append(["Total Theatres", len(theatre_data)])
    summary.append(["Total Shows", len(all_results)])
    summary.append(["Overall Occupancy (%)", tc_occ])
    summary.append(["Total Potential Gross (INR)", tc_gross])
    summary.append(["Total Booked Gross (INR)", tc_bgross])
    summary.append(["Report Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    file_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"bms_multistate_{file_ts}.xlsx"
    filepath = os.path.join(reports_dir, filename)
    wb.save(filepath)
    print(f"✅ Excel report saved at: {filepath}")

# ---------------- MAIN EXECUTION ----------------
if __name__ == "__main__":
    if not os.path.exists(BMS_CONFIG_PATH):
        print(f"❌ BMS config not found at {BMS_CONFIG_PATH}")
        exit()

    with open(BMS_CONFIG_PATH, 'r', encoding='utf-8') as f:
        bms_config = json.load(f)

    all_results = []
    last_valid_url = "" 

    print(f"🚀 Starting BMS Scraping for States: {', '.join(INPUT_STATE_LIST)}")

    for state in INPUT_STATE_LIST:
        cities = bms_config.get(state, [])
        if not cities:
            print(f"⚠️ No cities found for {state}")
            continue
            
        print(f"📍 Fetching {len(cities)} cities for {state}...")

        for city_obj in cities:
            city_name = city_obj['name']
            city_slug = city_obj['slug']
            
            print(f"\n🌍 Fetching: {city_name} ({state})...")
            
            # FRESH DRIVER PER CITY
            driver = get_driver()
            try:
                url = MOVIE_URL_TEMPLATE.format(city=city_slug)
                last_valid_url = url 
                
                results_city, total_city = process_movie(driver, url, city_name, state)
                
                all_results.extend(results_city)
                print(f"   💰 {city_name} Total: ₹{total_city}")

            except Exception as e:
                print(f"   ❌ Critical Error for {city_name}: {e}")
            finally:
                driver.quit() # Close driver to prevent blocking
            
            time.sleep(2) # Short pause between cities

    if all_results:
        # 1. Generate Excel
        generate_excel(all_results)
        
        # 2. Generate Multi-City Image Report
        img_path = f"reports/bms_multistate_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        generate_multi_city_image_report(all_results, last_valid_url, img_path)
    else:
        print("No data collected.")