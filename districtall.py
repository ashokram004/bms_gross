import json
import os
import time
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from utils.generateDistrictMultiStateImageReport import generate_multi_state_image_report

# Add states for reporting
inputStateList = ["Andhra Pradesh"] 


# --- CONFIGURATION ---
CONFIG_PATH = os.path.join("utils", "district_cities_config.json")
MOVIE_BASE_URL = "https://www.district.in/movies/mana-shankara-varaprasad-garu-movie-tickets-in-"
SHOW_DATE = "2026-01-24"

def get_driver():
    options = Options()
    options.add_argument("--headless=new")
    options.add_argument("start-maximized")
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=options)

def extract_city_data(driver, state_name, city_name, city_slug, processed_ids):
    url = f"{MOVIE_BASE_URL}{city_slug}-MV203929?fromdate={SHOW_DATE}"
    print(f"🌐 [{state_name}] Fetching {city_name}...", end="\r")
    
    try:
        driver.get(url)
        time.sleep(2) 
        html = driver.page_source

        marker = 'id="__NEXT_DATA__"'
        start_idx = html.find(marker)
        if start_idx == -1: return []
        
        start_json = html.find('>', start_idx) + 1
        end_json = html.find('</script>', start_json)
        full_data = json.loads(html[start_json:end_json])
        
        sessions_container = full_data['props']['pageProps']['data']['serverState']['movieSessions']
        dynamic_key = list(sessions_container.keys())[0]
        
        # Use nearbyCinemas for city-specific targeting
        nearby_cinemas = sessions_container[dynamic_key]['pageData']['nearbyCinemas']
        
        city_results = []
        for cinema in nearby_cinemas:
            venue = cinema['cinemaInfo']['name']
            for s in cinema.get('sessions', []):
                
                # ✅ DUPLICATE CHECK: Skip if session ID was already processed in a previous city
                session_id = s.get('sid', '')
                if session_id in processed_ids:
                    continue
                processed_ids.add(session_id)

                b_gross, p_gross, b_tkts, t_tkts = 0, 0, 0, 0
                for area in s.get('areas', []):
                    total, avail, price = area['sTotal'], area['sAvail'], area['price']
                    booked = total - avail
                    b_tkts += booked
                    t_tkts += total
                    b_gross += (booked * price)
                    p_gross += (total * price)
                    
                occ = round((b_tkts / t_tkts) * 100, 2) if t_tkts > 0 else 0
                city_results.append({
                    "state": state_name, "city": city_name, "venue": venue,
                    "showTime": s['showTime'], "total_tickets": t_tkts,
                    "booked_tickets": b_tkts, "total_gross": p_gross, 
                    "booked_gross": b_gross, "occupancy": occ
                })
        return city_results
    except Exception:
        return []
    

def generate_consolidated_report(all_results):
    wb = Workbook()
    reports_dir = "reports"
    os.makedirs(reports_dir, exist_ok=True)

    # 1. STATE WISE SHEET (New)
    ws_state = wb.active
    ws_state.title = "State Wise Collections"
    ws_state.append(["State", "Cities", "Theatres", "Shows", "Total Seats", "Booked Seats", "Total Gross ₹", "Booked Gross ₹", "Occ %"])
    
    state_map = {}
    city_tracker = {} # To count unique cities per state
    theatre_tracker = {} # To count unique theatres per state

    for r in all_results:
        st = r["state"]
        if st not in state_map:
            state_map[st] = {"shows":0, "t_seats":0, "b_seats":0, "p_gross":0, "b_gross":0}
            city_tracker[st] = set()
            theatre_tracker[st] = set()
        
        d = state_map[st]
        d["shows"] += 1
        d["t_seats"] += r["total_tickets"]
        d["b_seats"] += r["booked_tickets"]
        d["p_gross"] += r["total_gross"]
        d["b_gross"] += r["booked_gross"]
        city_tracker[st].add(r["city"])
        theatre_tracker[st].add(r["venue"])

    for st, d in state_map.items():
        avg_occ = round((d["b_seats"] / d["t_seats"]) * 100, 2) if d["t_seats"] > 0 else 0
        ws_state.append([st, len(city_tracker[st]), len(theatre_tracker[st]), d["shows"], d["t_seats"], d["b_seats"], d["p_gross"], d["b_gross"], avg_occ])

    # 2. CITY WISE SHEET (New)
    ws_city = wb.create_sheet(title="City Wise Collections")
    ws_city.append(["State", "City", "Theatres", "Shows", "Total Seats", "Booked Seats", "Total Gross ₹", "Booked Gross ₹", "Occ %"])
    
    city_map = {}
    city_theatre_tracker = {}

    for r in all_results:
        key = (r["state"], r["city"])
        if key not in city_map:
            city_map[key] = {"shows":0, "t_seats":0, "b_seats":0, "p_gross":0, "b_gross":0}
            city_theatre_tracker[key] = set()
        
        d = city_map[key]
        d["shows"] += 1
        d["t_seats"] += r["total_tickets"]
        d["b_seats"] += r["booked_tickets"]
        d["p_gross"] += r["total_gross"]
        d["b_gross"] += r["booked_gross"]
        city_theatre_tracker[key].add(r["venue"])

    for (st, ct), d in city_map.items():
        avg_occ = round((d["b_seats"] / d["t_seats"]) * 100, 2) if d["t_seats"] > 0 else 0
        ws_city.append([st, ct, len(city_theatre_tracker[(st, ct)]), d["shows"], d["t_seats"], d["b_seats"], d["p_gross"], d["b_gross"], avg_occ])

    # 3. THEATRE WISE SHEET
    ws1 = wb.create_sheet(title="Theatre Wise Collections")
    ws1.append(["State", "City", "Venue", "Shows", "Total Seats", "Booked Seats", "Total Gross ₹", "Booked Gross ₹", "Occ %"])
    
    theatre_map = {}
    for r in all_results:
        key = (r["state"], r["city"], r["venue"])
        if key not in theatre_map:
            theatre_map[key] = {"shows":0, "t_seats":0, "b_seats":0, "p_gross":0, "b_gross":0}
        d = theatre_map[key]
        d["shows"] += 1
        d["t_seats"] += r["total_tickets"]
        d["b_seats"] += r["booked_tickets"]
        d["p_gross"] += r["total_gross"]
        d["b_gross"] += r["booked_gross"]

    for (st, ct, vn), d in theatre_map.items():
        avg_occ = round((d["b_seats"] / d["t_seats"]) * 100, 2) if d["t_seats"] > 0 else 0
        ws1.append([st, ct, vn, d["shows"], d["t_seats"], d["b_seats"], d["p_gross"], d["b_gross"], avg_occ])

    # 4. SHOW WISE SHEET
    ws2 = wb.create_sheet(title="Show Wise Collections")
    ws2.append(["State", "City", "Venue", "Time", "Total Seats", "Booked Seats", "Total Gross ₹", "Booked Gross ₹", "Occ %"])
    for r in all_results:
        ws2.append([r["state"], r["city"], r["venue"], r["showTime"], r["total_tickets"], r["booked_tickets"], r["total_gross"], r["booked_gross"], r["occupancy"]])

    # 5. SUMMARY SHEET
    ws3 = wb.create_sheet(title="Summary")
    agg_p_gross = sum(r["total_gross"] for r in all_results)
    agg_b_gross = sum(r["booked_gross"] for r in all_results)
    agg_t_seats = sum(r["total_tickets"] for r in all_results)
    agg_b_seats = sum(r["booked_tickets"] for r in all_results)
    overall_occ = round((agg_b_seats / agg_t_seats) * 100, 2) if agg_t_seats > 0 else 0
    
    ws3.append(["Metric", "Value"])
    ws3.append(["States Processed", len(state_map)])
    ws3.append(["Total Cities", len(city_map)])
    ws3.append(["Total Theatres", len(theatre_map)])
    ws3.append(["Grand Total Potential Gross (₹)", agg_p_gross])
    ws3.append(["Grand Total Booked Gross (₹)", agg_b_gross])
    ws3.append(["Overall Occupancy (%)", overall_occ])
    ws3.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    filename = f"district_consolidated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(reports_dir, filename)
    wb.save(filepath)
    print(f"\n📊 Excel Report with 5 Sheets Saved: {filepath}")


if __name__ == "__main__":
    if not os.path.exists(CONFIG_PATH):
        raise FileNotFoundError(f"Missing config file at {CONFIG_PATH}")
        
    with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
        cities_config = json.load(f)
    
    driver = get_driver()
    final_data = []
    
    # ✅ Initialize the master session tracker
    processed_session_ids = set()
    
    try:
        for state in inputStateList:
            cities_to_fetch = cities_config.get(state, [])
            for city in cities_to_fetch:
                # Pass the tracker into the function
                res = extract_city_data(driver, state, city['name'], city['slug'], processed_session_ids)
                
                if res:
                    city_b_gross = sum(r['booked_gross'] for r in res)
                    city_p_gross = sum(r['total_gross'] for r in res)
                    city_b_seats = sum(r['booked_tickets'] for r in res)
                    city_t_seats = sum(r['total_tickets'] for r in res)
                    city_occ = round((city_b_seats / city_t_seats) * 100, 2) if city_t_seats > 0 else 0
                    
                    print(f"✅ {city['name']:<15} | Shows: {len(res):<3} | Total Gross: ₹{city_p_gross:<10,} | Booked Gross: ₹{city_b_gross:<10,} | Occ: {city_occ:>6}%")
                    final_data.extend(res)
                else:
                    print(f"⚪ {city['name']:<15} | No new shows found.                                                                           ")

    finally:
        driver.quit()

    if final_data:
        generate_consolidated_report(final_data)
        img_path = f"reports/district_multistate_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        
        # Now passing URL instead of hardcoded strings
        generate_multi_state_image_report(final_data, MOVIE_BASE_URL + "city?fromdate=" + SHOW_DATE, img_path)