import json
import time
from base64 import b64decode
from Crypto.Cipher import AES
from Crypto.Util.Padding import unpad
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
from fake_useragent import UserAgent
from datetime import datetime
from utils.generateImageReport import generate_city_image_report
import os

url = "https://in.bookmyshow.com/movies/hyderabad/mana-shankara-vara-prasad-garu/buytickets/ET00457184/20260124"

ENCRYPTION_KEY = "kYp3s6v9y$B&E)H+MbQeThWmZq4t7w!z"

# ✅ SEAT RULE
# 0 = empty/aisle/invalid
# 1 = available
# 2 = booked
BOOKED_STATES = {"2"}
SLEEP_TIME = 1


# ---------------- DRIVER ----------------
def get_driver():
    ua = UserAgent()
    options = Options()
    options.add_argument(f"user-agent={ua.random}")
    options.add_argument("--headless")
    options.add_argument("start-maximized")
    options.add_argument("--disable-web-security")
    options.add_argument("--disable-site-isolation-trials")
    options.add_argument("disable-csp")
    return webdriver.Chrome(options=options)


driver = get_driver()


# ---------------- INITIAL STATE ----------------
def extract_initial_state_from_page(url: str):
    driver.get(url)
    html = driver.page_source

    marker = "window.__INITIAL_STATE__"
    start = html.find(marker)
    if start == -1:
        raise ValueError("INITIAL_STATE not found")

    start = html.find("{", start)
    brace_count, end = 0, start

    while end < len(html):
        if html[end] == "{":
            brace_count += 1
        elif html[end] == "}":
            brace_count -= 1
        if brace_count == 0:
            break
        end += 1

    return json.loads(html[start:end + 1])


# ---------------- VENUES ----------------
def extract_venues(state):
    sbe = state["showtimesByEvent"]
    date_code = sbe["currentDateCode"]
    widgets = sbe["showDates"][date_code]["dynamic"]["data"]["showtimeWidgets"]

    for widget in widgets:
        if widget.get("type") == "groupList":
            for group in widget["data"]:
                if group.get("type") == "venueGroup":
                    return group["data"]
    return []


# ---------------- SEAT LAYOUT ----------------
def get_seat_layout(venue_code, session_id):
    global SLEEP_TIME
    api_url = "https://services-in.bookmyshow.com/doTrans.aspx"
    max_retries = 2

    for attempt in range(max_retries+1):
        js = """
        var callback = arguments[0];
        var xhr = new XMLHttpRequest();
        xhr.open("POST", "%s", true);
        xhr.setRequestHeader("Content-Type", "application/x-www-form-urlencoded");
        xhr.onload = function() { callback(xhr.responseText); };
        xhr.send(
            "strCommand=GETSEATLAYOUT" +
            "&strAppCode=WEB" +
            "&strVenueCode=%s" +
            "&lngTransactionIdentifier=0" +
            "&strParam1=%s" +
            "&strParam2=WEB" +
            "&strParam5=Y" +
            "&strFormat=json"
        );
        """ % (api_url, venue_code, session_id)
        response = driver.execute_async_script(js)

        data = json.loads(response)["BookMyShow"]

        if data.get("blnSuccess") == "true":
            return data.get("strData")
        else:
            exception = data.get("strException", "")
            if "Rate limit exceeded" in exception and attempt < max_retries:
                print(f"Rate limit hit, retrying in 60 seconds... (attempt {attempt + 1}/{max_retries + 1})")
                time.sleep(60)
                # Increase global sleep after first rate limit
                if SLEEP_TIME < 2.0:
                    SLEEP_TIME = 2.0
                    print("🐢 Increasing global sleep time to 1.0s")
                continue
            else:
                print(exception)
                return None


# ---------------- PRICE MAP ----------------
def extract_price_map_from_show(show):
    price_map = {}
    for cat in show["additionalData"].get("categories", []):
        price_map[cat["areaCatCode"]] = float(cat["curPrice"])
    return price_map


# ---------------- DECRYPT ----------------
def decrypt_data(enc):
    decoded = b64decode(enc)
    cipher = AES.new(ENCRYPTION_KEY.encode(), AES.MODE_CBC, iv=bytes(16))
    return unpad(cipher.decrypt(decoded), AES.block_size).decode()


# ---------------- CATEGORY MAP ----------------
def extract_category_map(decrypted):
    """
    Maps:
    A -> 0000000002
    B -> 0000000004
    C -> 0000000005
    """
    header = decrypted.split("||")[0]
    category_map = {}

    for part in header.split("|"):
        pieces = part.split(":")
        if len(pieces) >= 3:
            letter = pieces[1]
            area_code = pieces[2]
            category_map[letter] = area_code

    return category_map


# ---------------- CALCULATION ----------------
def calculate_show_collection(decrypted, price_map):
    header, rows_part = decrypted.split("||")
    rows = rows_part.split("|")

    category_map = extract_category_map(decrypted)

    seats_map = {}
    booked_map = {}

    for row in rows:
        if not row:
            continue

        parts = row.split(":")

        # A000 / B000 / C000 → take first letter
        #Skip invalid row
        if len(parts) < 3:
            continue
        elif len(parts) > 3:
            block_letter = parts[3][0]
        else:
            block_letter = parts[2][0]

        area_code = category_map.get(block_letter)

        if not area_code:
            continue

        for seat in parts:
            #Skip invalid seats
            if len(seat) < 2:
                continue

            status = seat[1]

            # count total seats (ignore aisle/void only if explicitly 0+0. 1 means available, 2 means filled)
            if seat[0] == block_letter and status in ("1", "2"):
                seats_map[area_code] = seats_map.get(area_code, 0) + 1

            if status in BOOKED_STATES:
                booked_map[area_code] = booked_map.get(area_code, 0) + 1

    total_tickets = booked_tickets = total_gross = booked_gross = 0

    for area_code, total in seats_map.items():
        booked = booked_map.get(area_code, 0)
        price = price_map.get(area_code, 0)

        total_tickets += total
        booked_tickets += booked
        total_gross += total * price
        booked_gross += booked * price

    occupancy = round((booked_tickets / total_tickets) * 100, 2) if total_tickets else 0

    return {
        "total_tickets": total_tickets,
        "booked_tickets": booked_tickets,
        "occupancy": occupancy,
        "total_gross": int(total_gross),
        "booked_gross": int(booked_gross)
    }


# ---------------- MAIN PROCESS ----------------
def process_movie(url):
    state = extract_initial_state_from_page(url)
    venues = extract_venues(state)

    results = []
    grand_total = 0
    curShowNo = 0

    for venue in venues:
        venue_code = venue["additionalData"]["venueCode"]
        venue_name = venue["additionalData"]["venueName"]

        for show in venue.get("showtimes", []):
            curShowNo += 1
            session_id = show["additionalData"]["sessionId"]
            show_time = show["title"]

            try:
                price_map = extract_price_map_from_show(show)
                enc = get_seat_layout(venue_code, session_id)
                if not enc:
                    # 🔴 SOLD OUT FALLBACK (HEURISTIC)
                    FALLBACK_TOTAL_SEATS = 200

                    if not price_map:
                        raise ValueError("Price map missing for sold-out show")

                    fallback_price = max(price_map.values())

                    data = {
                        "total_tickets": FALLBACK_TOTAL_SEATS,
                        "booked_tickets": FALLBACK_TOTAL_SEATS,
                        "occupancy": 100.0,
                        "total_gross": int(FALLBACK_TOTAL_SEATS * fallback_price),
                        "booked_gross": int(FALLBACK_TOTAL_SEATS * fallback_price)
                    }
                else:
                    decrypted = decrypt_data(enc)
                    data = calculate_show_collection(decrypted, price_map)

            except Exception as e:
                print(f"❌ Skipping {venue_name} | {show_time} : {e}")
                continue

            if data['total_tickets'] == 0:
                print(f"Something is wrong with this show data: {venue_name} | {show_time}")

            print(
                f"Show no: {curShowNo}\n"
                f"🎬 {venue_name} | {show_time}\n"
                f"   Seats: {data['total_tickets']} | "
                f"Booked: {data['booked_tickets']} | "
                f"Occ: {data['occupancy']}% | "
                f"Gross: ₹{data['booked_gross']}"
            )

            data.update({"venue": venue_name, "showTime": show_time})
            results.append(data)
            grand_total += data["booked_gross"]

            time.sleep(SLEEP_TIME)

    print(f"\n💰 TOTAL COLLECTION: ₹{grand_total}")
    return results, grand_total


# ---------------- EXCEL ----------------

def generate_excel(results, total):
    # ---------------- CREATE REPORTS FOLDER ----------------
    reports_dir = "reports"
    os.makedirs(reports_dir, exist_ok=True)

    wb = Workbook()

    # ================= SHEET 1 : THEATRE WISE COLLECTIONS =================
    theatre_sheet = wb.active
    theatre_sheet.title = "Theatre Wise Collections"
    headers2 = [
        "Venue", "Show count", "Total Seats",
        "Booked Seats", "Occupancy %",
        "Total Gross", "Booked Gross"
    ]
    theatre_sheet.append(headers2)

    theatre_data = {}
    for r in results:
        venue = r["venue"]
        if venue not in theatre_data:
            theatre_data[venue] = {
                "num_shows": 0,
                "total_tickets": 0,
                "booked_tickets": 0,
                "occupancies": [],
                "total_gross": 0,
                "booked_gross": 0
            }
        theatre_data[venue]["num_shows"] += 1
        theatre_data[venue]["total_tickets"] += r["total_tickets"]
        theatre_data[venue]["booked_tickets"] += r["booked_tickets"]
        theatre_data[venue]["occupancies"].append(r["occupancy"])
        theatre_data[venue]["total_gross"] += r["total_gross"]
        theatre_data[venue]["booked_gross"] += r["booked_gross"]

    for venue, data in theatre_data.items():
        num_shows = data["num_shows"]
        avg_occ = round(sum(data["occupancies"]) / num_shows, 2) if num_shows else 0
        theatre_sheet.append([
            venue,
            num_shows,
            data["total_tickets"],
            data["booked_tickets"],
            avg_occ,
            data["total_gross"],
            data["booked_gross"]
        ])

    # --------- AGGREGATES ROW ---------
    total_shows_overall = sum(data["num_shows"] for data in theatre_data.values())
    total_seats_theatre = sum(data["total_tickets"] for data in theatre_data.values())
    total_booked_seats_theatre = sum(data["booked_tickets"] for data in theatre_data.values())
    total_gross_theatre = sum(data["total_gross"] for data in theatre_data.values())
    total_booked_gross_theatre = sum(data["booked_gross"] for data in theatre_data.values())
    overall_occupancy_theatre = round((total_booked_seats_theatre / total_seats_theatre) * 100, 2) if total_seats_theatre else 0

    theatre_sheet.append([
        "TOTAL / AVG",
        total_shows_overall,
        total_seats_theatre,
        total_booked_seats_theatre,
        overall_occupancy_theatre,
        total_gross_theatre,
        total_booked_gross_theatre
    ])

    # ================= SHEET 2 : SHOW WISE COLLECTIONS =================
    sheet = wb.create_sheet(title="Show Wise Collections")

    headers = [
        "Venue", "Show Time", "Total Seats",
        "Booked Seats", "Occupancy %",
        "Total Gross", "Booked Gross"
    ]
    sheet.append(headers)

    for r in results:
        sheet.append([
            r["venue"],
            r["showTime"],
            r["total_tickets"],
            r["booked_tickets"],
            r["occupancy"],
            r["total_gross"],
            r["booked_gross"]
        ])

    # --------- AGGREGATES ROW ---------
    total_seats = sum(r["total_tickets"] for r in results)
    total_booked_seats = sum(r["booked_tickets"] for r in results)
    total_gross = sum(r["total_gross"] for r in results)
    total_booked_gross = sum(r["booked_gross"] for r in results)
    avg_occupancy = (
        round(sum(r["occupancy"] for r in results) / len(results), 2)
        if results else 0
    )

    sheet.append([
        "TOTAL / AVG",
        "-",
        total_seats,
        total_booked_seats,
        avg_occupancy,
        total_gross,
        total_booked_gross
    ])

    # ================= SHEET 3 : SUMMARY =================

    summary = wb.create_sheet(title="Summary")

    total_theatres = len(set(r["venue"] for r in results))
    total_shows = len(results)

    overall_occupancy = (
        round((total_booked_seats / total_seats) * 100, 2)
        if total_seats else 0
    )

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    summary.append(["Metric", "Value"])
    summary.append(["Total Theatres", total_theatres])
    summary.append(["Total Shows", total_shows])
    summary.append(["Overall Occupancy (%)", overall_occupancy])
    summary.append(["Total Potential Gross (₹)", total_gross])
    summary.append(["Total Booked Gross (₹)", total_booked_gross])
    summary.append(["Report Generated At", timestamp])
    

    # ================= SAVE FILE =================
    file_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"bms_collections_{file_ts}.xlsx"
    filepath = os.path.join(reports_dir, filename)

    wb.save(filepath)
    print(f"📊 Excel report generated: {filepath}")


# ---------------- RUN ----------------

results, total = process_movie(url)
generate_excel(results, total)

img_path = f"reports/bms_city_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
generate_city_image_report(results, url, img_path, "bms")

driver.quit()
