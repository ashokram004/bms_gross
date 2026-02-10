import json
import time
import os
import random
from base64 import b64decode
from concurrent.futures import ThreadPoolExecutor, as_completed
from itertools import cycle
from Crypto.Cipher import AES
from Crypto.Util.Padding import unpad
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from fake_useragent import UserAgent
from datetime import datetime, timedelta
import difflib
from collections import defaultdict

# --- IMPORT IMAGE GENERATORS ---
from utils.generateDistrictMultiStateImageReport import generate_multi_state_image_report
from utils.generateHybridStatesImageReport import generate_hybrid_image_report

# =========================== CONFIGURATION ===========================
INPUT_STATE_LIST = ["Andhra Pradesh", "Telangana"] 
SHOW_DATE = "2026-01-27"

# Config Paths
DISTRICT_CONFIG_PATH = os.path.join("utils", "district_cities_config.json")
BMS_CONFIG_PATH = os.path.join("utils", "bms_cities_config.json")

# Mapping Paths
DISTRICT_MAP_PATH = os.path.join("utils", "district_area_city_mapping.json")
BMS_MAP_PATH = os.path.join("utils", "bms_area_city_mapping.json")

# URLs
DISTRICT_URL_BASE = "https://www.district.in/movies/mana-shankara-varaprasad-garu-movie-tickets-in-"
BMS_URL_TEMPLATE = "https://in.bookmyshow.com/movies/{city}/mana-shankara-vara-prasad-garu/buytickets/ET00457184/20260127"

# BMS Settings
ENCRYPTION_KEY = "kYp3s6v9y$B&E)H+MbQeThWmZq4t7w!z"
BOOKED_STATES = {"2"}
SLEEP_TIME = 1.0        # Sleep between shows inside a worker
MAX_WORKERS = 3         # Number of parallel browsers for BMS

processed_sids = set()

# PROXY LIST
PROXY_LIST = [] 
proxy_pool = cycle(PROXY_LIST) if PROXY_LIST else None

# =========================== MAPPING SYSTEM ===========================

def load_mapping_dict(file_path):
    """Returns {(state, city_name): reporting_city} or empty dict if fails."""
    mapping = {}
    if os.path.exists(file_path):
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                for state, cities in data.items():
                    if isinstance(cities, list):
                        for entry in cities:
                            if "name" in entry and "reporting_city" in entry:
                                mapping[(state, entry["name"])] = entry["reporting_city"]
        except Exception as e:
            print(f"Warning: Could not parse mapping {file_path}: {e}")
    return mapping

# Pre-load mappings globally
DISTRICT_CITY_MAP = load_mapping_dict(DISTRICT_MAP_PATH)
BMS_CITY_MAP = load_mapping_dict(BMS_MAP_PATH)

def get_normalized_city_name(state, raw_city, source):
    """
    If the state exists in mapping and city is found, returns reporting_city.
    Otherwise, returns raw_city (ensures cross-state safety).
    """
    lookup = DISTRICT_CITY_MAP if source == "district" else BMS_CITY_MAP
    return lookup.get((state, raw_city), raw_city)

# =========================== CORE FUNCTIONS ===========================

def get_driver(proxy=None):
    ua = UserAgent()
    options = Options()
    options.add_argument(f"user-agent={ua.random}")
    options.add_argument("--headless=new") 
    options.add_argument("start-maximized")
    options.add_argument("--disable-web-security")
    options.add_argument("--disable-site-isolation-trials")
    options.add_argument("disable-csp")
    options.add_argument("--disable-blink-features=AutomationControlled")
    
    prefs = {
        "profile.managed_default_content_settings.images": 2,
        "profile.default_content_setting_values.notifications": 2
    }
    options.add_experimental_option("prefs", prefs)

    if proxy:
        options.add_argument(f'--proxy-server={proxy}')

    return webdriver.Chrome(options=options)

# ================= NEW: TIME NORMALIZATION HELPERS =================
def district_gmt_to_ist(dt_str):
    gmt = datetime.fromisoformat(dt_str)
    ist = gmt + timedelta(hours=5, minutes=30)
    return ist.strftime("%Y-%m-%d %H:%M")

def normalize_bms_time(show_date, show_time):
    dt = datetime.strptime(f"{show_date} {show_time}", "%Y-%m-%d %I:%M %p")
    return dt.strftime("%Y-%m-%d %H:%M")

def build_seat_signature(seat_map):
    """
    Build a source-agnostic seat signature.
    Ignores category names/codes.
    Uses only sorted seat counts.
    """
    counts = sorted(seat_map.values())
    return "|".join(str(c) for c in counts)

# ================= DISTRICT LOGIC =================
def fetch_district_data(driver):
    print("\nSTARTING DISTRICT APP PROCESS...")
    if not os.path.exists(DISTRICT_CONFIG_PATH):
        print(f"District Config missing: {DISTRICT_CONFIG_PATH}")
        return []

    with open(DISTRICT_CONFIG_PATH, 'r', encoding='utf-8') as f:
        config = json.load(f)
    
    results = []
    processed_sids = set() 
    
    for state in INPUT_STATE_LIST:
        cities = config.get(state, [])
        if not cities: continue

        for city in cities:
            url = f"{DISTRICT_URL_BASE}{city['slug']}-MV203929?fromdate={SHOW_DATE}"
            print(f"[{state}] Fetching {city['name']}...", end="\r")
            
            try:
                driver.get(url)
                time.sleep(1)
                html = driver.page_source
                
                marker = 'id="__NEXT_DATA__"'
                idx = html.find(marker)
                if idx == -1: continue
                
                start = html.find('>', idx) + 1
                end = html.find('</script>', start)
                data = json.loads(html[start:end])
                
                sessions = data['props']['pageProps']['data']['serverState']['movieSessions']
                key = list(sessions.keys())[0]
                cinemas = sessions[key]['pageData']['nearbyCinemas']

                # --- NORMALIZE CITY ---
                reporting_city = get_normalized_city_name(state, city['name'], "district")

                city_res = []
                for cin in cinemas:
                    venue = cin['cinemaInfo']['name']
                    for s in cin.get('sessions', []):
                        sid = str(s.get('sid', ''))

                        if sid in processed_sids: continue
                        processed_sids.add(sid)
                        
                        b_gross, p_gross, b_tkts, t_tkts = 0, 0, 0, 0
                        seat_map = {}
                        price_seat_map = defaultdict(int)
                        price_seat_list = []

                        for a in s.get('areas', []):
                            tot, av, pr = a['sTotal'], a['sAvail'], a['price']
                            bk = tot - av
                            seat_map[a['label']] = tot
                            b_tkts += bk; t_tkts += tot
                            b_gross += (bk*pr); p_gross += (tot*pr)
                            price_seat_map[float(pr)] += tot
                            price_seat_list.append((float(pr), tot))
                        
                        occ = round((b_tkts/t_tkts)*100, 2) if t_tkts else 0
                        normalized_time = district_gmt_to_ist(s['showTime'])
                        
                        city_res.append({
                            "source": "district", "sid": sid,
                            "state": state, "city": reporting_city, "venue": venue,
                            "showTime": s['showTime'], "normalized_show_time": normalized_time,
                            "seat_category_map": seat_map, "price_seat_map": dict(price_seat_map),
                            "price_seat_signature": sorted(price_seat_list),
                            "seat_signature": build_seat_signature(seat_map),
                            "total_tickets": abs(t_tkts),
                            "booked_tickets": min(abs(b_tkts), abs(t_tkts)),
                            "total_gross": abs(p_gross),
                            "booked_gross": min(abs(int(b_gross)), abs(int(p_gross))),
                            "occupancy": min(100, abs(occ)),
                            "is_fallback": False  # District is always real
                        })
                
                if city_res:
                    gross = sum(x['booked_gross'] for x in city_res)
                    print(f"✅ {city['name']:<15} -> {reporting_city:<15} | Shows: {len(city_res):<3} | Gross: ₹{gross:<10,}")
                    results.extend(city_res)
            except Exception: pass
            
    return results

# ================= BMS LOGIC =================

def extract_initial_state_from_page(driver, url):
    try:
        driver.get(url)
        time.sleep(2)
        html = driver.page_source
        marker = "window.__INITIAL_STATE__"
        start = html.find(marker)
        if start == -1: return None 
        start = html.find("{", start)
        brace_count, end = 0, start
        while end < len(html):
            if html[end] == "{": brace_count += 1
            elif html[end] == "}": brace_count -= 1
            if brace_count == 0: break
            end += 1
        return json.loads(html[start:end + 1])
    except: return None

def extract_venues(state):
    if not state: return []
    try:
        sbe = state.get("showtimesByEvent")
        if not sbe: return []
        date_code = sbe.get("currentDateCode")
        if not date_code: return []
        widgets = sbe["showDates"][date_code]["dynamic"]["data"]["showtimeWidgets"]
        for w in widgets:
            if w.get("type") == "groupList":
                for g in w["data"]:
                    if g.get("type") == "venueGroup": return g["data"]
    except: pass
    return []

def get_seat_layout(driver, venue_code, session_id):
    """
    Returns tuple: (EncryptedDataString, ErrorMessage)
    """
    api_url = "https://services-in.bookmyshow.com/doTrans.aspx"
    max_retries = 2
    js = """
        var cb = arguments[0]; var x = new XMLHttpRequest();
        x.open("POST", "%s", true);
        x.setRequestHeader("Content-Type", "application/x-www-form-urlencoded");
        x.onload = function() { cb(x.responseText); };
        x.send("strCommand=GETSEATLAYOUT&strAppCode=WEB&strVenueCode=%s&lngTransactionIdentifier=0&strParam1=%s&strParam2=WEB&strParam5=Y&strFormat=json");
    """ % (api_url, venue_code, session_id)

    for i in range(max_retries + 1):
        try:
            resp = driver.execute_async_script(js)
            data = json.loads(resp).get("BookMyShow", {})
            
            # 1. Success
            if data.get("blnSuccess") == "true": 
                return data.get("strData"), None
            
            # 2. Extract Error
            error_msg = data.get("strException", "")
            
            # 3. Handle Rate Limit
            if "Rate limit" in error_msg:
                if i < max_retries:
                    time.sleep(60)
                    continue
                else:
                    return None, "Rate limit exceeded"
            
            # 4. Return other errors (Sold Out, etc)
            return None, error_msg
            
        except Exception as e:
            return None, str(e)
            
    return None, "Unknown Error"

def decrypt_data(enc):
    decoded = b64decode(enc)
    cipher = AES.new(ENCRYPTION_KEY.encode(), AES.MODE_CBC, iv=bytes(16))
    return unpad(cipher.decrypt(decoded), AES.block_size).decode()

def calculate_show_collection(decrypted, price_map):
    header, rows_part = decrypted.split("||")
    rows = rows_part.split("|")
    cat_map = {}
    for p in header.split("|"):
        parts = p.split(":")
        if len(parts) >= 3: cat_map[parts[1]] = parts[2]

    seats, booked = {}, {}
    for row in rows:
        if not row: continue
        parts = row.split(":")
        if len(parts) < 3: continue
        block = parts[3][0] if len(parts) > 3 else parts[2][0]
        area = cat_map.get(block)
        if not area: continue

        for seat in parts:
            if len(seat) < 2: continue
            status = seat[1]
            if seat[0] == block and status in ("1", "2"):
                seats[area] = seats.get(area, 0) + 1
            if seat[0] == block and status in BOOKED_STATES:
                booked[area] = booked.get(area, 0) + 1

    t_tkts, b_tkts, t_gross, b_gross = 0, 0, 0, 0
    for area, total in seats.items():
        bk = booked.get(area, 0)
        pr = price_map.get(area, 0)
        t_tkts += total; b_tkts += bk
        t_gross += total * pr; b_gross += bk * pr

    occ = round((b_tkts / t_tkts) * 100, 2) if t_tkts else 0
    return t_tkts, b_tkts, int(t_gross), int(b_gross), occ, seats

def process_single_city(task_data):
    """ Worker Function for BMS Parallel Execution """
    city_name, city_slug, state_name, district_sids = task_data
    current_proxy = next(proxy_pool) if proxy_pool else None
    
    # --- NORMALIZE CITY ---
    reporting_city = get_normalized_city_name(state_name, city_name, "bms")

    print(f"Starting BMS: {city_name} -> {reporting_city}...")
    url = BMS_URL_TEMPLATE.format(city=city_slug)
    driver = get_driver(current_proxy)
    city_results = []
    city_total = 0

    try:
        state_data = extract_initial_state_from_page(driver, url)
        if not state_data: return [], 0, None
        venues = extract_venues(state_data)
        
        for venue in venues:
            v_name = venue["additionalData"]["venueName"]
            v_code = venue["additionalData"]["venueCode"]
            
            # 1. Init Capacity Map
            screen_capacity_map = {}

            # 2. Get and Sort Shows (Available first)
            shows = venue.get("showtimes", [])
            shows.sort(key=lambda s: s["additionalData"].get("availStatus", "0"), reverse=True)

            for show in shows:
                sid = str(show["additionalData"]["sessionId"])
                show_time = show["title"]
                
                # 3. Determine Screen Name
                raw_screen = show.get("screenAttr", "")
                screenName = raw_screen if raw_screen else "Main Screen"
                
                
                if sid in processed_sids: continue
                processed_sids.add(sid)

                # OPTIMIZATION: Skip if already found in District
                if sid in district_sids:
                    print(f"   ⏭️  Skipping {sid} (Found in District)")
                    continue

                soldOut = False
                seat_map = {}
                is_fallback = False
                price_seat_map = {}

                try:
                    cats = show["additionalData"].get("categories", [])
                    price_map = {c["areaCatCode"]: float(c["curPrice"]) for c in cats}
                    
                    enc, error_msg = get_seat_layout(driver, v_code, sid)
                    
                    data = None
                    if not enc:
                        if not price_map: continue
                        max_price = max(price_map.values())
                        is_fallback = True

                        # Populate price_seat_map for fallback matching
                        for p in price_map.values():
                            price_seat_map[float(p)] = 0
                        
                        # SMART FALLBACK LOGIC
                        if error_msg and "sold out" in error_msg.lower():
                            # Case 1: Sold Out -> Try Recovery or Fallback
                            recovered_capacity = None
                            recovered_seat_map = None
                            
                            try:
                                base_sid = int(sid)
                                # Try offsets +7 down to +1
                                for offset in range(7, 0, -1):
                                    target_sid = str(base_sid + offset)
                                    time.sleep(1)
                                    n_enc, n_err = get_seat_layout(driver, v_code, target_sid)
                                    if n_enc:
                                        n_dec = decrypt_data(n_enc)
                                        n_res = calculate_show_collection(n_dec, {})
                                        if n_res[0] > 0:
                                            recovered_capacity = n_res[0]
                                            recovered_seat_map = n_res[5]
                                            print(f"   ✨ Fixed SoldOut {sid} using {target_sid} (Cap: {recovered_capacity})")
                                            break
                            except Exception: pass
                            
                            if recovered_capacity:
                                calc_gross = 0
                                for ac, count in recovered_seat_map.items():
                                    calc_gross += count * price_map.get(ac, 0)
                                
                                if calc_gross > 0:
                                    t_tkts = recovered_capacity; b_tkts = recovered_capacity
                                    t_gross = calc_gross; b_gross = calc_gross
                                    screen_capacity_map[screenName] = t_tkts
                                    
                                    # Update maps for better data
                                    seat_map = recovered_seat_map
                                    ps_map = defaultdict(int)
                                    for ac, count in seat_map.items():
                                        ps_map[float(price_map.get(ac, 0))] += count
                                    price_seat_map = dict(ps_map)
                                else:
                                    recovered_capacity = None

                            if not recovered_capacity:
                                if screenName in screen_capacity_map:
                                    FALLBACK_SEATS = screen_capacity_map[screenName]
                                else:
                                    FALLBACK_SEATS = 400
                                t_tkts = FALLBACK_SEATS; b_tkts = FALLBACK_SEATS
                                t_gross = int(t_tkts * max_price); b_gross = t_gross

                            occ = 100.0
                            soldOut = True
                            
                            data = {
                                "total_tickets": t_tkts, "booked_tickets": b_tkts,
                                "total_gross": t_gross, "booked_gross": b_gross, "occupancy": occ
                            }
                        elif error_msg and "Rate limit" in error_msg:
                            print(f"   Skipping {v_name[:15]} due to Rate Limit")
                            continue 
                        else:
                            # Case 3: Other -> 50%
                            t_tkts = 400; b_tkts = 200
                            t_gross = int(t_tkts * max_price)
                            b_gross = int(b_tkts * max_price)
                            occ = 50.0
                            data = {
                                "total_tickets": t_tkts, "booked_tickets": b_tkts,
                                "total_gross": t_gross, "booked_gross": b_gross, "occupancy": occ
                            }
                    else:
                        decrypted = decrypt_data(enc)
                        res = calculate_show_collection(decrypted, price_map)
                        data = {
                            "total_tickets": abs(res[0]), "booked_tickets": min(abs(res[1]), abs(res[0])),
                            "total_gross": abs(res[2]), "booked_gross": min(abs(res[3]), abs(res[2])), "occupancy": min(100, abs(res[4]))
                        }
                        seat_map = res[5]
                        
                        # Cache Capacity if successful
                        if data["total_tickets"] > 0:
                            # Build Price Seat Map for Matching
                            ps_map = defaultdict(int)
                            ps_list = []
                            for ac, count in seat_map.items():
                                pr = float(price_map.get(ac, 0))
                                ps_map[pr] += count
                                ps_list.append((pr, count))
                            price_seat_map = dict(ps_map)
                            data["price_seat_signature"] = sorted(ps_list)
                            screen_capacity_map[screenName] = data["total_tickets"]

                    if data and data['total_tickets'] > 0:
                        normalized_time = normalize_bms_time(SHOW_DATE, show_time)
                        tag = "(SOLD OUT)" if soldOut else ""
                        print(f"   [{city_name[:10]}] {v_name[:15]:<15} | {show_time} | Occ: {data['occupancy']:>5}% | {data['booked_gross']:<8} {tag}")

                        data.update({
                            "source": "bms", "sid": sid,
                            "state": state_name, "city": reporting_city,
                            "venue": v_name, "showTime": show_time,
                            "normalized_show_time": normalized_time,
                            "seat_category_map": seat_map, "price_seat_map": price_seat_map,
                            "price_seat_signature": data.get("price_seat_signature", []),
                            "seat_signature": build_seat_signature(seat_map),
                            "is_fallback": is_fallback
                        })
                        city_results.append(data)
                        city_total += data["booked_gross"]

                except Exception: continue
                time.sleep(SLEEP_TIME) 
    except Exception: pass
    finally: driver.quit()
    return city_results, city_total, url

# ================= PART 3: EXCEL GENERATOR =================
def generate_consolidated_excel(all_results, filename):
    print("\nGenerating Consolidated Excel Report...")
    wb = Workbook()
    reports_dir = "reports"
    os.makedirs(reports_dir, exist_ok=True)

    # 1. STATE WISE
    ws_state = wb.active
    ws_state.title = "State Wise"
    ws_state.append(["State", "Cities", "Theatres", "Shows", "Total Seats", "Booked Seats", "Total Gross", "Booked Gross", "Occ %"])
    state_map, city_tracker, theatre_tracker = {}, {}, {}

    for r in all_results:
        st = r["state"]
        if st not in state_map:
            state_map[st] = {"shows":0, "t_seats":0, "b_seats":0, "p_gross":0, "b_gross":0}
            city_tracker[st] = set(); theatre_tracker[st] = set()
        
        d = state_map[st]
        d["shows"] += 1; d["t_seats"] += r["total_tickets"]; d["b_seats"] += r["booked_tickets"]
        d["p_gross"] += r["total_gross"]; d["b_gross"] += r["booked_gross"]
        city_tracker[st].add(r["city"]); theatre_tracker[st].add(r["venue"])

    for st, d in state_map.items():
        avg_occ = round((d["b_seats"] / d["t_seats"]) * 100, 2) if d["t_seats"] > 0 else 0
        ws_state.append([st, len(city_tracker[st]), len(theatre_tracker[st]), d["shows"], d["t_seats"], d["b_seats"], d["p_gross"], d["b_gross"], avg_occ])

    # 2. CITY WISE (Aggregates Reporting Cities)
    ws_city = wb.create_sheet(title="City Wise")
    ws_city.append(["State", "City", "Theatres", "Shows", "Total Seats", "Booked Seats", "Total Gross", "Booked Gross", "Occ %"])
    city_map, city_theatre_tracker = {}, {}

    for r in all_results:
        k = (r["state"], r["city"])
        if k not in city_map:
            city_map[k] = {"shows":0, "t_seats":0, "b_seats":0, "p_gross":0, "b_gross":0}
            city_theatre_tracker[k] = set()
        d = city_map[k]
        d["shows"] += 1; d["t_seats"] += r["total_tickets"]; d["b_seats"] += r["booked_tickets"]
        d["p_gross"] += r["total_gross"]; d["b_gross"] += r["booked_gross"]
        city_theatre_tracker[k].add(r["venue"])

    for (st, ct), d in city_map.items():
        avg_occ = round((d["b_seats"] / d["t_seats"]) * 100, 2) if d["t_seats"] > 0 else 0
        ws_city.append([st, ct, len(city_theatre_tracker[(st, ct)]), d["shows"], d["t_seats"], d["b_seats"], d["p_gross"], d["b_gross"], avg_occ])

    # 3. THEATRE WISE
    ws_th = wb.create_sheet(title="Theatre Wise")
    ws_th.append(["Source", "State", "City", "Venue", "Shows", "Total Seats", "Booked Seats", "Total Gross", "Booked Gross", "Occ %"])
    th_map = {}
    for r in all_results:
        k = (r["source"], r["state"], r["city"], r["venue"])
        if k not in th_map: th_map[k] = {"shows":0, "t_seats":0, "b_seats":0, "p_gross":0, "b_gross":0}
        d = th_map[k]
        d["shows"] += 1; d["t_seats"] += r["total_tickets"]; d["b_seats"] += r["booked_tickets"]
        d["p_gross"] += r["total_gross"]; d["b_gross"] += r["booked_gross"]

    for (src, st, ct, vn), d in th_map.items():
        avg_occ = round((d["b_seats"] / d["t_seats"]) * 100, 2) if d["t_seats"] > 0 else 0
        ws_th.append([src, st, ct, vn, d["shows"], d["t_seats"], d["b_seats"], d["p_gross"], d["b_gross"], avg_occ])

    # 4. SHOW WISE
    ws_show = wb.create_sheet(title="Show Wise")
    ws_show.append(["Source", "State", "City", "Venue", "Time", "SID", "Total Seats", "Booked Seats", "Total Gross", "Booked Gross", "Occ %"])
    for r in all_results:
        ws_show.append([r["source"], r["state"], r["city"], r["venue"], r["showTime"], r["sid"], r["total_tickets"], r["booked_tickets"], r["total_gross"], r["booked_gross"], r["occupancy"]])

    # 5. SUMMARY
    ws_sum = wb.create_sheet(title="Summary")
    agg_p_gross = sum(r["total_gross"] for r in all_results)
    agg_b_gross = sum(r["booked_gross"] for r in all_results)
    agg_t_seats = sum(r["total_tickets"] for r in all_results)
    agg_b_seats = sum(r["booked_tickets"] for r in all_results)
    overall_occ = round((agg_b_seats / agg_t_seats) * 100, 2) if agg_t_seats > 0 else 0
    
    ws_sum.append(["Metric", "Value"])
    ws_sum.append(["States Processed", len(state_map)])
    ws_sum.append(["Total Cities", len(city_map)])
    ws_sum.append(["Total Shows", len(all_results)])
    ws_sum.append(["Total Booked Gross", agg_b_gross])
    ws_sum.append(["Overall Occupancy %", overall_occ])
    ws_sum.append(["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    path = os.path.join(reports_dir, filename)
    wb.save(path)
    print(f"Consolidated Excel Saved: {path}")

# ================= MAIN EXECUTION FLOW =================
if __name__ == "__main__":
    # 1. DISTRICT
    d_driver = get_driver()
    district_data = []

    try:
        district_data = fetch_district_data(d_driver)
        for r in district_data:
            if r["sid"]: processed_sids.add(r["sid"])
        
        if district_data:
            ts = datetime.now().strftime("%H%M")
            generate_consolidated_excel(district_data, f"District_Only_{ts}.xlsx")
    finally:
        d_driver.quit() 

    # 2. BMS
    if os.path.exists(BMS_CONFIG_PATH):
        with open(BMS_CONFIG_PATH, 'r', encoding='utf-8') as f:
            bms_config = json.load(f)

        district_known_sids = {r['sid'] for r in district_data if r['sid']}

        bms_tasks = []
        for state in INPUT_STATE_LIST:
            for city in bms_config.get(state, []):
                bms_tasks.append((city['name'], city['slug'], state, district_known_sids))

        print(f"Launching {MAX_WORKERS} Workers for {len(bms_tasks)} BMS cities...")
        bms_data = []
        last_valid_url = ""
        executor = ThreadPoolExecutor(max_workers=MAX_WORKERS)
        futures = [executor.submit(process_single_city, task) for task in bms_tasks]

        try:
            for future in as_completed(futures):
                res, total, url = future.result()
                if res:
                    bms_data.extend(res)
                    if url: last_valid_url = url
        except KeyboardInterrupt:
            executor.shutdown(wait=False)

    # 3. MERGE LOGIC
    print("\nMerging Data Sources...")
    final_data = []
    SEAT_TOLERANCE = 5

    # Index District data by (state, city, normalized_time) for fast lookup
    district_index = defaultdict(list)
    for r in district_data:
        key = (r['state'], r['city'], r['normalized_show_time'])
        district_index[key].append(r)

    for bms in bms_data:
        key = (bms['state'], bms['city'], bms['normalized_show_time'])
        candidates = district_index.get(key, [])
        match_found = None

        # 1. Try EXACT SID Match
        for cand in candidates:
            if cand['sid'] == bms['sid']:
                match_found = cand
                break
        
        # 2. Try Price Seat Signature Match (if not fallback)
        if not match_found and not bms.get('is_fallback', False):
            b_sig = bms.get('price_seat_signature', [])
            bms_venue_clean = bms['venue'].lower()
            
            for cand in candidates:
                d_sig = cand.get('price_seat_signature', [])
                if not b_sig or not d_sig: continue
                if len(b_sig) != len(d_sig): continue
                
                if all(bp == dp and abs(bs - ds) <= SEAT_TOLERANCE for (bp, bs), (dp, ds) in zip(b_sig, d_sig)):
                    # Also check fuzzy venue name to avoid false positives
                    ratio = difflib.SequenceMatcher(None, bms_venue_clean, cand['venue'].lower()).ratio()
                    if ratio > 0.4:
                        match_found = cand
                        break
        
        # 3. Try FUZZY Venue Match + Strict Price Match (if fallback)
        if not match_found and candidates:
            best_ratio = 0
            best_cand = None
            bms_venue_clean = bms['venue'].lower()
            b_prices = set(bms.get('price_seat_map', {}).keys())
            
            for cand in candidates:
                # Check price match (Strict Price Category Match)
                d_prices = set(cand.get('price_seat_map', {}).keys())
                if b_prices != d_prices: continue
                
                ratio = difflib.SequenceMatcher(None, bms_venue_clean, cand['venue'].lower()).ratio()
                if ratio > 0.5 and ratio > best_ratio:
                    best_ratio = ratio
                    best_cand = cand
            
            if best_cand:
                match_found = best_cand

        if match_found:
            candidates.remove(match_found)
            if bms.get('is_fallback', False):
                final_data.append(match_found)
            else:
                if bms['booked_gross'] > match_found['booked_gross']:
                    # Update District object with BMS stats to preserve Venue Name
                    match_found.update({
                        'total_tickets': bms['total_tickets'],
                        'booked_tickets': bms['booked_tickets'],
                        'total_gross': bms['total_gross'],
                        'booked_gross': bms['booked_gross'],
                        'occupancy': bms['occupancy'],
                        'seat_category_map': bms['seat_category_map'],
                        'price_seat_map': bms['price_seat_map'],
                        'seat_signature': bms['seat_signature']
                    })
                    final_data.append(match_found)
                else:
                    final_data.append(match_found)
        else:
            final_data.append(bms)

    # Add remaining unmatched District shows
    for sublist in district_index.values():
        final_data.extend(sublist)

    if final_data:
        ts_final = datetime.now().strftime("%Y%m%d_%H%M%S")
        generate_consolidated_excel(final_data, f"Total_States_Report_{ts_final}.xlsx")
        
        ref_url_final = last_valid_url if last_valid_url else (DISTRICT_URL_BASE + "city")
        generate_hybrid_image_report(final_data, BMS_URL_TEMPLATE, f"reports/Total_States_Report_{ts_final}.png", "bms")
    else:
        print("No data found.")