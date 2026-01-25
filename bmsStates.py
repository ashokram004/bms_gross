import json
import time
import os
import random
from base64 import b64decode
from concurrent.futures import ThreadPoolExecutor, as_completed
from itertools import cycle
from Crypto.Cipher import AES
from Crypto.Util.Padding import unpad
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from fake_useragent import UserAgent
from datetime import datetime

# --- IMPORT IMAGE GENERATOR ---
from utils.generateBMSMultiStateImageReport import generate_multi_state_image_report

# =========================== CONFIGURATION ===========================
INPUT_STATE_LIST = ["Andhra Pradesh"]
BMS_CONFIG_PATH = os.path.join("utils", "bms_cities_config.json")
MOVIE_URL_TEMPLATE = "https://in.bookmyshow.com/movies/{city}/mana-shankara-vara-prasad-garu/buytickets/ET00457184/20260125"

# AES Key
ENCRYPTION_KEY = "kYp3s6v9y$B&E)H+MbQeThWmZq4t7w!z"
BOOKED_STATES = {"2"}

# Tuning
SLEEP_TIME = 1.0        # Sleep between shows inside a worker
MAX_WORKERS = 3         # ⚡ Number of parallel browsers

# 🛡️ PROXY LIST (Optional)
PROXY_LIST = [] 
proxy_pool = cycle(PROXY_LIST) if PROXY_LIST else None

# =========================== CORE FUNCTIONS ===========================

def get_driver(proxy=None):
    ua = UserAgent()
    options = Options()
    
    # Anti-Bot Headers & Headless
    options.add_argument(f"user-agent={ua.random}")
    options.add_argument("--headless=new") 
    options.add_argument("start-maximized")
    options.add_argument("--disable-web-security")
    options.add_argument("--disable-site-isolation-trials")
    options.add_argument("disable-csp")
    options.add_argument("--disable-blink-features=AutomationControlled")
    
    # Speed Optimization: Block Images (2)
    prefs = {
        "profile.managed_default_content_settings.images": 2,
        "profile.default_content_setting_values.notifications": 2
    }
    options.add_experimental_option("prefs", prefs)

    if proxy:
        options.add_argument(f'--proxy-server={proxy}')

    return webdriver.Chrome(options=options)

def extract_initial_state_from_page(driver, url: str):
    try:
        driver.get(url)
        time.sleep(2) # React Hydration wait
        html = driver.page_source

        marker = "window.__INITIAL_STATE__"
        start = html.find(marker)
        if start == -1: return None 

        start = html.find("{", start)
        brace_count, end = 0, start

        while end < len(html):
            if html[end] == "{": brace_count += 1
            elif html[end] == "}": brace_count -= 1
            if brace_count == 0: break
            end += 1

        return json.loads(html[start:end + 1])
    except Exception:
        return None

def extract_venues(state):
    if not state: return []
    try:
        sbe = state.get("showtimesByEvent")
        if not sbe: return []
        date_code = sbe.get("currentDateCode")
        if not date_code: return []
        
        widgets = sbe["showDates"][date_code]["dynamic"]["data"]["showtimeWidgets"]
        for widget in widgets:
            if widget.get("type") == "groupList":
                for group in widget["data"]:
                    if group.get("type") == "venueGroup":
                        return group["data"]
    except Exception: pass
    return []

def get_seat_layout(driver, venue_code, session_id):
    api_url = "https://services-in.bookmyshow.com/doTrans.aspx"
    max_retries = 2

    js_script = """
        var callback = arguments[0];
        var xhr = new XMLHttpRequest();
        xhr.open("POST", "%s", true);
        xhr.setRequestHeader("Content-Type", "application/x-www-form-urlencoded");
        xhr.onload = function() { callback(xhr.responseText); };
        xhr.send("strCommand=GETSEATLAYOUT&strAppCode=WEB&strVenueCode=%s&lngTransactionIdentifier=0&strParam1=%s&strParam2=WEB&strParam5=Y&strFormat=json");
    """ % (api_url, venue_code, session_id)

    for attempt in range(max_retries + 1):
        try:
            response = driver.execute_async_script(js_script)
            data = json.loads(response).get("BookMyShow", {})
            
            if data.get("blnSuccess") == "true":
                return data.get("strData")
            
            exception = data.get("strException", "")
            if "Rate limit" in exception and attempt < max_retries:
                time.sleep(60) 
                continue
            return None
        except Exception: return None
    return None

def decrypt_data(enc):
    decoded = b64decode(enc)
    cipher = AES.new(ENCRYPTION_KEY.encode(), AES.MODE_CBC, iv=bytes(16))
    return unpad(cipher.decrypt(decoded), AES.block_size).decode()

def extract_price_map_from_show(show):
    price_map = {}
    for cat in show["additionalData"].get("categories", []):
        price_map[cat["areaCatCode"]] = float(cat["curPrice"])
    return price_map

def extract_category_map(decrypted):
    header = decrypted.split("||")[0]
    category_map = {}
    for part in header.split("|"):
        pieces = part.split(":")
        if len(pieces) >= 3: category_map[pieces[1]] = pieces[2]
    return category_map

def calculate_show_collection(decrypted, price_map):
    header, rows_part = decrypted.split("||")
    rows = rows_part.split("|")
    category_map = extract_category_map(decrypted)

    seats_map, booked_map = {}, {}

    for row in rows:
        if not row: continue
        parts = row.split(":")
        if len(parts) < 3: continue
        block_letter = parts[3][0] if len(parts) > 3 else parts[2][0]
        area_code = category_map.get(block_letter)
        if not area_code: continue

        for seat in parts:
            if len(seat) < 2: continue
            status = seat[1]
            if seat[0] == block_letter and status in ("1", "2"):
                seats_map[area_code] = seats_map.get(area_code, 0) + 1
            if status in BOOKED_STATES:
                booked_map[area_code] = booked_map.get(area_code, 0) + 1

    t_tkts, b_tkts, t_gross, b_gross = 0, 0, 0, 0
    for area, total in seats_map.items():
        booked = booked_map.get(area, 0)
        price = price_map.get(area, 0)
        t_tkts += total; b_tkts += booked
        t_gross += total * price; b_gross += booked * price

    occ = round((b_tkts / t_tkts) * 100, 2) if t_tkts else 0
    return t_tkts, b_tkts, int(t_gross), int(b_gross), occ

# =========================== WORKER FUNCTION ===========================

def process_single_city(task_data):
    """
    Worker function: Launches browser -> Scrapes 1 City -> Quits.
    """
    city_name, city_slug, state_name = task_data
    
    # 1. Select Proxy (if available)
    current_proxy = next(proxy_pool) if proxy_pool else None
    
    print(f"🚀 Starting: {city_name}...")
    url = MOVIE_URL_TEMPLATE.format(city=city_slug)
    
    driver = get_driver(current_proxy)

    city_results = []
    city_total = 0

    try:
        state = extract_initial_state_from_page(driver, url)
        if not state:
            print(f"   ⚠️ No initial state found for {city_name}")
            return [], 0, None

        venues = extract_venues(state)
        
        for venue in venues:
            venue_name = venue["additionalData"]["venueName"]
            venue_code = venue["additionalData"]["venueCode"]

            for show in venue.get("showtimes", []):
                session_id = show["additionalData"]["sessionId"]
                show_time = show["title"]
                soldOut = False

                try:
                    price_map = extract_price_map_from_show(show)
                    enc = get_seat_layout(driver, venue_code, session_id)
                    
                    data = None
                    if not enc:
                        # Fallback Logic
                        FALLBACK_TOTAL_SEATS = 200
                        if not price_map: continue
                        fallback_price = max(price_map.values())
                        data = {
                            "total_tickets": FALLBACK_TOTAL_SEATS,
                            "booked_tickets": FALLBACK_TOTAL_SEATS,
                            "occupancy": 100.0,
                            "total_gross": int(FALLBACK_TOTAL_SEATS * fallback_price),
                            "booked_gross": int(FALLBACK_TOTAL_SEATS * fallback_price)
                        }
                        soldOut = True
                    else:
                        decrypted = decrypt_data(enc)
                        # calculate_show_collection returns a tuple: (t_tkts, b_tkts, t_gross, b_gross, occ)
                        res_tuple = calculate_show_collection(decrypted, price_map)
                        data = {
                            "total_tickets": res_tuple[0],
                            "booked_tickets": res_tuple[1],
                            "total_gross": res_tuple[2],
                            "booked_gross": res_tuple[3],
                            "occupancy": res_tuple[4]
                        }

                    if data and data['total_tickets'] > 0:
                        tag = "(SOLD OUT)" if soldOut else ""
                        
                        # --- 🟢 LOGGING SHOW DETAILS ---
                        print(f"   [{city_name[:10]}] 🎬 {venue_name[:20]:<20} | {show_time} | Occ: {data['occupancy']:>5}% | Gross: ₹{data['booked_gross']:<8} {tag}")

                        data.update({
                            "state": state_name, "city": city_name,
                            "venue": venue_name, "showTime": show_time,
                            "sid": str(session_id)
                        })
                        city_results.append(data)
                        city_total += data["booked_gross"]

                except Exception: continue
                time.sleep(SLEEP_TIME) 

    except Exception as e:
        print(f"   ❌ Critical Error for {city_name}: {e}")
    finally:
        driver.quit() 
        
    print(f"✅ Finished: {city_name} | Total: ₹{city_total}")
    return city_results, city_total, url

# =========================== EXCEL REPORTING ===========================

def generate_excel(all_results):
    print("\n📊 Generating Excel Report...")
    reports_dir = "reports"
    os.makedirs(reports_dir, exist_ok=True)
    wb = Workbook()

    # SHEET 1: CITY WISE
    city_sheet = wb.active
    city_sheet.title = "City Wise Collections"
    city_sheet.append(["State", "City", "Show Count", "Total Seats", "Booked Seats", "Occupancy %", "Total Gross", "Booked Gross"])

    city_totals = {}
    for r in all_results:
        key = (r["state"], r["city"])
        if key not in city_totals:
            city_totals[key] = {
                "show_count": 0, "total_seats": 0, "booked_seats": 0,
                "occupancies": [], "total_gross": 0, "booked_gross": 0
            }
        city_totals[key]["show_count"] += 1
        city_totals[key]["total_seats"] += r["total_tickets"]
        city_totals[key]["booked_seats"] += r["booked_tickets"]
        city_totals[key]["occupancies"].append(r["occupancy"])
        city_totals[key]["total_gross"] += r["total_gross"]
        city_totals[key]["booked_gross"] += r["booked_gross"]

    for (st, ct), data in city_totals.items():
        avg_occ = round(sum(data["occupancies"]) / data["show_count"], 2) if data["show_count"] else 0
        city_sheet.append([
            st, ct, data["show_count"], data["total_seats"], data["booked_seats"],
            avg_occ, data["total_gross"], data["booked_gross"]
        ])

    tc_shows = sum(d["show_count"] for d in city_totals.values())
    tc_seats = sum(d["total_seats"] for d in city_totals.values())
    tc_booked = sum(d["booked_seats"] for d in city_totals.values())
    tc_gross = sum(d["total_gross"] for d in city_totals.values())
    tc_bgross = sum(d["booked_gross"] for d in city_totals.values())
    tc_occ = round((tc_booked / tc_seats) * 100, 2) if tc_seats else 0
    city_sheet.append(["TOTAL", "", tc_shows, tc_seats, tc_booked, tc_occ, tc_gross, tc_bgross])

    # SHEET 2: THEATRE WISE
    theatre_sheet = wb.create_sheet(title="Theatre Wise Collections")
    theatre_sheet.append(["State", "City", "Venue", "Show count", "Total Seats", "Booked Seats", "Occupancy %", "Total Gross", "Booked Gross"])

    theatre_data = {}
    for r in all_results:
        key = (r["state"], r["city"], r["venue"])
        if key not in theatre_data:
            theatre_data[key] = {
                "num_shows": 0, "total_tickets": 0, "booked_tickets": 0,
                "occupancies": [], "total_gross": 0, "booked_gross": 0
            }
        theatre_data[key]["num_shows"] += 1
        theatre_data[key]["total_tickets"] += r["total_tickets"]
        theatre_data[key]["booked_tickets"] += r["booked_tickets"]
        theatre_data[key]["occupancies"].append(r["occupancy"])
        theatre_data[key]["total_gross"] += r["total_gross"]
        theatre_data[key]["booked_gross"] += r["booked_gross"]

    for key, data in theatre_data.items():
        avg_occ = round(sum(data["occupancies"]) / data["num_shows"], 2) if data["num_shows"] else 0
        theatre_sheet.append([
            key[0], key[1], key[2], data["num_shows"], data["total_tickets"], data["booked_tickets"],
            avg_occ, data["total_gross"], data["booked_gross"]
        ])

    # SHEET 3: SHOW WISE
    show_sheet = wb.create_sheet(title="Show Wise Collections")
    show_sheet.append(["State", "City", "Venue", "Show Time", "Total Seats", "Booked Seats", "Occupancy %", "Total Gross", "Booked Gross"])
    for r in all_results:
        show_sheet.append([
            r["state"], r["city"], r["venue"], r["showTime"], r["total_tickets"], r["booked_tickets"],
            r["occupancy"], r["total_gross"], r["booked_gross"]
        ])

    # SHEET 4: SUMMARY
    summary = wb.create_sheet(title="Summary")
    summary.append(["Metric", "Value"])
    summary.append(["Total Cities", len(city_totals)])
    summary.append(["Total Theatres", len(theatre_data)])
    summary.append(["Total Shows", len(all_results)])
    summary.append(["Overall Occupancy (%)", tc_occ])
    summary.append(["Total Potential Gross (INR)", tc_gross])
    summary.append(["Total Booked Gross (INR)", tc_bgross])
    summary.append(["Report Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    file_ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"bms_states_report_{file_ts}.xlsx"
    filepath = os.path.join(reports_dir, filename)
    wb.save(filepath)
    print(f"✅ Excel report saved at: {filepath}")

# =========================== MAIN EXECUTION ===========================

if __name__ == "__main__":
    if not os.path.exists(BMS_CONFIG_PATH):
        print("❌ Config file missing.")
        exit()

    with open(BMS_CONFIG_PATH, 'r', encoding='utf-8') as f:
        bms_config = json.load(f)

    tasks = []
    print(f"📋 Reading config for states: {INPUT_STATE_LIST}")
    for state in INPUT_STATE_LIST:
        cities = bms_config.get(state, [])
        for city in cities:
            tasks.append((city['name'], city['slug'], state))

    print(f"🔥 Launching {MAX_WORKERS} Parallel Workers for {len(tasks)} cities...")
    
    global_results = []
    last_valid_url = ""
    start_time = time.time()

    executor = ThreadPoolExecutor(max_workers=MAX_WORKERS)
    futures = []

    try:
        # Submit tasks
        for task in tasks:
            future = executor.submit(process_single_city, task)
            futures.append(future)
        
        # Collect results as they finish
        for future in as_completed(futures):
            try:
                res, total, url = future.result()
                if res:
                    global_results.extend(res)
                    if url: last_valid_url = url
            except Exception as exc:
                print(f"💥 Thread Exception: {exc}")

    except KeyboardInterrupt:
        print("\n\n🛑 STOPPING! Cancelling pending tasks...")
        executor.shutdown(wait=False, cancel_futures=True)

    duration = time.time() - start_time
    print(f"\n🏁 ALL DONE in {round(duration, 2)}s")

    # 3. Generate Reports
    if global_results:
        generate_excel(global_results)
        
        # Use valid url for metadata in image report
        if not last_valid_url and tasks:
             last_valid_url = MOVIE_URL_TEMPLATE.format(city=tasks[0][1])
             
        img_path = f"reports/bms_states_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        generate_multi_state_image_report(global_results, last_valid_url, img_path)
    else:
        print("❌ No data collected.")