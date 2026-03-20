import json
import os
import time
import random
import shutil
import threading
import requests
from datetime import datetime, timedelta
from base64 import b64decode
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()
from Crypto.Cipher import AES
from Crypto.Util.Padding import unpad
from fake_useragent import UserAgent
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
import difflib
from collections import defaultdict, deque
import math
from concurrent.futures import ThreadPoolExecutor, as_completed, wait

from utils.generatePremiumCityImageReport import generate_premium_city_image_report
from utils.generateHybridCityHTMLReport import generate_hybrid_city_html_report
from utils.sendReportEmail import send_collection_report

# =============================================================================
# ── CONFIGURATION ─────────────────────────────────────────────────────────────
# =============================================================================

SHOW_DATE = "2026-03-19"

DISTRICT_URL_TEMPLATE = (
    "https://www.district.in/movies/ustaad-bhagat-singh-movie-tickets-in-{city}-MV161614"
    "?frmtid=tvqjmjqme&fromdate=" + SHOW_DATE
)
BMS_URL_TEMPLATE = (
    "https://in.bookmyshow.com/movies/{city}/ustaad-bhagat-singh/buytickets/ET00339939/20260319"
)

DISTRICT_CITIES = [
    "hyderabad"
]
BMS_CITIES = [
    "hyderabad"
]

BMS_KEY      = "kYp3s6v9y$B&E)H+MbQeThWmZq4t7w!z"
BOOKED_CODES = {"2"}
SLEEP_TIME   = 0.5
VENUE_WORKERS = 3   # parallel venue workers per city (both District & BMS)
BMS_RATE_LIMIT_WAIT = 30

# =============================================================================
# ── HELPERS ───────────────────────────────────────────────────────────────────
# =============================================================================

def extract_movie_name_from_url(url):
    try:
        if '/movies/' in url and '/buytickets/' in url:
            parts      = url.split('/movies/')[1].split('/buytickets/')[0].split('/')
            movie_slug = parts[-1] if len(parts) > 1 else parts[0]
            return movie_slug.replace('-', ' ').title()
        if '/movies/' in url and '-movie-tickets-in-' in url:
            movie_slug = url.split('/movies/')[1].split('-movie-tickets-in-')[0]
            return movie_slug.replace('-', ' ').title()
    except Exception as e:
        print(f"Could not extract movie name from URL: {e}")
    return "Movie Collection"


# =============================================================================
# ── HTTP SESSION (thread-local, connection-pooled) ────────────────────────────
# =============================================================================

_thread_local = threading.local()

def get_http_session():
    if not hasattr(_thread_local, 'session'):
        s = requests.Session()
        ua = UserAgent()
        s.headers.update({
            'User-Agent': ua.random,
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
        })
        adapter = requests.adapters.HTTPAdapter(
            pool_connections=20,
            pool_maxsize=20,
            max_retries=requests.adapters.Retry(total=2, backoff_factor=0.5,
                                                 status_forcelist=[502, 503, 504]),
        )
        s.mount('https://', adapter)
        s.mount('http://', adapter)
        _thread_local.session = s
    return _thread_local.session


def get_driver():
    ua = UserAgent()
    options = Options()
    options.add_argument(f"user-agent={ua.random}")
    options.add_argument("--headless=new")
    options.add_argument("start-maximized")
    options.add_argument("--disable-web-security")
    options.add_argument("--disable-site-isolation-trials")
    options.add_argument("disable-csp")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    prefs = {
        "profile.managed_default_content_settings.images": 2,
        "profile.default_content_setting_values.notifications": 2,
    }
    options.add_experimental_option("prefs", prefs)
    driver = webdriver.Chrome(options=options)
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined});"
    })
    return driver


def district_gmt_to_ist(dt_str):
    gmt = datetime.fromisoformat(dt_str)
    ist = gmt + timedelta(hours=5, minutes=30)
    return ist.strftime("%Y-%m-%d %H:%M")

def normalize_bms_time(show_date, show_time):
    dt = datetime.strptime(f"{show_date} {show_time}", "%Y-%m-%d %I:%M %p")
    return dt.strftime("%Y-%m-%d %H:%M")

def build_seat_signature(seat_map):
    return "|".join(str(c) for c in sorted(seat_map.values()))


# =============================================================================
# ── DISTRICT ──────────────────────────────────────────────────────────────────
# =============================================================================

def get_district_seat_layout_http(cinema_id, session_id):
    """Direct HTTP POST for District seat layout API — no Selenium needed."""
    api_url = "https://www.district.in/gw/consumer/movies/v1/select-seat"
    params  = {
        "version": "3", "site_id": "1", "channel": "mweb",
        "child_site_id": "1", "platform": "district",
    }
    payload = {"cinemaId": int(cinema_id), "sessionId": str(session_id)}
    headers = {
        "Content-Type": "application/json",
        "x-guest-token": str(random.randint(1, 9999999999)),
        "Origin": "https://www.district.in",
        "Referer": "https://www.district.in/",
    }
    try:
        session = get_http_session()
        resp = session.post(api_url, params=params, json=payload,
                            headers=headers, timeout=10)
        if resp.status_code == 200:
            return resp.json()
    except Exception:
        pass
    return None


def district_process_venue_http(cin, city_name):
    """Processes a single District venue via HTTP — no Selenium driver needed.
    Uses global _global_district_sids set to skip already-processed SIDs across all workers.
    """
    results = []
    try:
        venue = cin['entityName']
        for s in cin.get('sessions', []):
            sid = str(s.get('sid', ''))
            cid = s.get('cid')

            # Check if SID already processed (thread-safe global check)
            with _global_district_sids_lock:
                if sid in _global_district_sids:
                    continue
                _global_district_sids.add(sid)

            code_to_label  = {}
            default_prices = {}
            for a in s.get('areas', []):
                code_to_label[a['code']]  = a['label']
                default_prices[a['code']] = float(a['price'])

            b_gross, p_gross, b_tkts, t_tkts = 0, 0, 0, 0
            seat_map        = defaultdict(int)
            label_price_map = {}
            layout_res      = None

            if cid:
                layout_res = get_district_seat_layout_http(cid, sid)

            if layout_res and 'seatLayout' in layout_res:
                for area in layout_res['seatLayout'].get('colAreas', {}).get('objArea', []):
                    area_code = area.get('AreaCode')
                    label     = code_to_label.get(area_code, area_code)
                    price     = float(area.get('AreaPrice') or default_prices.get(area_code, 0))
                    label_price_map[label] = price
                    for row in area.get('objRow', []):
                        for seat in row.get('objSeat', []):
                            status = seat.get('SeatStatus')
                            t_tkts += 1; p_gross += price
                            seat_map[label] += 1
                            if status != '0' and status != 0:
                                b_tkts += 1; b_gross += price
            else:
                for a in s.get('areas', []):
                    tot, av, pr = a['sTotal'], a['sAvail'], a['price']
                    bk = tot - av
                    seat_map[a['label']]       = tot
                    label_price_map[a['label']] = float(pr)
                    b_tkts += bk; t_tkts += tot
                    b_gross += bk * pr; p_gross += tot * pr

            price_seat_map  = defaultdict(int)
            price_seat_list = []
            for label, count in seat_map.items():
                pr = label_price_map.get(label, 0.0)
                price_seat_map[float(pr)] += count
                price_seat_list.append((float(pr), count))

            occ             = round((b_tkts / t_tkts) * 100, 2) if t_tkts else 0
            normalized_time = district_gmt_to_ist(s['showTime'])

            print(
                f"   🎬 [District][{city_name}] {venue[:20]:<20} | {normalized_time} | "
                f"Occ: {occ:>5}% | Gross: ₹{int(b_gross):<8,}"
            )

            results.append({
                "source":               "district",
                "sid":                  sid,
                "city":                 city_name,
                "venue":                venue,
                "cinema_id":            str(cid) if cid else "",
                "showTime":             s['showTime'],
                "normalized_show_time": normalized_time,
                "seat_category_map":    dict(seat_map),
                "price_seat_map":       dict(price_seat_map),
                "price_seat_signature": sorted(price_seat_list),
                "seat_signature":       build_seat_signature(seat_map),
                "total_tickets":        abs(t_tkts),
                "booked_tickets":       min(abs(b_tkts), abs(t_tkts)),
                "total_gross":          abs(p_gross),
                "booked_gross":         min(abs(int(b_gross)), abs(int(p_gross))),
                "occupancy":            min(100, abs(occ)),
                "is_fallback":          False,
            })
    except Exception as e:
        print(f"❌ [District][{city_name}] Venue worker error: {e}")
    return results


def fetch_district_city(city_name, city_slug, city_index, total_cities):
    """
    Fetches all District data for one city via HTTP (no Selenium).
    Step 1: HTTP GET page to extract __NEXT_DATA__ JSON.
    Step 2: VENUE_WORKERS parallel HTTP calls for seat layouts.
    """
    url = DISTRICT_URL_TEMPLATE.replace("{city}", city_slug)
    print(f"\n🏙️  [District] City {city_index}/{total_cities}: {city_name}")

    # Step 1: HTTP page scrape
    cinemas = []
    for attempt in range(2):
        try:
            session = get_http_session()
            resp    = session.get(url, timeout=15)

            if resp.status_code == 403 and attempt == 0:
                time.sleep(10)
                if hasattr(_thread_local, 'session'):
                    delattr(_thread_local, 'session')
                continue

            if resp.status_code != 200:
                print(f"   ⚠️  [District][{city_name}] HTTP {resp.status_code}")
                return []
            html = resp.text

            marker = 'id="__NEXT_DATA__"'
            idx    = html.find(marker)
            if idx == -1:
                print(f"   ⚠️  [District][{city_name}] No page data found.")
                return []

            start    = html.find('>', idx) + 1
            end      = html.find('</script>', start)
            data     = json.loads(html[start:end])
            sessions = data['props']['pageProps']['data']['serverState']['movieSessions']
            if not sessions:
                return []
            key      = list(sessions.keys())[0]
            cinemas  = sessions[key].get('arrangedSessions', [])
            break
        except Exception as e:
            print(f"❌ [District][{city_name}] Page scrape error: {e}")
            return []

    if not cinemas:
        print(f"   ⚠️  [District][{city_name}] No sessions found.")
        return []

    # Step 2: Parallel HTTP venue processing (no drivers needed)
    city_results   = []
    with ThreadPoolExecutor(max_workers=VENUE_WORKERS) as executor:
        futures = [
            executor.submit(district_process_venue_http, cin, city_name)
            for cin in cinemas
        ]
        for future in as_completed(futures):
            try:
                city_results.extend(future.result())
            except Exception as e:
                print(f"❌ [District][{city_name}] Venue worker exception: {e}")

    gross = sum(r['booked_gross'] for r in city_results)
    print(f"   ✅ [District] {city_name:<20} | Shows: {len(city_results):<3} | Gross: ₹{gross:<10,}")
    return city_results


def run_district(city_pairs):
    """
    Main District runner. Fully HTTP-based (no Selenium).
    Processes cities sequentially, VENUE_WORKERS parallel HTTP calls per city.
    """
    all_results  = []
    total        = len(city_pairs)
    print(f"\n🚀 [District] Starting — {total} cities, {VENUE_WORKERS} venue workers per city (HTTP)")

    for idx, (city_name, city_slug) in enumerate(city_pairs, 1):
        results = fetch_district_city(city_name, city_slug, idx, total)
        all_results.extend(results)

    print(f"\n✅ [District] Done — {len(all_results)} total shows across {total} cities.")
    return all_results


# =============================================================================
# ── BMS ───────────────────────────────────────────────────────────────────────
# =============================================================================

def decrypt_data(enc):
    try:
        decoded = b64decode(enc)
        cipher  = AES.new(BMS_KEY.encode(), AES.MODE_CBC, iv=bytes(16))
        return unpad(cipher.decrypt(decoded), AES.block_size).decode()
    except:
        return None

def calculate_bms_collection(decrypted, price_map):
    header, rows_part = decrypted.split("||")
    rows = rows_part.split("|")

    cat_map         = {}
    local_price_map = price_map.copy()
    last_price      = 0.0

    for p in header.split("|"):
        parts = p.split(":")
        if len(parts) >= 3:
            cat_map[parts[1]] = parts[2]
            current_price = local_price_map.get(parts[2], 0.0)
            if current_price > 0:   last_price = current_price
            elif last_price > 0:    local_price_map[parts[2]] = last_price

    seats, booked = {}, {}
    for row in rows:
        if not row: continue
        parts = row.split(":")
        if len(parts) < 3: continue
        block = parts[3][0] if len(parts) > 3 else parts[2][0]
        area  = cat_map.get(block)
        if not area: continue
        for seat in parts:
            if len(seat) < 2: continue
            status = seat[1]
            if seat[0] == block and status in ("1", "2"):
                seats[area] = seats.get(area, 0) + 1
            if seat[0] == block and status in BOOKED_CODES:
                booked[area] = booked.get(area, 0) + 1

    t_tkts, b_tkts, t_gross, b_gross = 0, 0, 0, 0
    for area, total in seats.items():
        bk = booked.get(area, 0); pr = local_price_map.get(area, 0)
        t_tkts += total; b_tkts += bk
        t_gross += total * pr; b_gross += bk * pr

    occ = round((b_tkts / t_tkts) * 100, 2) if t_tkts else 0
    return t_tkts, b_tkts, int(t_gross), int(b_gross), occ, seats, local_price_map


def get_seat_layout(driver, venue_code, session_id):
    api = "https://services-in.bookmyshow.com/doTrans.aspx"
    js  = f"""
    var cb = arguments[0]; var x = new XMLHttpRequest();
    x.open("POST", "{api}", true);
    x.setRequestHeader("Content-Type", "application/x-www-form-urlencoded");
    x.onload = function() {{ cb(x.responseText); }};
    x.send("strCommand=GETSEATLAYOUT&strAppCode=WEB&strVenueCode={venue_code}&strParam1={session_id}&strParam2=WEB&strParam5=Y&strFormat=json");
    """
    while True:
        try:
            driver.set_script_timeout(10)
            resp = driver.execute_async_script(js)
            j    = json.loads(resp).get("BookMyShow", {})
            if j.get("blnSuccess") == "true":
                return j.get("strData"), None
            err = j.get("strException", "")
            if any(kw in err.lower() for kw in ["rate limit", "connectivity issue", "high demand"]):
                time.sleep(BMS_RATE_LIMIT_WAIT)
                continue
            return None, err
        except Exception as e:
            if "timeout" in str(e).lower():
                continue  # retry immediately
            return None, str(e)


def get_seat_layout_http(venue_code, session_id):
    """HTTP-based BMS seat layout call — no Selenium driver needed.
    Much faster and allows parallel venue processing."""
    api_url = "https://services-in.bookmyshow.com/doTrans.aspx"
    payload = (
        f"strCommand=GETSEATLAYOUT&strAppCode=WEB&strVenueCode={venue_code}"
        f"&lngTransactionIdentifier=0&strParam1={session_id}"
        f"&strParam2=WEB&strParam5=Y&strFormat=json"
    )
    headers = {
        "Content-Type": "application/x-www-form-urlencoded",
        "User-Agent": UserAgent().random,
        "Origin": "https://in.bookmyshow.com",
        "Referer": "https://in.bookmyshow.com/",
    }
    while True:
        try:
            resp = requests.post(api_url, data=payload, headers=headers, timeout=15)
            if resp.status_code == 429:
                print(f"      ⏳ HTTP 429 — waiting {BMS_RATE_LIMIT_WAIT}s...")
                time.sleep(BMS_RATE_LIMIT_WAIT)
                continue
            if resp.status_code != 200:
                return None, f"HTTP {resp.status_code}"
            data = resp.json().get("BookMyShow", {})
            if data.get("blnSuccess") == "true":
                return data.get("strData"), None
            error_msg = data.get("strException", "")
            if any(kw in error_msg.lower() for kw in ["rate limit", "connectivity issue", "high demand"]):
                print(f"      ⏳ Rate limit response — waiting {BMS_RATE_LIMIT_WAIT}s...")
                time.sleep(BMS_RATE_LIMIT_WAIT)
                continue
            return None, error_msg
        except Exception as e:
            if "timeout" in str(e).lower():
                continue  # retry immediately
            return None, str(e).split('\n')[0]


# Global: detect whether HTTP seat layout works (tested once)
_bms_http_tested = False
_bms_http_works  = False
_bms_http_lock   = threading.Lock()

# Global BMS SIDs set (shared across all cities/workers)
_global_bms_sids = set()
_global_bms_sids_lock = threading.Lock()

# Global District SIDs set (shared across all cities/workers)
_global_district_sids = set()
_global_district_sids_lock = threading.Lock()


def bms_fetch_single_venue(venue, city_name, use_http=False):
    """Processes a single BMS venue. Uses HTTP if available, otherwise Selenium driver.
    Clean version: no fallback/recovery logic. Skips shows on any error except rate limit."""
    results = []
    driver  = None if use_http else get_driver()

    try:
        if True:  # single venue
            v_name = venue["additionalData"]["venueName"]
            v_code = venue["additionalData"]["venueCode"]

            shows      = venue.get("showtimes", [])
            shows.sort(key=lambda s: s["additionalData"].get("availStatus", "0"), reverse=True)
            show_queue = deque(shows)

            while show_queue:
                show      = show_queue.popleft()
                sid       = str(show["additionalData"]["sessionId"])
                show_time = show["title"]

                # Check if SID already processed (thread-safe global check)
                with _global_bms_sids_lock:
                    if sid in _global_bms_sids:
                        continue
                    _global_bms_sids.add(sid)

                try:
                    cats      = show["additionalData"].get("categories", [])
                    price_map = {c["areaCatCode"]: float(c["curPrice"]) for c in cats}
                    enc, error_msg = get_seat_layout_http(v_code, sid) if use_http else get_seat_layout(driver, v_code, sid)

                    if not enc:
                        # get_seat_layout_http already retried rate limits with 30s waits
                        print(f"      ⏭️  [BMS][{city_name}] Skipping {sid}: {error_msg}")
                        continue

                    decrypted = decrypt_data(enc)
                    res       = calculate_bms_collection(decrypted, price_map)
                    data      = {
                        "total_tickets":  abs(res[0]),
                        "booked_tickets": min(abs(res[1]), abs(res[0])),
                        "total_gross":    abs(res[2]),
                        "booked_gross":   min(abs(res[3]), abs(res[2])),
                        "occupancy":      min(100, abs(res[4])),
                    }
                    seat_map        = res[5]
                    final_price_map = res[6]
                    price_seat_map  = {}

                    if data["total_tickets"] > 0:
                        ps_map  = defaultdict(int); ps_list = []
                        for ac, count in seat_map.items():
                            pr = float(final_price_map.get(ac, 0))
                            ps_map[pr] += count; ps_list.append((pr, count))
                        price_seat_map             = dict(ps_map)
                        data["price_seat_signature"] = sorted(ps_list)

                    if data and data['total_tickets'] > 0:
                        normalized_time = normalize_bms_time(SHOW_DATE, show_time)
                        print(
                            f"   🎬 [BMS][{city_name}] {v_name[:15]:<15} | {normalized_time} | "
                            f"Occ: {data['occupancy']:>5}% | Gross: ₹{data['booked_gross']:<8,}"
                        )
                        data.update({
                            "source":               "bms",
                            "sid":                  sid,
                            "city":                 city_name,
                            "venue":                v_name,
                            "venue_code":           v_code,
                            "showTime":             show_time,
                            "normalized_show_time": normalized_time,
                            "seat_category_map":    seat_map,
                            "price_seat_map":       price_seat_map,
                            "price_seat_signature": data.get("price_seat_signature", []),
                            "seat_signature":       build_seat_signature(seat_map),
                            "is_fallback":          False,
                        })
                        results.append(data)

                except Exception:
                    continue

    except Exception as e:
        print(f"❌ [BMS] Venue worker error for {city_name}: {e}")
    finally:
        if driver:
            driver.quit()

    return results


def fetch_bms_city(city_name, city_slug, city_index, total_cities):
    """
    Fetches all BMS data for one city.
    Step 1: Fresh stealth driver scrapes BMS page for venue list.
    Step 2: If HTTP seat layout works, quits driver early and
            processes venues in parallel via HTTP. Otherwise Selenium fallback.
    """
    global _bms_http_tested, _bms_http_works

    url = BMS_URL_TEMPLATE.replace("{city}", city_slug)
    print(f"\n🏙️  [BMS] City {city_index}/{total_cities}: {city_name}")

    # Step 1: Scrape page for venue list (fresh driver per city for Cloudflare)
    page_driver = get_driver()
    venues      = []
    try:
        page_driver.set_script_timeout(20)
        page_driver.get(url)
        time.sleep(1)
        html = page_driver.page_source

        marker = "window.__INITIAL_STATE__"
        start  = html.find(marker)
        if start == -1:
            # Retry once with extra wait
            time.sleep(1)
            html  = page_driver.page_source
            start = html.find(marker)
            if start == -1:
                print(f"   ⚠️  [BMS][{city_name}] Could not find initial state.")
                return []

        start = html.find("{", start)
        brace = 0; end = start
        while end < len(html):
            if html[end] == "{": brace += 1
            elif html[end] == "}": brace -= 1
            if brace == 0: break
            end += 1

        state_data = json.loads(html[start:end + 1])
        sbe        = state_data.get("showtimesByEvent")
        dc         = sbe.get("currentDateCode")
        widgets    = sbe["showDates"][dc]["dynamic"]["data"]["showtimeWidgets"]
        for w in widgets:
            if w.get("type") == "groupList":
                for g in w["data"]:
                    if g.get("type") == "venueGroup":
                        venues = g["data"]
    except Exception as e:
        print(f"❌ [BMS][{city_name}] Page scrape error: {e}")
        return []
    finally:
        page_driver.quit()

    if not venues:
        print(f"   ⚠️  [BMS][{city_name}] No venues found.")
        return []

    # Probe HTTP mode once (first city with venues)
    with _bms_http_lock:
        if not _bms_http_tested:
            for v in venues:
                shows = v.get("showtimes", [])
                if not shows:
                    continue
                test_vcode = v["additionalData"]["venueCode"]
                test_sid   = str(shows[0]["additionalData"]["sessionId"])
                enc, err   = get_seat_layout_http(test_vcode, test_sid)
                _bms_http_works = enc is not None or (err and "sold out" in err.lower())
                break
            _bms_http_tested = True
            mode = "parallel HTTP" if _bms_http_works else "Selenium"
            print(f"   🔍 [BMS] Seat layout mode: {mode}")

    # Step 2: Process venues sequentially (speed test: no threading)
    city_results = []
    for venue in venues:
        city_results.extend(
            bms_fetch_single_venue(venue, city_name, _bms_http_works)
        )

    gross = sum(r['booked_gross'] for r in city_results)
    print(f"   ✅ [BMS] {city_name:<20} | Shows: {len(city_results):<3} | Gross: ₹{gross:<10,}")
    return city_results


def run_bms(city_pairs):
    """
    Main BMS runner — SPEED TEST version.
    Sequential processing: no threading, no sleeps (except 30s on rate limit).
    """
    all_results = []
    total       = len(city_pairs)
    print(f"\n🚀 [BMS] Starting SPEED TEST — {total} cities, sequential (no threading, no sleeps)")
    print(f"   Rate limit strategy: wait 30s and retry\n")

    for idx, (city_name, city_slug) in enumerate(city_pairs, 1):
        results = fetch_bms_city(city_name, city_slug, idx, total)
        all_results.extend(results)

    print(f"\n✅ [BMS] Done — {len(all_results)} total shows across {total} cities.")
    return all_results


# =============================================================================
# ── VENUE MAPPING ─────────────────────────────────────────────────────────────
# =============================================================================

VENUE_MAP = {}   # BMS VenueCode → District cinema_id

def load_venue_mapping():
    global VENUE_MAP
    mapping_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'utils', 'venue_mapping.json')
    if os.path.exists(mapping_path):
        with open(mapping_path, encoding='utf-8') as f:
            data = json.load(f)
        VENUE_MAP = data.get('bms_to_district', {})
        print(f"📍 Loaded venue mapping: {len(VENUE_MAP)} BMS→District pairs")
    else:
        print("⚠️  Venue mapping not found (utils/venue_mapping.json). Venue-based matching disabled.")

def _is_same_venue(bms_show, dist_show):
    """Check venue mapping. Returns True/False if mapping exists, 'unmapped' if BMS venue not in map."""
    bms_code = bms_show.get('venue_code', '')
    dist_cid = dist_show.get('cinema_id', '')
    if not bms_code or not dist_cid:
        return 'unmapped'
    mapped_dist = VENUE_MAP.get(bms_code)
    if mapped_dist is None:
        return 'unmapped'
    return mapped_dist == dist_cid

# =============================================================================
# ── DEDUP + MERGE ─────────────────────────────────────────────────────────────
# =============================================================================

def dedup_same_platform(records, source_label):
    seen    = {}
    best    = []
    dropped = 0
    for r in records:
        sid = r.get('sid', '')
        if not sid:
            best.append(r); continue
        if sid not in seen:
            seen[sid] = len(best)
            best.append(r)
        else:
            dropped += 1
            existing = best[seen[sid]]
            if r.get('booked_gross', 0) > existing.get('booked_gross', 0):
                best[seen[sid]] = r
                print(f"   ♻️  [{source_label}] Replaced lower-gross duplicate SID {sid}")
            else:
                print(f"   ♻️  [{source_label}] Dropped duplicate SID {sid}")
    if dropped:
        print(f"   ✅ [{source_label}] Dedup removed {dropped} duplicate(s). {len(best)} unique shows remain.")
    return best


def merge_data(dist_data, bms_data):
    dist_data = dedup_same_platform(dist_data, "District")
    bms_data  = dedup_same_platform(bms_data,  "BMS")

    print(f"\n🔄 Merging {len(dist_data)} District + {len(bms_data)} BMS shows...")
    final_data     = []
    SEAT_TOLERANCE = 5

    district_by_time = defaultdict(list)
    for r in dist_data:
        district_by_time[(r['city'], r['normalized_show_time'])].append(r)

    for bms in bms_data:
        key        = (bms['city'], bms['normalized_show_time'])
        candidates = district_by_time.get(key, [])
        match      = None

        # 1. Exact SID
        for c in candidates:
            if c['sid'] == bms['sid']:
                match = c; print(f"   🔗 SID Match: {bms['sid']}"); break

        # 2. Price + Seat signature + Venue Map
        if not match and not bms.get('is_fallback', False):
            b_sig = bms.get('price_seat_signature', [])
            for c in candidates:
                d_sig = c.get('price_seat_signature', [])
                if not b_sig or not d_sig or len(b_sig) != len(d_sig): continue
                if all(bp == dp and abs(bs - ds) <= SEAT_TOLERANCE
                       for (bp, bs), (dp, ds) in zip(b_sig, d_sig)):
                    if _is_same_venue(bms, c) is True:
                        match = c
                        print(f"   🔗 Price/Seat Sig + Venue Map: {bms['venue']} == {c['venue']}")
                        break

        # 3. Seat signature + Venue Map
        if not match and not bms.get('is_fallback', False):
            b_seats = sorted(bms.get('seat_category_map', {}).values())
            for c in candidates:
                d_seats = sorted(c.get('seat_category_map', {}).values())
                if not b_seats or not d_seats or len(b_seats) != len(d_seats): continue
                if all(abs(bs - ds) <= SEAT_TOLERANCE for bs, ds in zip(b_seats, d_seats)):
                    if _is_same_venue(bms, c) is True:
                        match = c
                        print(f"   🔗 Seat Sig + Venue Map: {bms['venue']} == {c['venue']}")
                        break

        # 4. Venue map + strict price set
        if not match and candidates:
            b_prices = {p for p in bms.get('price_seat_map', {}).keys() if p > 0}
            for c in candidates:
                d_prices = {p for p in c.get('price_seat_map', {}).keys() if p > 0}
                if b_prices != d_prices: continue
                if _is_same_venue(bms, c) is True:
                    match = c
                    print(f"   🔗 Venue Map + Price: {bms['venue']} == {c['venue']}")
                    break

        # 5. Venue map only (seats/prices may differ between platforms)
        if not match and candidates:
            for c in candidates:
                if _is_same_venue(bms, c) is True:
                    match = c
                    print(f"   🔗 Venue Map Only: {bms['venue']} == {c['venue']}")
                    break

        if match:
            candidates.remove(match)
            if not bms.get('is_fallback', False) and bms['booked_gross'] > match['booked_gross']:
                for k in ('total_tickets','booked_tickets','total_gross','booked_gross',
                          'occupancy','seat_category_map','price_seat_map','seat_signature'):
                    match[k] = bms[k]
            final_data.append(match)
        else:
            final_data.append(bms)

    for sublist in district_by_time.values():
        final_data.extend(sublist)

    print(f"✅ Merge complete — {len(final_data)} final shows.")
    return final_data


# =============================================================================
# ── EXCEL ─────────────────────────────────────────────────────────────────────
# =============================================================================

def generate_excel(data, filename):
    wb          = Workbook()
    reports_dir = "reports"
    os.makedirs(reports_dir, exist_ok=True)

    # Theatre Wise
    ws_th = wb.active
    ws_th.title = "Theatre Wise"
    ws_th.append(["Venue","Shows","Total Seats","Booked Seats","Total Gross","Booked Gross","Occ %"])
    th_map = {}
    for r in data:
        k = r["venue"]
        if k not in th_map:
            th_map[k] = {"shows":0,"t_seats":0,"b_seats":0,"p_gross":0,"b_gross":0}
        d = th_map[k]
        d["shows"]+=1; d["t_seats"]+=r["total_tickets"]; d["b_seats"]+=r["booked_tickets"]
        d["p_gross"]+=r["total_gross"]; d["b_gross"]+=r["booked_gross"]
    for v, d in th_map.items():
        occ = round((d["b_seats"]/d["t_seats"])*100,2) if d["t_seats"] else 0
        ws_th.append([v,d["shows"],d["t_seats"],d["b_seats"],d["p_gross"],d["b_gross"],occ])

    # City Wise
    ws_city = wb.create_sheet(title="City Wise")
    ws_city.append(["City","Theatres","Shows","Total Seats","Booked Seats","Total Gross","Booked Gross","Occ %"])
    city_map = {}
    for r in data:
        k = r["city"]
        if k not in city_map:
            city_map[k] = {"shows":0,"t_seats":0,"b_seats":0,"p_gross":0,"b_gross":0,"venues":set()}
        d = city_map[k]
        d["shows"]+=1; d["t_seats"]+=r["total_tickets"]; d["b_seats"]+=r["booked_tickets"]
        d["p_gross"]+=r["total_gross"]; d["b_gross"]+=r["booked_gross"]; d["venues"].add(r["venue"])
    for city, d in city_map.items():
        occ = round((d["b_seats"]/d["t_seats"])*100,2) if d["t_seats"] else 0
        ws_city.append([city,len(d["venues"]),d["shows"],d["t_seats"],d["b_seats"],d["p_gross"],d["b_gross"],occ])

    # Show Wise
    ws_show = wb.create_sheet(title="Show Wise")
    ws_show.append(["Source","City","Venue","Time","SID","Total Seats","Booked Seats","Total Gross","Booked Gross","Occ %"])
    for r in data:
        ws_show.append([r["source"],r["city"],r["venue"],r["normalized_show_time"],r["sid"],
                        r["total_tickets"],r["booked_tickets"],r["total_gross"],r["booked_gross"],r["occupancy"]])

    # Summary
    ws_sum = wb.create_sheet(title="Summary")
    agg_t  = sum(r["total_tickets"]  for r in data)
    agg_b  = sum(r["booked_tickets"] for r in data)
    agg_bg = sum(r["booked_gross"]   for r in data)
    occ    = round((agg_b/agg_t)*100,2) if agg_t else 0
    ws_sum.append(["Metric","Value"])
    for row in [("Total Cities",len(city_map)),("Total Theatres",len(th_map)),
                ("Total Shows",len(data)),("Booked Gross",agg_bg),
                ("Occupancy %",occ),("Generated At",datetime.now().strftime("%Y-%m-%d %H:%M:%S"))]:
        ws_sum.append(list(row))

    path = os.path.join(reports_dir, filename)
    wb.save(path)
    print(f"📊 Excel saved: {path}")


# =============================================================================
# ── REPORT MANAGEMENT ────────────────────────────────────────────────────────
# =============================================================================

def get_report_base_name(movie_name, show_date, report_type):
    """Generate meaningful report name: MovieName_18Mar_CitiesReport"""
    slug = movie_name.replace(" ", "_")
    date_str = datetime.strptime(show_date, "%Y-%m-%d").strftime("%d%b")
    return f"{slug}_{date_str}_{report_type}Report"


def load_previous_report_data(base_name, reports_dir="reports"):
    """Load previously saved show data for merging."""
    path = os.path.join(reports_dir, f"{base_name}_data.json")
    if os.path.exists(path):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            print(f"📂 Loaded previous data: {len(data)} shows from {path}")
            return data
        except Exception as e:
            print(f"⚠️  Could not load previous data: {e}")
    return None


def save_report_data(final_data, base_name, reports_dir="reports"):
    """Save show data as JSON for future merging."""
    path = os.path.join(reports_dir, f"{base_name}_data.json")
    serializable = []
    for show in final_data:
        s = {}
        for k, v in show.items():
            s[k] = list(v) if isinstance(v, set) else v
        serializable.append(s)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(serializable, f, ensure_ascii=False, indent=2)
    print(f"💾 Data saved: {path}")


def merge_with_previous_data(new_data, old_data):
    """Merge current run with previous saved data.
    - Matches by (city, venue, normalized_show_time)
    - Existing shows: updated with new values
    - New shows: added
    - Old-only shows (e.g. morning): preserved
    """
    if not old_data:
        return new_data

    new_by_key = {}
    for show in new_data:
        key = (show["city"], show["venue"], show["normalized_show_time"])
        new_by_key[key] = show

    merged = list(new_data)
    preserved = 0
    for show in old_data:
        key = (show["city"], show["venue"], show["normalized_show_time"])
        if key not in new_by_key:
            merged.append(show)
            preserved += 1

    if preserved:
        print(f"   📎 Preserved {preserved} shows from previous run")
    return merged


def archive_previous_reports(base_name, reports_dir="reports", old_reports_dir="old_reports"):
    """Move existing aggregated reports to old_reports/ before overriding."""
    os.makedirs(old_reports_dir, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    moved = 0
    for ext in ['.xlsx', '.png', '.html', '_data.json']:
        src = os.path.join(reports_dir, f"{base_name}{ext}")
        if os.path.exists(src):
            dest = os.path.join(old_reports_dir, f"{base_name}_{ts}{ext}")
            shutil.move(src, dest)
            moved += 1
    if moved:
        print(f"   📁 Archived {moved} previous report files to {old_reports_dir}/")


# =============================================================================
# ── MAIN ──────────────────────────────────────────────────────────────────────
# =============================================================================

if __name__ == "__main__":
    # Build city pairs — city name derived from District slug
    district_pairs = [
        (slug.replace("-", " ").title(), slug)
        for slug in DISTRICT_CITIES
    ]
    bms_pairs = [
        (d_slug.replace("-", " ").title(), b_slug)
        for d_slug, b_slug in zip(DISTRICT_CITIES, BMS_CITIES)
    ]

    # Clear global SIDs sets for fresh run
    with _global_district_sids_lock:
        _global_district_sids.clear()
    with _global_bms_sids_lock:
        _global_bms_sids.clear()

    total = len(district_pairs)
    print(f"🎬 Starting run — {total} cities")
    print(f"   District + BMS run IN PARALLEL (2 threads)")
    print(f"   Each city processed sequentially, {VENUE_WORKERS} venue workers per city\n")

    all_dist_data = []
    all_bms_data  = []

    # District and BMS run fully in parallel — each is one thread
    with ThreadPoolExecutor(max_workers=2) as pool:
        district_future = pool.submit(run_district, district_pairs)
        bms_future      = pool.submit(run_bms,      bms_pairs)

        all_dist_data = district_future.result()
        all_bms_data  = bms_future.result()

    print(f"\n📋 Both sources done.")
    print(f"   District: {len(all_dist_data)} shows")
    print(f"   BMS:      {len(all_bms_data)} shows")

    # Load venue mapping and merge
    load_venue_mapping()
    final_data = merge_data(all_dist_data, all_bms_data)

    # Reports
    if final_data:
        movie_name      = extract_movie_name_from_url(DISTRICT_URL_TEMPLATE.replace("{city}", DISTRICT_CITIES[0]))
        show_date_fmt   = datetime.strptime(SHOW_DATE, "%Y-%m-%d").strftime("%d %b %Y")
        base_name       = get_report_base_name(movie_name, SHOW_DATE, "Cities")
        is_show_day     = SHOW_DATE == datetime.now().strftime("%Y-%m-%d")
        current_run_data = list(final_data)  # snapshot before merging

        # Load previous data and merge (preserves shows from earlier runs)
        old_data = load_previous_report_data(base_name)
        if old_data:
            final_data = merge_with_previous_data(final_data, old_data)

        print(f"\n📦 Generating reports — {len(final_data)} shows across {total} cities...")

        # Archive previous aggregated reports before overriding
        archive_previous_reports(base_name)

        # ── Aggregated report (cumulative) in reports/ ──
        generate_excel(final_data, f"{base_name}.xlsx")
        generate_premium_city_image_report(
            final_data, f"reports/{base_name}.png",
            movie_name=movie_name, show_date=show_date_fmt,
        )
        generate_hybrid_city_html_report(
            final_data,
            DISTRICT_URL_TEMPLATE.replace("{city}", DISTRICT_CITIES[0]),
            f"reports/{base_name}.html",
        )
        save_report_data(final_data, base_name)

        # ── Current-run-only report (booking update) in old_reports/ ──
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        snapshot_name = f"{base_name}_{ts}"
        os.makedirs("old_reports", exist_ok=True)
        generate_excel(current_run_data, f"{snapshot_name}.xlsx")
        shutil.move(f"reports/{snapshot_name}.xlsx", f"old_reports/{snapshot_name}.xlsx")
        generate_premium_city_image_report(
            current_run_data, f"old_reports/{snapshot_name}.png",
            movie_name=movie_name, show_date=show_date_fmt,
        )
        generate_hybrid_city_html_report(
            current_run_data,
            DISTRICT_URL_TEMPLATE.replace("{city}", DISTRICT_CITIES[0]),
            f"old_reports/{snapshot_name}.html",
        )
        print(f"   📋 Booking update saved to old_reports/{snapshot_name}.*")

        # ── Email ──
        aggregated_files = [
            f"reports/{base_name}.xlsx",
            f"reports/{base_name}.png",
            f"reports/{base_name}.html",
        ]
        snapshot_files = [
            f"old_reports/{snapshot_name}.xlsx",
            f"old_reports/{snapshot_name}.png",
            f"old_reports/{snapshot_name}.html",
        ]

        if is_show_day and old_data:
            send_collection_report(
                report_type="cities",
                movie_name=movie_name,
                show_date=show_date_fmt,
                subject_label="Tracked Gross + Advance Sales",
                attachment_paths=aggregated_files + snapshot_files,
                sections=[
                    {
                        "label": "A. Tracked Gross + Advance Sales (Full Day)",
                        "note":  "Cumulative gross tracked so far, including advance sales for remaining shows today.",
                        "files": aggregated_files,
                    },
                    {
                        "label": "B. Advance Sales (Remaining Shows)",
                        "note":  "Current advance booking status for shows yet to begin.",
                        "files": snapshot_files,
                    },
                ],
            )
        else:
            send_collection_report(
                report_type="cities",
                movie_name=movie_name,
                show_date=show_date_fmt,
                subject_label="Advance Sales",
                attachment_paths=aggregated_files,
            )

        print(f"\n🏁 All done. Report: {base_name}")
    else:
        print("❌ No data collected.")