import json
import time
import os
import sys
import random
import threading
import queue
import requests
from base64 import b64decode
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()
from concurrent.futures import ThreadPoolExecutor, as_completed
from Crypto.Cipher import AES
from Crypto.Util.Padding import unpad
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
from fake_useragent import UserAgent
from datetime import datetime, timedelta
import difflib
from collections import defaultdict, deque

from utils.generatePremiumStatesImageReport import generate_premium_states_image_report
from utils.generateHybridStatesHTMLReport import generate_hybrid_states_html_report
from utils.sendReportEmail import send_collection_report

# =============================================================================
# ── CONFIGURATION ─────────────────────────────────────────────────────────────
# =============================================================================

INPUT_STATE_LIST = [
    'Andhra Pradesh','Karnataka','Telangana'
]

DISTRICT_CONFIG_PATH = os.path.join("utils", "district_cities_config.json")
BMS_CONFIG_PATH      = os.path.join("utils", "bms_cities_config.json")
DISTRICT_MAP_PATH    = os.path.join("utils", "district_area_city_mapping.json")
BMS_MAP_PATH         = os.path.join("utils", "bms_area_city_mapping.json")

DISTRICT_URL          = "https://www.district.in/movies/ustaad-bhagat-singh-movie-tickets-in-{city}-MV161614"
SHOW_DATE             = "2026-03-19"
DISTRICT_URL_TEMPLATE = DISTRICT_URL + "?frmtid=TVQjMJQmE&fromdate=" + SHOW_DATE
BMS_URL_TEMPLATE      = "https://in.bookmyshow.com/movies/{city}/ustaad-bhagat-singh/buytickets/ET00339939/20260319"

ENCRYPTION_KEY = "kYp3s6v9y$B&E)H+MbQeThWmZq4t7w!z"
BOOKED_STATES  = {"2"}

# ── PERFORMANCE TUNING ──
DISTRICT_CITY_WORKERS = 8     # parallel city workers for District (HTTP)
BMS_WORKERS           = 5     # parallel BMS workers (each creates fresh driver per city)
VENUE_WORKERS         = 3     # parallel venue workers per city (HTTP seat layout calls)
DISTRICT_RATE         = 5     # max requests/second to district.in (conservative to avoid 403)
BMS_RATE              = 8     # max requests/second to bookmyshow.com
BMS_RATE_LIMIT_WAIT   = 30    # seconds to wait on BMS rate limit

# Pre-generate a User-Agent for HTTP seat layout calls (avoid per-call overhead)
_BMS_HTTP_UA = UserAgent().random


# =============================================================================
# ── RATE LIMITER (thread-safe, non-blocking scheduling) ───────────────────────
# =============================================================================

class RateLimiter:
    def __init__(self, rate):
        self.min_interval = 1.0 / rate
        self._lock = threading.Lock()
        self._next_slot = 0.0

    def acquire(self):
        with self._lock:
            now = time.monotonic()
            if self._next_slot <= now:
                self._next_slot = now + self.min_interval
                return
            wait_until = self._next_slot
            self._next_slot = wait_until + self.min_interval
        sleep_time = wait_until - time.monotonic()
        if sleep_time > 0:
            time.sleep(sleep_time)


district_limiter = RateLimiter(DISTRICT_RATE)
bms_limiter      = RateLimiter(BMS_RATE)


# =============================================================================
# ── HTTP SESSION (for District — thread-local, connection-pooled) ─────────────
# =============================================================================

_thread_local = threading.local()

def get_http_session():
    if not hasattr(_thread_local, 'session'):
        s = requests.Session()
        ua = UserAgent()
        s.headers.update({
            'User-Agent': ua.random,
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.9',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Cache-Control': 'no-cache',
        })
        adapter = requests.adapters.HTTPAdapter(
            pool_connections=20,
            pool_maxsize=20,
            max_retries=requests.adapters.Retry(total=2, backoff_factor=0.5,
                                                 status_forcelist=[502, 503, 504]),
        )
        s.mount('https://', adapter)
        s.mount('http://', adapter)
        _thread_local.session = s
    return _thread_local.session


# =============================================================================
# ── SELENIUM DRIVER POOL (for BMS — reusable Chrome instances) ────────────────
# =============================================================================

def _create_chrome_driver():
    ua = UserAgent()
    options = Options()
    options.add_argument(f"user-agent={ua.random}")
    options.add_argument("--headless=new")
    options.add_argument("start-maximized")
    options.add_argument("--disable-web-security")
    options.add_argument("--disable-site-isolation-trials")
    options.add_argument("disable-csp")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option("useAutomationExtension", False)
    prefs = {
        "profile.managed_default_content_settings.images": 2,
        "profile.default_content_setting_values.notifications": 2,
    }
    options.add_experimental_option("prefs", prefs)
    driver = webdriver.Chrome(options=options)
    driver.execute_cdp_cmd("Page.addScriptToEvaluateOnNewDocument", {
        "source": "Object.defineProperty(navigator, 'webdriver', {get: () => undefined});"
    })
    return driver





# =============================================================================
# ── GLOBAL BMS SIDs SET (shared across all workers) ──────────────────────────
# =============================================================================

_global_bms_sids = set()
_global_bms_sids_lock = threading.Lock()


# =============================================================================
# ── GLOBAL DISTRICT SIDs SET (shared across all workers) ────────────────────
# =============================================================================

_global_district_sids = set()
_global_district_sids_lock = threading.Lock()


# =============================================================================
# ── HELPERS ───────────────────────────────────────────────────────────────────
# =============================================================================

def extract_movie_name_from_url(url):
    try:
        if '/movies/' in url and '/buytickets/' in url:
            parts      = url.split('/movies/')[1].split('/buytickets/')[0].split('/')
            movie_slug = parts[-1] if len(parts) > 1 else parts[0]
            return movie_slug.replace('-', ' ').title()
        if '/movies/' in url and '-movie-tickets-in-' in url:
            movie_slug = url.split('/movies/')[1].split('-movie-tickets-in-')[0]
            return movie_slug.replace('-', ' ').title()
    except Exception as e:
        print(f"Could not extract movie name from URL: {e}")
    return "Movie Collection"


def load_mapping_dict(file_path):
    mapping = {}
    if os.path.exists(file_path):
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                for state, cities in data.items():
                    if isinstance(cities, list):
                        for entry in cities:
                            if "name" in entry and "reporting_city" in entry:
                                mapping[(state, entry["name"])] = entry["reporting_city"]
        except Exception as e:
            print(f"Warning: Could not parse mapping {file_path}: {e}")
    return mapping

DISTRICT_CITY_MAP = load_mapping_dict(DISTRICT_MAP_PATH)
BMS_CITY_MAP      = load_mapping_dict(BMS_MAP_PATH)

def get_normalized_city_name(state, raw_city, source):
    lookup = DISTRICT_CITY_MAP if source == "district" else BMS_CITY_MAP
    return lookup.get((state, raw_city), raw_city)


def district_gmt_to_ist(dt_str):
    gmt = datetime.fromisoformat(dt_str)
    ist = gmt + timedelta(hours=5, minutes=30)
    return ist.strftime("%Y-%m-%d %H:%M")

def normalize_bms_time(show_date, show_time):
    dt = datetime.strptime(f"{show_date} {show_time}", "%Y-%m-%d %I:%M %p")
    return dt.strftime("%Y-%m-%d %H:%M")

def build_seat_signature(seat_map):
    return "|".join(str(c) for c in sorted(seat_map.values()))


# =============================================================================
# ── DISTRICT (HTTP-based, no Selenium) ────────────────────────────────────────
# =============================================================================

def get_district_seat_layout(cinema_id, session_id):
    """Direct HTTP POST for District seat layout API."""
    api_url = "https://www.district.in/gw/consumer/movies/v1/select-seat"
    params  = {
        "version": "3", "site_id": "1", "channel": "mweb",
        "child_site_id": "1", "platform": "district",
    }
    payload = {"cinemaId": int(cinema_id), "sessionId": str(session_id)}
    headers = {
        "Content-Type": "application/json",
        "x-guest-token": str(random.randint(1, 9999999999)),
        "Origin": "https://www.district.in",
        "Referer": "https://www.district.in/",
    }
    try:
        district_limiter.acquire()
        session = get_http_session()
        resp = session.post(api_url, params=params, json=payload,
                            headers=headers, timeout=10)
        if resp.status_code == 200:
            return resp.json()
    except Exception:
        pass
    return None


def process_district_venue(cin, state, city_name, reporting_city):
    """Processes all shows for a single District venue using HTTP.
    Uses global _global_district_sids set to skip already-processed SIDs across all workers.
    """
    results = []
    # venue = cin['entityName']
    venue = cin['cinemaInfo']['name']
    for s in cin.get('sessions', []):
        sid = str(s.get('sid', ''))
        cid = s.get('cid')

        # Check if SID already processed (thread-safe global check)
        with _global_district_sids_lock:
            if sid in _global_district_sids:
                continue
            _global_district_sids.add(sid)

        price_map     = {}
        code_to_label = {}
        for area in s.get('areas', []):
            price_map[area['code']]     = float(area['price'])
            code_to_label[area['code']] = area['label']

        b_gross, p_gross, b_tkts, t_tkts = 0, 0, 0, 0
        seat_map       = defaultdict(int)
        price_seat_map = defaultdict(int)
        layout_res     = None

        if cid:
            layout_res = get_district_seat_layout(cid, sid)

        if layout_res and 'seatLayout' in layout_res:
            for area in layout_res['seatLayout'].get('colAreas', {}).get('objArea', []):
                area_code = area.get('AreaCode')
                price     = area.get('AreaPrice', price_map.get(area_code, 0))
                label     = code_to_label.get(area_code, area_code)
                for row in area.get('objRow', []):
                    for seat in row.get('objSeat', []):
                        status = seat.get('SeatStatus')
                        t_tkts += 1; p_gross += price
                        seat_map[label] += 1
                        price_seat_map[float(price)] += 1
                        if status != '0' and status != 0:
                            b_tkts += 1; b_gross += price
        else:
            for a in s.get('areas', []):
                tot, av, pr = a['sTotal'], a['sAvail'], a['price']
                bk = tot - av
                seat_map[a['label']]      = tot
                b_tkts += bk; t_tkts += tot
                b_gross += bk * pr; p_gross += tot * pr
                price_seat_map[float(pr)] += tot

        price_seat_list = sorted(price_seat_map.items())
        occ             = round((b_tkts / t_tkts) * 100, 2) if t_tkts else 0
        normalized_time = district_gmt_to_ist(s['showTime'])

        results.append({
            "source":               "district",
            "sid":                  sid,
            "state":                state,
            "city":                 reporting_city,
            "venue":                venue,
            "cinema_id":            str(cid) if cid else "",
            "showTime":             s['showTime'],
            "normalized_show_time": normalized_time,
            "seat_category_map":    dict(seat_map),
            "price_seat_map":       dict(price_seat_map),
            "price_seat_signature": price_seat_list,
            "seat_signature":       build_seat_signature(seat_map),
            "total_tickets":        abs(t_tkts),
            "booked_tickets":       min(abs(b_tkts), abs(t_tkts)),
            "total_gross":          abs(p_gross),
            "booked_gross":         min(abs(int(b_gross)), abs(int(p_gross))),
            "occupancy":            min(100, abs(occ)),
            "is_fallback":          False,
        })
    return results


def fetch_district_city(state, city, city_counter_str):
    """
    Fetches all District data for one city via HTTP.
    Step 1: HTTP GET page to extract __NEXT_DATA__ JSON.
    Step 2: Process each venue's shows (seat layout via HTTP POST).
    """
    city_name      = city['name']
    slug           = city.get('slug')
    reporting_city = get_normalized_city_name(state, city_name, "district")

    # Skip cities with null/empty slug
    if not slug:
        print(f"   ⚠️  [District] {city_counter_str} {city_name:<15} — skipped (no slug)")
        return []

    url = DISTRICT_URL_TEMPLATE.format(city=slug)

    # Step 1: Fetch page HTML via HTTP (with 403 retry)
    cinemas = []
    for attempt in range(2):
        try:
            district_limiter.acquire()
            session = get_http_session()
            resp    = session.get(url, timeout=15)

            if resp.status_code == 403 and attempt == 0:
                # Rate limited — wait and retry with fresh session (new User-Agent)
                time.sleep(15)
                if hasattr(_thread_local, 'session'):
                    delattr(_thread_local, 'session')
                continue

            if resp.status_code != 200:
                print(f"   ⚠️  [District] {city_counter_str} {city_name:<15} — HTTP {resp.status_code}")
                return []
            html = resp.text

            marker = 'id="__NEXT_DATA__"'
            idx    = html.find(marker)
            if idx == -1:
                return []

            start    = html.find('>', idx) + 1
            end      = html.find('</script>', start)
            data     = json.loads(html[start:end])
            sessions = data['props']['pageProps']['data']['serverState']['movieSessions']
            if not sessions:
                return []  # Movie not showing in this city
            key      = list(sessions.keys())[0]

            # Use arrangedSessions (same as city reporter)
            # cinemas = sessions[key].get('arrangedSessions', [])
            cinemas = sessions[key]['pageData']['nearbyCinemas']
            break  # Success
        except Exception as e:
            print(f"   ❌ [District] {city_counter_str} {city_name:<15} — Error: {e}")
            return []

    if not cinemas:
        return []

    # Step 2: Process each venue sequentially (HTTP calls are fast, no driver overhead)
    city_results   = []
    for cin in cinemas:
        results = process_district_venue(cin, state, city_name, reporting_city)
        city_results.extend(results)

    gross = sum(r['booked_gross'] for r in city_results)
    if city_results:
        print(f"   ✅ [District] {city_counter_str} {city_name:<15} → {reporting_city:<15} | Shows: {len(city_results):<3} | Gross: ₹{gross:<10,}")
    return city_results


def run_district(all_cities):
    """
    Main District runner. Processes cities in PARALLEL with DISTRICT_CITY_WORKERS threads.
    Rate limiter controls request frequency globally.
    """
    all_results = []
    total       = len(all_cities)
    completed   = [0]
    lock        = threading.Lock()
    print(f"\n🚀 [District] Starting — {total} cities, {DISTRICT_CITY_WORKERS} parallel workers, {DISTRICT_RATE} req/sec\n")

    def _wrapped(state, city, idx, total):
        counter_str = f"[{idx}/{total}]"
        results = fetch_district_city(state, city, counter_str)
        with lock:
            completed[0] += 1
            if completed[0] % 50 == 0:
                print(f"   📊 [District] Progress: {completed[0]}/{total} cities done")
        return results

    with ThreadPoolExecutor(max_workers=DISTRICT_CITY_WORKERS) as executor:
        futures = {
            executor.submit(_wrapped, state, city, idx, total): city['name']
            for idx, (state, city) in enumerate(all_cities, 1)
        }
        for future in as_completed(futures):
            try:
                all_results.extend(future.result())
            except Exception as e:
                print(f"❌ [District] City worker error: {e}")

    print(f"\n✅ [District] Done — {len(all_results)} total shows across {total} cities.")
    return all_results


# =============================================================================
# ── BMS (Selenium driver pool — Cloudflare requires real browser) ─────────────
# =============================================================================

def extract_initial_state_from_page(driver, url):
    """Load BMS page in Selenium and extract window.__INITIAL_STATE__ JSON."""
    try:
        driver.get(url)
        time.sleep(1)
        html   = driver.page_source
        marker = "window.__INITIAL_STATE__"
        start  = html.find(marker)
        if start == -1:
            # Retry once with extra wait for slower pages
            time.sleep(1)
            html  = driver.page_source
            start = html.find(marker)
            if start == -1:
                return None
        start = html.find("{", start)
        brace_count = 0; end = start
        while end < len(html):
            if html[end] == "{":   brace_count += 1
            elif html[end] == "}": brace_count -= 1
            if brace_count == 0:   break
            end += 1
        return json.loads(html[start:end + 1])
    except Exception:
        return None

def extract_venues(state):
    if not state:
        return []
    try:
        sbe       = state.get("showtimesByEvent")
        if not sbe: return []
        date_code = sbe.get("currentDateCode")
        if not date_code: return []
        widgets   = sbe["showDates"][date_code]["dynamic"]["data"]["showtimeWidgets"]
        for w in widgets:
            if w.get("type") == "groupList":
                for g in w["data"]:
                    if g.get("type") == "venueGroup":
                        return g["data"]
    except Exception:
        pass
    return []

def get_seat_layout(driver, venue_code, session_id):
    """Selenium XHR for BMS seat layout API (bypasses Cloudflare)."""
    api_url = "https://services-in.bookmyshow.com/doTrans.aspx"
    js = """
        var cb = arguments[0]; var x = new XMLHttpRequest();
        x.open("POST", "%s", true);
        x.setRequestHeader("Content-Type", "application/x-www-form-urlencoded");
        x.onload = function() { cb(x.responseText); };
        x.onerror = function() { cb(null); };
        x.send("strCommand=GETSEATLAYOUT&strAppCode=WEB&strVenueCode=%s&lngTransactionIdentifier=0&strParam1=%s&strParam2=WEB&strParam5=Y&strFormat=json");
    """ % (api_url, venue_code, session_id)

    max_retries = 2
    for i in range(max_retries + 1):
        try:
            bms_limiter.acquire()
            driver.set_script_timeout(20)
            resp = driver.execute_async_script(js)
            if not resp:
                return None, "Empty response"
            data = json.loads(resp).get("BookMyShow", {})
            if data.get("blnSuccess") == "true":
                return data.get("strData"), None
            error_msg = data.get("strException", "")
            # Retry on transient errors
            if any(kw in error_msg.lower() for kw in ["rate limit", "connectivity issue", "high demand"]):
                if i < max_retries:
                    wait = BMS_RATE_LIMIT_WAIT if "rate limit" in error_msg.lower() else 5
                    time.sleep(wait)
                    continue
            return None, error_msg
        except Exception as e:
            err_line = str(e).split('\n')[0]  # First line only, skip stacktrace
            if i < max_retries and "timeout" in err_line.lower():
                time.sleep(2)
                continue  # Retry on script timeout
            return None, err_line
    return None, "Unknown Error"


def get_seat_layout_http(venue_code, session_id):
    """HTTP-based seat layout call — no Selenium driver needed.
    Much faster than Selenium XHR and allows parallel venue processing.
    Uses thread-local session for HTTP keep-alive connection pooling."""
    api_url = "https://services-in.bookmyshow.com/doTrans.aspx"
    payload = (
        f"strCommand=GETSEATLAYOUT&strAppCode=WEB&strVenueCode={venue_code}"
        f"&lngTransactionIdentifier=0&strParam1={session_id}"
        f"&strParam2=WEB&strParam5=Y&strFormat=json"
    )

    # Thread-local session for connection pooling (keep-alive reuse)
    if not hasattr(_thread_local, 'bms_session'):
        s = requests.Session()
        s.headers.update({
            "Content-Type": "application/x-www-form-urlencoded",
            "User-Agent": _BMS_HTTP_UA,
            "Origin": "https://in.bookmyshow.com",
            "Referer": "https://in.bookmyshow.com/",
        })
        _thread_local.bms_session = s

    session = _thread_local.bms_session
    max_retries = 2
    for i in range(max_retries + 1):
        try:
            bms_limiter.acquire()
            resp = session.post(api_url, data=payload, timeout=15)
            if resp.status_code != 200:
                return None, f"HTTP {resp.status_code}"
            data = resp.json().get("BookMyShow", {})
            if data.get("blnSuccess") == "true":
                return data.get("strData"), None
            error_msg = data.get("strException", "")
            if any(kw in error_msg.lower() for kw in ["rate limit", "connectivity issue", "high demand"]):
                if i < max_retries:
                    wait = BMS_RATE_LIMIT_WAIT if "rate limit" in error_msg.lower() else 5
                    time.sleep(wait)
                    continue
            return None, error_msg
        except Exception as e:
            if i < max_retries and "timeout" in str(e).lower():
                time.sleep(2)
                continue
            return None, str(e).split('\n')[0]
    return None, "Unknown Error"


# Global: detect whether HTTP seat layout works (tested once, used for all cities)
_bms_http_tested = False
_bms_http_works  = False
_bms_http_lock   = threading.Lock()


def decrypt_data(enc):
    decoded = b64decode(enc)
    cipher  = AES.new(ENCRYPTION_KEY.encode(), AES.MODE_CBC, iv=bytes(16))
    return unpad(cipher.decrypt(decoded), AES.block_size).decode()

def calculate_show_collection(decrypted, price_map):
    header, rows_part = decrypted.split("||")
    rows              = rows_part.split("|")

    cat_map         = {}
    local_price_map = price_map.copy()
    last_price      = 0.0

    for p in header.split("|"):
        parts = p.split(":")
        if len(parts) >= 3:
            cat_map[parts[1]] = parts[2]
            current_price = local_price_map.get(parts[2], 0.0)
            if current_price > 0:  last_price = current_price
            elif last_price > 0:   local_price_map[parts[2]] = last_price

    seats, booked = {}, {}
    for row in rows:
        if not row: continue
        parts = row.split(":")
        if len(parts) < 3: continue
        block = parts[3][0] if len(parts) > 3 else parts[2][0]
        area  = cat_map.get(block)
        if not area: continue
        for seat in parts:
            if len(seat) < 2: continue
            status = seat[1]
            if seat[0] == block and status in ("1", "2"):
                seats[area] = seats.get(area, 0) + 1
            if seat[0] == block and status in BOOKED_STATES:
                booked[area] = booked.get(area, 0) + 1

    t_tkts, b_tkts, t_gross, b_gross = 0, 0, 0, 0
    for area, total in seats.items():
        bk = booked.get(area, 0); pr = local_price_map.get(area, 0)
        t_tkts += total; b_tkts += bk
        t_gross += total * pr; b_gross += bk * pr

    occ = round((b_tkts / t_tkts) * 100, 2) if t_tkts else 0
    return t_tkts, b_tkts, int(t_gross), int(b_gross), occ, seats, local_price_map


def process_bms_venue(driver, venue, city_name, reporting_city, state_name, use_http=False):
    """
    Processes ONE BMS venue. Uses HTTP seat layout if use_http=True (faster, parallelizable),
    otherwise uses Selenium XHR via the provided driver.
    Preserves all business logic: sold-out recovery, screen caching, deferred SIDs.
    Uses global _global_bms_sids set to skip already-processed SIDs across all workers.
    """
    results            = []
    screen_details_map = {}

    try:
        v_name = venue["additionalData"]["venueName"]
        v_code = venue["additionalData"]["venueCode"]

        shows      = venue.get("showtimes", [])
        shows.sort(key=lambda s: s["additionalData"].get("availStatus", "0"), reverse=True)
        show_queue    = deque(shows)
        deferred_sids = set()

        while show_queue:
            show      = show_queue.popleft()
            sid       = str(show["additionalData"]["sessionId"])
            show_time = show["title"]

            raw_screen = show.get("screenAttr", "")
            screenName = raw_screen if raw_screen else "Main Screen"

            # Check if SID already processed (thread-safe global check)
            with _global_bms_sids_lock:
                if sid in _global_bms_sids:
                    continue
                _global_bms_sids.add(sid)

            soldOut        = False
            seat_map       = {}
            is_fallback    = False
            price_seat_map = {}

            try:
                cats      = show["additionalData"].get("categories", [])
                price_map = {c["areaCatCode"]: float(c["curPrice"]) for c in cats}
                enc, error_msg = get_seat_layout_http(v_code, sid) if use_http else get_seat_layout(driver, v_code, sid)
                data           = None

                if not enc:
                    if not price_map: continue
                    max_price   = max(price_map.values())
                    is_fallback = True
                    for p in price_map.values():
                        price_seat_map[float(p)] = 0

                    if error_msg and "sold out" in error_msg.lower():
                        # print(f"      🔴 [BMS][{city_name}] Sold Out: {sid}. Checking recovery...")
                        recovered_capacity = None
                        recovered_seat_map = None

                        if screenName in screen_details_map:
                            recovered_seat_map = screen_details_map[screenName]
                            recovered_capacity = sum(recovered_seat_map.values())
                            # print(f"         ⚡ Using cached layout ({recovered_capacity} seats)")

                        if not recovered_capacity:
                            try:
                                base_sid = int(sid)
                                for offset in range(7, 0, -1):
                                    target_sid = str(base_sid + offset)
                                    n_enc, _ = get_seat_layout_http(v_code, target_sid) if use_http else get_seat_layout(driver, v_code, target_sid)
                                    if n_enc:
                                        n_dec = decrypt_data(n_enc)
                                        n_res = calculate_show_collection(n_dec, {})
                                        if n_res[0] > 0:
                                            recovered_capacity = n_res[0]
                                            recovered_seat_map = n_res[5]
                                            # print(f"         ✨ Recovered using {target_sid}")
                                            break
                            except Exception:
                                pass

                        if recovered_capacity:
                            calc_gross = sum(count * price_map.get(ac, 0) for ac, count in recovered_seat_map.items())
                            if calc_gross > 0:
                                t_tkts = b_tkts = recovered_capacity
                                t_gross = b_gross = calc_gross
                                screen_details_map[screenName] = recovered_seat_map
                                seat_map    = recovered_seat_map
                                is_fallback = False
                                ps_map      = defaultdict(int)
                                for ac, count in seat_map.items():
                                    ps_map[float(price_map.get(ac, 0))] += count
                                price_seat_map = dict(ps_map)
                            else:
                                recovered_capacity = None

                        if not recovered_capacity:
                            FALLBACK_SEATS = 400
                            t_tkts  = b_tkts  = FALLBACK_SEATS
                            t_gross = b_gross = int(FALLBACK_SEATS * max_price)
                            # print(f"         ❌ [BMS][{city_name}] Recovery failed. Using default {FALLBACK_SEATS}.")

                        occ     = 100.0
                        soldOut = True
                        data    = {"total_tickets": t_tkts, "booked_tickets": b_tkts,
                                   "total_gross": t_gross, "booked_gross": b_gross, "occupancy": occ}

                    elif error_msg and "Rate limit" in error_msg:
                        print(f"      🚫 [BMS][{city_name}] Rate Limit for {v_name[:15]}")
                        continue

                    else:
                        print(f"      ⚠️  [BMS][{city_name}] Error for {sid}: {error_msg}")
                        if screenName in screen_details_map:
                            cached = screen_details_map[screenName]
                            seat_map     = cached
                            t_tkts       = sum(cached.values())
                            b_tkts       = int(t_tkts * 0.5)
                            ps_map       = defaultdict(int); t_gross_calc = 0
                            for ac, count in cached.items():
                                pr = float(price_map.get(ac, 0))
                                ps_map[pr] += count; t_gross_calc += count * pr
                            price_seat_map = dict(ps_map)
                            t_gross = int(t_gross_calc); b_gross = int(t_gross * 0.5)
                            occ = 50.0; is_fallback = False
                            # print(f"         ⚡ Smart Fallback: {screenName} ({t_tkts} seats)")
                        elif sid not in deferred_sids and len(show_queue) > 0:
                            deferred_sids.add(sid)
                            with _global_bms_sids_lock:
                                _global_bms_sids.discard(sid)
                            show_queue.append(show)
                            continue
                        else:
                            t_tkts  = 400; b_tkts  = 200
                            t_gross = int(400 * max_price); b_gross = int(200 * max_price)
                            occ     = 50.0
                            # print(f"         ❌ [BMS][{city_name}] Hard Fallback 400/200")
                        data = {"total_tickets": t_tkts, "booked_tickets": b_tkts,
                                "total_gross": t_gross, "booked_gross": b_gross, "occupancy": occ}
                else:
                    decrypted = decrypt_data(enc)
                    res       = calculate_show_collection(decrypted, price_map)
                    data      = {
                        "total_tickets":  abs(res[0]),
                        "booked_tickets": min(abs(res[1]), abs(res[0])),
                        "total_gross":    abs(res[2]),
                        "booked_gross":   min(abs(res[3]), abs(res[2])),
                        "occupancy":      min(100, abs(res[4])),
                    }
                    seat_map        = res[5]
                    final_price_map = res[6]

                    if data["total_tickets"] > 0:
                        ps_map  = defaultdict(int); ps_list = []
                        for ac, count in seat_map.items():
                            pr = float(final_price_map.get(ac, 0))
                            ps_map[pr] += count; ps_list.append((pr, count))
                        price_seat_map             = dict(ps_map)
                        data["price_seat_signature"] = sorted(ps_list)
                        screen_details_map[screenName] = seat_map

                if data and data['total_tickets'] > 0:
                    normalized_time = normalize_bms_time(SHOW_DATE, show_time)
                    tag = "(SOLD OUT)" if soldOut else ""
                    # print(
                    #     f"   🎬 [BMS][{city_name}] {v_name[:15]:<15} | {normalized_time} | "
                    #     f"Occ: {data['occupancy']:>5}% | Gross: ₹{data['booked_gross']:<8,} {tag}"
                    # )
                    data.update({
                        "source":               "bms",
                        "sid":                  sid,
                        "state":                state_name,
                        "city":                 reporting_city,
                        "venue":                v_name,
                        "venue_code":           v_code,
                        "showTime":             show_time,
                        "normalized_show_time": normalized_time,
                        "seat_category_map":    seat_map,
                        "price_seat_map":       price_seat_map,
                        "price_seat_signature": data.get("price_seat_signature", []),
                        "seat_signature":       build_seat_signature(seat_map),
                        "is_fallback":          is_fallback,
                    })
                    results.append(data)

            except Exception:
                continue

    except Exception as e:
        print(f"❌ [BMS] Venue worker error for {city_name}: {e}")

    return results


def fetch_bms_city(state_name, city_name, city_slug, city_counter_str):
    """
    Fetches all BMS data for one city.
    Creates a FRESH Chrome driver per city (Cloudflare blocks reused sessions).
    If HTTP seat layout API works, quits driver early and processes venues in parallel.
    """
    global _bms_http_tested, _bms_http_works

    reporting_city = get_normalized_city_name(state_name, city_name, "bms")
    url            = BMS_URL_TEMPLATE.format(city=city_slug)
    driver         = None

    try:
        bms_limiter.acquire()
        driver = _create_chrome_driver()
        driver.set_script_timeout(20)

        state_data = extract_initial_state_from_page(driver, url)
        venues     = extract_venues(state_data) if state_data else []

        if not venues:
            return []

        # Probe HTTP mode once (first city with venues triggers the test)
        with _bms_http_lock:
            if not _bms_http_tested:
                for v in venues:
                    shows = v.get("showtimes", [])
                    if not shows:
                        continue
                    test_vcode = v["additionalData"]["venueCode"]
                    test_sid   = str(shows[0]["additionalData"]["sessionId"])
                    enc, err   = get_seat_layout_http(test_vcode, test_sid)
                    _bms_http_works = enc is not None or (err and "sold out" in err.lower())
                    break
                _bms_http_tested = True
                mode = "parallel HTTP" if _bms_http_works else "Selenium"
                print(f"   🔍 [BMS] Seat layout mode: {mode}")

        if _bms_http_works:
            # HTTP works — quit driver early and process venues in parallel
            driver.quit()
            driver = None

            city_results = []
            with ThreadPoolExecutor(max_workers=VENUE_WORKERS) as venue_pool:
                futures = [
                    venue_pool.submit(
                        process_bms_venue, None, v, city_name,
                        reporting_city, state_name, True
                    )
                    for v in venues
                ]
                for f in as_completed(futures):
                    try:
                        city_results.extend(f.result())
                    except Exception:
                        pass
        else:
            # HTTP unavailable — sequential Selenium mode (same driver)
            city_results   = []
            for venue in venues:
                results = process_bms_venue(driver, venue, city_name, reporting_city, state_name)
                city_results.extend(results)

        gross = sum(r['booked_gross'] for r in city_results)
        if city_results:
            print(f"   ✅ [BMS] {city_counter_str} {city_name:<15} → {reporting_city:<15} | Shows: {len(city_results):<3} | Gross: ₹{gross:<10,}")
        return city_results

    except Exception as e:
        print(f"   ❌ [BMS] {city_counter_str} {city_name:<15} — Error: {e}")
        return []
    finally:
        if driver:
            try:
                driver.quit()
            except Exception:
                pass


def run_bms(all_cities):
    """
    Main BMS runner. Fresh Chrome driver created per city (Cloudflare blocks reused sessions).
    Cities processed in parallel by BMS_WORKERS threads.
    """
    all_results = []
    total       = len(all_cities)
    completed   = [0]
    lock        = threading.Lock()

    print(f"\n🚀 [BMS] Starting — {total} cities, {BMS_WORKERS} parallel workers, {BMS_RATE} req/sec")
    print(f"   Fresh Chrome driver per city | {VENUE_WORKERS} venue workers per city\n")

    def _process(args):
        idx, (state, city_name, city_slug) = args
        counter_str = f"[{idx}/{total}]"
        results = fetch_bms_city(state, city_name, city_slug, counter_str)
        with lock:
            completed[0] += 1
            if completed[0] % 50 == 0:
                print(f"   📊 [BMS] Progress: {completed[0]}/{total} cities done")
        return results

    try:
        with ThreadPoolExecutor(max_workers=BMS_WORKERS) as executor:
            futures = {
                executor.submit(_process, (i + 1, ct)): ct[1]
                for i, ct in enumerate(all_cities)
            }
            for future in as_completed(futures):
                try:
                    all_results.extend(future.result())
                except Exception as e:
                    print(f"❌ [BMS] Worker error: {e}")
    except Exception as e:
        print(f"❌ [BMS] Fatal error: {e}")

    print(f"\n✅ [BMS] Done — {len(all_results)} total shows across {total} cities.")
    return all_results


# =============================================================================
# ── VENUE MAPPING ─────────────────────────────────────────────────────────────
# =============================================================================

VENUE_MAP = {}   # BMS VenueCode → District cinema_id

def load_venue_mapping():
    global VENUE_MAP
    mapping_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'utils', 'venue_mapping.json')
    if os.path.exists(mapping_path):
        with open(mapping_path, encoding='utf-8') as f:
            data = json.load(f)
        VENUE_MAP = data.get('bms_to_district', {})
        print(f"📍 Loaded venue mapping: {len(VENUE_MAP)} BMS→District pairs")
    else:
        print("⚠️  Venue mapping not found (utils/venue_mapping.json). Venue-based matching disabled.")

def _is_same_venue(bms_show, dist_show):
    """Check venue mapping. Returns True/False if mapping exists, 'unmapped' if BMS venue not in map."""
    bms_code = bms_show.get('venue_code', '')
    dist_cid = dist_show.get('cinema_id', '')
    if not bms_code or not dist_cid:
        return 'unmapped'
    mapped_dist = VENUE_MAP.get(bms_code)
    if mapped_dist is None:
        return 'unmapped'
    return mapped_dist == dist_cid

# =============================================================================
# ── DEDUP + MERGE ─────────────────────────────────────────────────────────────
# =============================================================================

def dedup_same_platform(records, source_label):
    seen    = {}
    best    = []
    dropped = 0
    for r in records:
        sid = r.get('sid', '')
        if not sid:
            best.append(r); continue
        if sid not in seen:
            seen[sid] = len(best)
            best.append(r)
        else:
            dropped += 1
            existing = best[seen[sid]]
            if r.get('booked_gross', 0) > existing.get('booked_gross', 0):
                best[seen[sid]] = r
                print(f"   ♻️  [{source_label}] Replaced lower-gross duplicate SID {sid}")
            else:
                print(f"   ♻️  [{source_label}] Dropped duplicate SID {sid}")
    if dropped:
        print(f"   ✅ [{source_label}] Dedup removed {dropped} duplicate(s). {len(best)} unique shows remain.")
    return best


def merge_data(all_dist_data, all_bms_data):
    all_dist_data = dedup_same_platform(all_dist_data, "District")
    all_bms_data  = dedup_same_platform(all_bms_data,  "BMS")

    print(f"\n🔄 Merging {len(all_dist_data)} District + {len(all_bms_data)} BMS shows...")

    final_data     = []
    SEAT_TOLERANCE = 5

    district_index = defaultdict(list)
    for r in all_dist_data:
        district_index[(r['state'], r['city'], r['normalized_show_time'])].append(r)

    for bms in all_bms_data:
        key        = (bms['state'], bms['city'], bms['normalized_show_time'])
        candidates = district_index.get(key, [])
        match      = None

        # 1. Exact SID
        for c in candidates:
            if c['sid'] == bms['sid']:
                match = c; print(f"   🔗 SID Match: {bms['sid']}"); break

        # 2. Price + Seat signature + Venue Map
        if not match and not bms.get('is_fallback', False):
            b_sig = bms.get('price_seat_signature', [])
            for c in candidates:
                d_sig = c.get('price_seat_signature', [])
                if not b_sig or not d_sig or len(b_sig) != len(d_sig): continue
                if all(bp == dp and abs(bs - ds) <= SEAT_TOLERANCE
                       for (bp, bs), (dp, ds) in zip(b_sig, d_sig)):
                    if _is_same_venue(bms, c) is True:
                        match = c
                        print(f"   🔗 Price/Seat Sig + Venue Map: {bms['venue']} == {c['venue']}")
                        break

        # 3. Seat signature + Venue Map
        if not match and not bms.get('is_fallback', False):
            b_seats = sorted(bms.get('seat_category_map', {}).values())
            for c in candidates:
                d_seats = sorted(c.get('seat_category_map', {}).values())
                if not b_seats or not d_seats or len(b_seats) != len(d_seats): continue
                if all(abs(bs - ds) <= SEAT_TOLERANCE for bs, ds in zip(b_seats, d_seats)):
                    if _is_same_venue(bms, c) is True:
                        match = c
                        print(f"   🔗 Seat Sig + Venue Map: {bms['venue']} == {c['venue']}")
                        break

        # 4. Venue map + strict price set
        if not match and candidates:
            b_prices = {p for p in bms.get('price_seat_map', {}).keys() if p > 0}
            for c in candidates:
                d_prices = {p for p in c.get('price_seat_map', {}).keys() if p > 0}
                if b_prices != d_prices: continue
                if _is_same_venue(bms, c) is True:
                    match = c
                    print(f"   🔗 Venue Map + Price: {bms['venue']} == {c['venue']}")
                    break

        # 5. Venue map only (seats/prices may differ between platforms)
        if not match and candidates:
            for c in candidates:
                if _is_same_venue(bms, c) is True:
                    match = c
                    print(f"   🔗 Venue Map Only: {bms['venue']} == {c['venue']}")
                    break

        if match:
            candidates.remove(match)
            if not bms.get('is_fallback', False) and bms['booked_gross'] > match['booked_gross']:
                match.update({
                    'total_tickets':     bms['total_tickets'],
                    'booked_tickets':    bms['booked_tickets'],
                    'total_gross':       bms['total_gross'],
                    'booked_gross':      bms['booked_gross'],
                    'occupancy':         bms['occupancy'],
                    'seat_category_map': bms['seat_category_map'],
                    'price_seat_map':    bms['price_seat_map'],
                    'seat_signature':    bms['seat_signature'],
                })
            final_data.append(match)
        else:
            final_data.append(bms)

    for sublist in district_index.values():
        final_data.extend(sublist)

    print(f"✅ Merge complete — {len(final_data)} final shows.")
    return final_data


# =============================================================================
# ── EXCEL GENERATOR ───────────────────────────────────────────────────────────
# =============================================================================

def generate_consolidated_excel(all_results, filename):
    print("\nGenerating Consolidated Excel Report...")
    wb          = Workbook()
    reports_dir = "reports"
    os.makedirs(reports_dir, exist_ok=True)

    # State Wise
    ws_state = wb.active
    ws_state.title = "State Wise"
    ws_state.append(["State","Cities","Theatres","Shows","Total Seats","Booked Seats","Total Gross","Booked Gross","Occ %"])
    state_map, city_tracker, theatre_tracker = {}, {}, {}
    for r in all_results:
        st = r["state"]
        if st not in state_map:
            state_map[st]    = {"shows":0,"t_seats":0,"b_seats":0,"p_gross":0,"b_gross":0}
            city_tracker[st] = set(); theatre_tracker[st] = set()
        d = state_map[st]
        d["shows"]+=1; d["t_seats"]+=r["total_tickets"]; d["b_seats"]+=r["booked_tickets"]
        d["p_gross"]+=r["total_gross"]; d["b_gross"]+=r["booked_gross"]
        city_tracker[st].add(r["city"]); theatre_tracker[st].add(r["venue"])
    for st, d in state_map.items():
        occ = round((d["b_seats"]/d["t_seats"])*100,2) if d["t_seats"] else 0
        ws_state.append([st,len(city_tracker[st]),len(theatre_tracker[st]),d["shows"],
                         d["t_seats"],d["b_seats"],d["p_gross"],d["b_gross"],occ])

    # City Wise
    ws_city = wb.create_sheet(title="City Wise")
    ws_city.append(["State","City","Theatres","Shows","Total Seats","Booked Seats","Total Gross","Booked Gross","Occ %"])
    city_map, city_theatre_tracker = {}, {}
    for r in all_results:
        k = (r["state"],r["city"])
        if k not in city_map:
            city_map[k]             = {"shows":0,"t_seats":0,"b_seats":0,"p_gross":0,"b_gross":0}
            city_theatre_tracker[k] = set()
        d = city_map[k]
        d["shows"]+=1; d["t_seats"]+=r["total_tickets"]; d["b_seats"]+=r["booked_tickets"]
        d["p_gross"]+=r["total_gross"]; d["b_gross"]+=r["booked_gross"]
        city_theatre_tracker[k].add(r["venue"])
    for (st,ct), d in city_map.items():
        occ = round((d["b_seats"]/d["t_seats"])*100,2) if d["t_seats"] else 0
        ws_city.append([st,ct,len(city_theatre_tracker[(st,ct)]),d["shows"],
                        d["t_seats"],d["b_seats"],d["p_gross"],d["b_gross"],occ])

    # Theatre Wise
    ws_th = wb.create_sheet(title="Theatre Wise")
    ws_th.append(["Source","State","City","Venue","Shows","Total Seats","Booked Seats","Total Gross","Booked Gross","Occ %"])
    th_map = {}
    for r in all_results:
        k = (r["source"],r["state"],r["city"],r["venue"])
        if k not in th_map:
            th_map[k] = {"shows":0,"t_seats":0,"b_seats":0,"p_gross":0,"b_gross":0}
        d = th_map[k]
        d["shows"]+=1; d["t_seats"]+=r["total_tickets"]; d["b_seats"]+=r["booked_tickets"]
        d["p_gross"]+=r["total_gross"]; d["b_gross"]+=r["booked_gross"]
    for (src,st,ct,vn), d in th_map.items():
        occ = round((d["b_seats"]/d["t_seats"])*100,2) if d["t_seats"] else 0
        ws_th.append([src,st,ct,vn,d["shows"],d["t_seats"],d["b_seats"],d["p_gross"],d["b_gross"],occ])

    # Show Wise
    ws_show = wb.create_sheet(title="Show Wise")
    ws_show.append(["Source","State","City","Venue","Time","SID","Total Seats","Booked Seats","Total Gross","Booked Gross","Occ %"])
    for r in all_results:
        ws_show.append([r["source"],r["state"],r["city"],r["venue"],
                        r["normalized_show_time"],r["sid"],
                        r["total_tickets"],r["booked_tickets"],
                        r["total_gross"],r["booked_gross"],r["occupancy"]])

    # Summary
    ws_sum  = wb.create_sheet(title="Summary")
    agg_t   = sum(r["total_tickets"]  for r in all_results)
    agg_b   = sum(r["booked_tickets"] for r in all_results)
    agg_bg  = sum(r["booked_gross"]   for r in all_results)
    occ     = round((agg_b/agg_t)*100,2) if agg_t else 0
    ws_sum.append(["Metric","Value"])
    for row in [("States",len(state_map)),("Cities",len(city_map)),("Shows",len(all_results)),
                ("Booked Gross",agg_bg),("Occupancy %",occ),
                ("Generated At",datetime.now().strftime("%Y-%m-%d %H:%M:%S"))]:
        ws_sum.append(list(row))

    path = os.path.join(reports_dir, filename)
    wb.save(path)
    print(f"📊 Excel saved: {path}")


# =============================================================================
# ── MAIN ──────────────────────────────────────────────────────────────────────
# =============================================================================

if __name__ == "__main__":
    # ── LOG FILE SETUP ──
    os.makedirs("logs", exist_ok=True)
    log_path = os.path.join("logs", f"run_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log")
    _log_file = open(log_path, 'w', encoding='utf-8')

    class _Tee:
        """Writes to both console and log file simultaneously."""
        def __init__(self, *streams):
            self.streams = streams
        def write(self, text):
            for s in self.streams:
                try:
                    s.write(text)
                    s.flush()
                except Exception:
                    pass
        def flush(self):
            for s in self.streams:
                try:
                    s.flush()
                except Exception:
                    pass

    sys.stdout = _Tee(sys.__stdout__, _log_file)
    sys.stderr = _Tee(sys.__stderr__, _log_file)
    print(f"📝 Logging to: {log_path}\n")

    if not os.path.exists(DISTRICT_CONFIG_PATH) or not os.path.exists(BMS_CONFIG_PATH):
        print("❌ Config files missing. Exiting.")
        exit(1)

    with open(DISTRICT_CONFIG_PATH, 'r', encoding='utf-8') as f:
        district_config = json.load(f)
    with open(BMS_CONFIG_PATH, 'r', encoding='utf-8') as f:
        bms_config = json.load(f)

    district_cities = [
        (state, city)
        for state in INPUT_STATE_LIST
        for city in district_config.get(state, [])
    ]
    bms_cities = [
        (state, city['name'], city['slug'])
        for state in INPUT_STATE_LIST
        for city in bms_config.get(state, [])
    ]

    total_d = len(district_cities)
    total_b = len(bms_cities)

    # Clear global SIDs sets for fresh run
    with _global_district_sids_lock:
        _global_district_sids.clear()
    with _global_bms_sids_lock:
        _global_bms_sids.clear()

    print(f"🎬 Starting run — District: {total_d} cities | BMS: {total_b} cities")
    print(f"   Both sources run in PARALLEL (2 threads)")
    print(f"   District: {DISTRICT_CITY_WORKERS} HTTP workers | BMS: {BMS_WORKERS} workers (fresh driver/city)")
    print(f"   Rate limits: District {DISTRICT_RATE} req/s | BMS {BMS_RATE} req/s")
    print(f"   District: pure HTTP | BMS: Selenium (Cloudflare bypass)\n")

    start_time = time.monotonic()

    all_dist_data = []
    all_bms_data  = []

    # District and BMS run fully in parallel — each is one thread
    with ThreadPoolExecutor(max_workers=2) as pool:
        district_future = pool.submit(run_district, district_cities)
        bms_future      = pool.submit(run_bms,      bms_cities)

        all_dist_data = district_future.result()
        all_bms_data  = bms_future.result()

    elapsed = time.monotonic() - start_time
    print(f"\n📋 Both sources done in {elapsed/60:.1f} minutes.")
    print(f"   District: {len(all_dist_data)} shows")
    print(f"   BMS:      {len(all_bms_data)} shows")

    # Save District-only intermediate Excel
    if all_dist_data:
        ts_mid = datetime.now().strftime("%H%M")
        generate_consolidated_excel(all_dist_data, f"District_Only_{ts_mid}.xlsx")

    # Load venue mapping and merge
    load_venue_mapping()
    final_data = merge_data(all_dist_data, all_bms_data)

    # Reports
    if final_data:
        ts            = datetime.now().strftime("%Y%m%d_%H%M%S")
        movie_name    = extract_movie_name_from_url(DISTRICT_URL_TEMPLATE.format(city="city"))
        show_date_fmt = datetime.strptime(SHOW_DATE, "%Y-%m-%d").strftime("%d %b %Y")

        generate_consolidated_excel(final_data, f"Total_States_Report_{ts}.xlsx")

        generate_premium_states_image_report(
            final_data,
            f"reports/Total_States_Report_Premium_{ts}.png",
            movie_name=movie_name,
            show_date=show_date_fmt,
        )
        generate_hybrid_states_html_report(
            final_data,
            f"reports/Total_States_Report_{ts}.html",
            movie_name=movie_name,
            show_date=show_date_fmt,
        )

        # Email reports
        send_collection_report(
            report_type="states",
            movie_name=movie_name,
            show_date=show_date_fmt,
            attachment_paths=[
                f"reports/Total_States_Report_{ts}.xlsx",
                f"reports/Total_States_Report_Premium_{ts}.png",
                f"reports/Total_States_Report_{ts}.html",
            ],
        )

        total_elapsed = time.monotonic() - start_time
        print(f"\n🏁 All done in {total_elapsed/60:.1f} minutes. Reports saved with timestamp {ts}")
    else:
        print("❌ No data found.")
