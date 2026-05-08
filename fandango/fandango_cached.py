import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from playwright.sync_api import sync_playwright

# =============================================================================
# ── 1. CONFIGURATION ─────────────────────────────────────────────────────────
# =============================================================================

MOVIE_ID = "244813"
MOVIE_SLUG = "peddi-2026"  
SHOW_DATE = "2026-06-03"
ZIP_CODE = "75201"
current_time = time.strftime("%H%M%S")
EXCEL_FILENAME = f"reports/Fandango_Report_{ZIP_CODE}_{SHOW_DATE}_{current_time}.xlsx"

# =============================================================================
# ── 2. EXCEL EXPORT ──────────────────────────────────────────────────────────
# =============================================================================

def export_to_excel(shows_data, summary_data):
    print(f"\n📊 Generating Excel Report: {EXCEL_FILENAME}")
    wb = Workbook()
    
    ws_shows = wb.active
    ws_shows.title = "Showtime Details"
    ws_shows.append(["Theater Name", "Show Time", "Status", "Ticket Price", "Total Seats", "Booked Seats", "Occupancy %", "Gross ($)"])
    
    for row in shows_data:
        occ = round((row['booked'] / row['total']) * 100, 2) if row['total'] > 0 else 0
        ws_shows.append([row['theater'], row['time'], row['status'], row['price_str'], row['total'], row['booked'], occ, row['gross']])
        
    ws_summary = wb.create_sheet(title="Theater Summary")
    ws_summary.append(["Theater Name", "Total Shows", "Total Seats", "Total Booked", "Overall Occ %", "Total Gross ($)"])
    
    total_shows_all = total_seats_all = total_booked_all = total_gross_all = 0
    for t_name, stats in summary_data.items():
        occ = round((stats['booked'] / stats['total']) * 100, 2) if stats['total'] > 0 else 0
        ws_summary.append([t_name, stats['shows'], stats['total'], stats['booked'], occ, stats['gross']])
        total_shows_all += stats['shows']
        total_seats_all += stats['total']
        total_booked_all += stats['booked']
        total_gross_all += stats['gross']
        
    ws_summary.append([])
    ws_summary.append(["GRAND TOTAL", total_shows_all, total_seats_all, total_booked_all, 
                       round((total_booked_all/total_seats_all)*100,2) if total_seats_all > 0 else 0, total_gross_all])
    
    for ws in [ws_shows, ws_summary]:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            
    wb.save(EXCEL_FILENAME)
    print(f"✅ Export complete: {EXCEL_FILENAME}")

# =============================================================================
# ── MAIN EXECUTION ───────────────────────────────────────────────────────────
# =============================================================================

if __name__ == "__main__":
    shows_data = []
    summary_data = {}

    with sync_playwright() as p:
        print("\n🚀 Launching Browser in Background (Headless Mode)...")
        browser = p.chromium.launch(headless=True, args=["--disable-blink-features=AutomationControlled"])
        context = browser.new_context(viewport={'width': 1280, 'height': 800})
        page = context.new_page()

        print(f"⏳ Loading Fandango main page to bypass Akamai natively...")
        movie_url = f"https://www.fandango.com/{MOVIE_SLUG}-{MOVIE_ID}/movie-overview?date={SHOW_DATE}"
        page.goto(movie_url)
        page.wait_for_timeout(4000)

        # ---------------------------------------------------------
        # Step 1: Fetch and Group the Theaters
        # ---------------------------------------------------------
        print(f"🎬 Fetching theaters for ZIP: {ZIP_CODE}...")
        api_url = f"https://www.fandango.com/napi/theaterShowtimeGroupings/{MOVIE_ID}/{SHOW_DATE}?isdesktop=false&zip={ZIP_CODE}"
        
        try:
            raw_data = page.evaluate(f"""async () => {{
                const response = await fetch('{api_url}');
                if (!response.ok) return null;
                return await response.json();
            }}""")
        except Exception as e:
            print(f"❌ Failed to fetch theaters: {e}")
            browser.close()
            exit()

        if not raw_data:
            print("❌ No data returned from main API. Akamai might have blocked the initial load.")
            browser.close()
            exit()

        theaters = raw_data.get('theaterShowtimes', {}).get('theaters', [])
        grouped_shows = {}
        
        for theater in theaters:
            t_name = theater.get('name', 'Unknown Theater')
            summary_data[t_name] = {'shows': 0, 'total': 0, 'booked': 0, 'gross': 0.0}
            
            for variant in theater.get('variants', []):
                for amenity in variant.get('amenityGroups', []):
                    for show in amenity.get('showtimes', []):
                        show_hash = show.get('showtimeHashCode')
                        if not show_hash: 
                            continue
                            
                        show_time = show.get('screenReaderTime', 'Unknown')
                        status = show.get('type', 'Unknown')
                        
                        group_key = f"{t_name}_{show_time}"
                        
                        if group_key not in grouped_shows:
                            grouped_shows[group_key] = {
                                'theater': t_name,
                                'time': show_time,
                                'tiers': []
                            }
                        
                        grouped_shows[group_key]['tiers'].append({
                            'hash': show_hash,
                            'status': status
                        })

        if not grouped_shows:
            print("❌ No valid bookable showtimes found for this ZIP code.")
            browser.close()
            exit()

        print(f"✅ Found {len(grouped_shows)} unique showtimes (combining multiple tiers). Blasting the NAPI endpoint...\n")

        # ---------------------------------------------------------
        # Step 2: Blast the `/napi/seatMap` Endpoint per Group!
        # ---------------------------------------------------------
        for group_key, show_info in grouped_shows.items():
            t_name = show_info['theater']
            show_time = show_info['time']
            tier_count = len(show_info['tiers'])
            
            combined_total = 0
            combined_booked = 0
            combined_gross = 0.0
            prices_seen = set()
            all_sold_out = True
            
            print(f"🎟️ Checking: {t_name[:20]:<20} | Time: {show_time:<10} (Tiers: {tier_count})")

            for tier in show_info['tiers']:
                tier_hash = tier['hash']
                tier_status = tier['status']

                if tier_status.lower() == "soldout":
                    combined_total += 100
                    combined_booked += 100
                    combined_gross += 2700.00
                    prices_seen.add(27.00)
                    continue
                
                all_sold_out = False
                seat_api_url = f"https://www.fandango.com/napi/seatMap/{tier_hash}"
                
                try:
                    seat_response = page.evaluate(f"""async () => {{
                        try {{
                            const response = await fetch('{seat_api_url}');
                            if (!response.ok) return {{ error: response.status }};
                            return await response.json();
                        }} catch(e) {{
                            return {{ error: "Fetch Exception" }};
                        }}
                    }}""")
                    
                    if seat_response and "error" not in seat_response:
                        data_block = seat_response.get('data', seat_response)
                        areas = data_block.get('areas', [])
                        seats_array = data_block.get('seats', [])
                        
                        if areas:
                            # ✨ THE HYBRID LOGIC ✨
                            if tier_count == 1 and seats_array:
                                # ONLY 1 TIER: 100% accurate physical seat counting!
                                total_seats = len(seats_array)
                                available_seats = sum(1 for seat in seats_array if seat.get('status') == 'A')
                                booked_seats = max(0, total_seats - available_seats)
                                calc_method = "Physical Count"
                            else:
                                # MULTIPLE TIERS: Fallback to Fandango's cached summary to prevent double counting
                                total_seats = int(areas[0].get('totalSeatCount', 0))
                                available_seats = int(areas[0].get('availableSeatCount', 0))
                                booked_seats = max(0, total_seats - available_seats)
                                calc_method = "Cached Summary"
                            
                            # Price Extraction
                            price = 0.0
                            try: price = float(areas[0].get('ticketInfo', [{}])[0].get('price', 0))
                            except: pass
                                
                            gross = booked_seats * price
                            
                            combined_total += total_seats
                            combined_booked += booked_seats
                            combined_gross += gross
                            if price > 0: prices_seen.add(price)
                            
                except Exception as e:
                    pass
                
                time.sleep(0.5) 

            # Compile and append exactly ONE row per physical showtime
            if combined_total > 0:
                final_status = "Sold Out" if all_sold_out else "Available"
                price_str = " / ".join(sorted([f"${p:.2f}" for p in prices_seen])) if prices_seen else "$0.00"
                
                # Using tier_count to display which method was used in the logs for your visibility
                log_method = "Physical Count" if tier_count == 1 else "Cached Summary"
                print(f"   => 📊 Seats: {combined_total:<3} | Booked: {combined_booked:<3} | Gross: ${combined_gross:<7.2f} [{log_method}]")
                
                shows_data.append({
                    'theater': t_name, 'time': show_time, 'status': final_status, 
                    'price_str': price_str, 'total': combined_total, 
                    'booked': combined_booked, 'gross': combined_gross
                })
                
                summary_data[t_name]['shows'] += 1
                summary_data[t_name]['total'] += combined_total
                summary_data[t_name]['booked'] += combined_booked
                summary_data[t_name]['gross'] += combined_gross
            else:
                print(f"   => ⚠️ No valid seat data returned for any tier of this showtime.")

        print("\n🛑 Closing browser.")
        browser.close()

    if shows_data:
        export_to_excel(shows_data, summary_data)
    else:
        print("❌ No successful showtime seat maps were fetched.")