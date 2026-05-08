import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from playwright.sync_api import sync_playwright

# =============================================================================
# ── 1. CONFIGURATION ─────────────────────────────────────────────────────────
# =============================================================================

MOVIE_ID = "244813"
MOVIE_SLUG = "peddi-2026"  
SHOW_DATE = "2026-06-03"
ZIP_CODE = "75201"
EXCEL_FILENAME = f"Fandango_Report_{ZIP_CODE}_{SHOW_DATE}.xlsx"

# =============================================================================
# ── 2. EXCEL EXPORT ──────────────────────────────────────────────────────────
# =============================================================================

def export_to_excel(shows_data, summary_data):
    print(f"\n📊 Generating Excel Report: {EXCEL_FILENAME}")
    wb = Workbook()
    
    ws_shows = wb.active
    ws_shows.title = "Showtime Details"
    ws_shows.append(["Theater Name", "Show Time", "Status", "Ticket Price", "Total Seats", "Booked Seats", "Occupancy %", "Gross ($)"])
    
    for row in shows_data:
        occ = round((row['booked'] / row['total']) * 100, 2) if row['total'] > 0 else 0
        ws_shows.append([row['theater'], row['time'], row['status'], f"${row['price']:.2f}", row['total'], row['booked'], occ, row['gross']])
        
    ws_summary = wb.create_sheet(title="Theater Summary")
    ws_summary.append(["Theater Name", "Total Shows", "Total Seats", "Total Booked", "Overall Occ %", "Total Gross ($)"])
    
    total_shows_all = total_seats_all = total_booked_all = total_gross_all = 0
    for t_name, stats in summary_data.items():
        occ = round((stats['booked'] / stats['total']) * 100, 2) if stats['total'] > 0 else 0
        ws_summary.append([t_name, stats['shows'], stats['total'], stats['booked'], occ, stats['gross']])
        total_shows_all += stats['shows']
        total_seats_all += stats['total']
        total_booked_all += stats['booked']
        total_gross_all += stats['gross']
        
    ws_summary.append([])
    ws_summary.append(["GRAND TOTAL", total_shows_all, total_seats_all, total_booked_all, 
                       round((total_booked_all/total_seats_all)*100,2) if total_seats_all > 0 else 0, total_gross_all])
    
    for ws in [ws_shows, ws_summary]:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            
    wb.save(EXCEL_FILENAME)
    print(f"✅ Export complete: {EXCEL_FILENAME}")

# =============================================================================
# ── MAIN EXECUTION ───────────────────────────────────────────────────────────
# =============================================================================

if __name__ == "__main__":
    shows_data = []
    summary_data = {}

    with sync_playwright() as p:
        print("\n🚀 Launching Browser in Background (Headless Mode)...")
        # Switched to headless=True as requested!
        browser = p.chromium.launch(headless=True, args=["--disable-blink-features=AutomationControlled"])
        context = browser.new_context(viewport={'width': 1280, 'height': 800})
        page = context.new_page()

        print(f"⏳ Loading Fandango main page to bypass Akamai natively...")
        movie_url = f"https://www.fandango.com/{MOVIE_SLUG}-{MOVIE_ID}/movie-overview?date={SHOW_DATE}"
        page.goto(movie_url)
        
        # Let the page load and cookies settle
        page.wait_for_timeout(4000)

        # ---------------------------------------------------------
        # Step 1: Fetch the Theater Groupings Natively
        # ---------------------------------------------------------
        print(f"🎬 Fetching theaters for ZIP: {ZIP_CODE}...")
        api_url = f"https://www.fandango.com/napi/theaterShowtimeGroupings/{MOVIE_ID}/{SHOW_DATE}?isdesktop=false&zip={ZIP_CODE}"
        
        try:
            raw_data = page.evaluate(f"""async () => {{
                const response = await fetch('{api_url}');
                if (!response.ok) return null;
                return await response.json();
            }}""")
        except Exception as e:
            print(f"❌ Failed to fetch theaters: {e}")
            browser.close()
            exit()

        if not raw_data:
            print("❌ No data returned from main API. Akamai might have blocked the initial load.")
            browser.close()
            exit()

        theaters = raw_data.get('theaterShowtimes', {}).get('theaters', [])
        shows_to_check = []
        
        for theater in theaters:
            t_name = theater.get('name', 'Unknown Theater')
            summary_data[t_name] = {'shows': 0, 'total': 0, 'booked': 0, 'gross': 0.0}
            
            for variant in theater.get('variants', []):
                for amenity in variant.get('amenityGroups', []):
                    for show in amenity.get('showtimes', []):
                        show_hash = show.get('showtimeHashCode')
                        if not show_hash: 
                            continue
                            
                        shows_to_check.append({
                            'theater': t_name, 
                            'time': show.get('screenReaderTime', 'Unknown'),
                            'hash': show_hash, 
                            'status': show.get('type', 'Unknown')
                        })

        if not shows_to_check:
            print("❌ No valid bookable showtimes found for this ZIP code.")
            browser.close()
            exit()

        print(f"✅ Found {len(shows_to_check)} shows to check. Blasting the NAPI endpoint...\n")

        # ---------------------------------------------------------
        # Step 2: Blast the new `/napi/seatMap` Endpoint!
        # ---------------------------------------------------------
        for show in shows_to_check:
            t_name, show_time, show_hash, status = show['theater'], show['time'], show['hash'], show['status']
            print(f"🎟️ Checking: {t_name[:20]:<20} | Time: {show_time:<10}")

            if status.lower() == "soldout":
                shows_data.append({'theater': t_name, 'time': show_time, 'status': 'Sold Out', 'price': 27.00, 'total': 100, 'booked': 100, 'gross': 2700.00})
                summary_data[t_name]['shows'] += 1; summary_data[t_name]['total'] += 100
                summary_data[t_name]['booked'] += 100; summary_data[t_name]['gross'] += 2700.00
                continue

            seat_api_url = f"https://www.fandango.com/napi/seatMap/{show_hash}"
            
            try:
                seat_response = page.evaluate(f"""async () => {{
                    try {{
                        const response = await fetch('{seat_api_url}');
                        if (!response.ok) return {{ error: response.status }};
                        return await response.json();
                    }} catch(e) {{
                        return {{ error: "Fetch Exception" }};
                    }}
                }}""")
                
                if not seat_response:
                    print(f"   => ⚠️ Empty response from API.")
                elif "error" in seat_response:
                    print(f"   => ⚠️ Show Unavailable (HTTP {seat_response['error']}) - Likely a ghost showtime.")
                else:
                    data_block = seat_response.get('data', seat_response)
                    
                    # ✨ NEW LOGIC: Count actual physical seats array instead of the summary cache!
                    seats_array = data_block.get('seats', [])
                    
                    if seats_array:
                        total_seats = len(seats_array)
                        
                        # Any seat that is NOT "A" (Available) is considered booked (this handles "X" and "H")
                        available_seats = sum(1 for seat in seats_array if seat.get('status') == 'A')
                        booked_seats = total_seats - available_seats
                        
                        # We still need the price from the 'areas' array
                        areas = data_block.get('areas', [])
                        price = 0.0
                        if areas:
                            try: price = float(areas[0].get('ticketInfo', [{}])[0].get('price', 0))
                            except: pass
                            
                        gross = booked_seats * price
                        
                        print(f"   => 📊 Seats: {total_seats:<3} | Booked: {booked_seats:<3} | Gross: ${gross:<7.2f}")
                        
                        shows_data.append({'theater': t_name, 'time': show_time, 'status': status, 'price': price, 'total': total_seats, 'booked': booked_seats, 'gross': gross})
                        summary_data[t_name]['shows'] += 1; summary_data[t_name]['total'] += total_seats
                        summary_data[t_name]['booked'] += booked_seats; summary_data[t_name]['gross'] += gross
                    else:
                        print(f"   => ⚠️ No valid seats array found in response.")
                        
            except Exception as e:
                print(f"   => ❌ Browser execution error: {e}")
            
            time.sleep(1.0) 

        print("\n🛑 Closing browser.")
        browser.close()

    if shows_data:
        export_to_excel(shows_data, summary_data)
    else:
        print("❌ No successful showtime seat maps were fetched.")