import json
import os
import time
import random
from datetime import datetime, timedelta
from base64 import b64decode
from Crypto.Cipher import AES
from Crypto.Util.Padding import unpad
from fake_useragent import UserAgent
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook
import difflib
from collections import defaultdict
import math
from concurrent.futures import ThreadPoolExecutor, as_completed

from utils.generateHybridCityImageReport import generate_hybrid_city_image_report

# ================= ORIGINAL (LEFT AS-IS) =================
processed_sids = set()

# ================= NEW: ISOLATED TRACKING =================
processed_district_sids = set()
processed_bms_sids = set()

# ================= CONFIGURATION =================
DISTRICT_URL = "https://www.district.in/movies/orange-2010-movie-tickets-in-hyderabad-MV160920"
BMS_URL = "https://in.bookmyshow.com/movies/hyderabad/orange/buytickets/ET00005527/20260209"

SHOW_DATE = "2026-02-09"
DISTRICT_FULL_URL = f"{DISTRICT_URL}?fromdate={SHOW_DATE}"

BMS_KEY = "kYp3s6v9y$B&E)H+MbQeThWmZq4t7w!z"
BOOKED_CODES = {"2"}
SLEEP_TIME = 1.0
MAX_WORKERS = 3

# ================= SHARED DRIVER =================
def get_driver():
    ua = UserAgent()
    options = Options()
    options.add_argument(f"user-agent={ua.random}")
    options.add_argument("--headless")
    options.add_argument("start-maximized")
    options.add_argument("--disable-web-security")
    options.add_argument("--disable-site-isolation-trials")
    options.add_argument("disable-csp")
    return webdriver.Chrome(options=options)

# ================= NEW: TIME NORMALIZATION HELPERS =================
def district_gmt_to_ist(dt_str):
    gmt = datetime.fromisoformat(dt_str)
    ist = gmt + timedelta(hours=5, minutes=30)
    return ist.strftime("%Y-%m-%d %H:%M")

def normalize_bms_time(show_date, show_time):
    dt = datetime.strptime(f"{show_date} {show_time}", "%Y-%m-%d %I:%M %p")
    return dt.strftime("%Y-%m-%d %H:%M")

def build_seat_signature(seat_map):
    """
    Build a source-agnostic seat signature.
    Ignores category names/codes.
    Uses only sorted seat counts.
    """
    counts = sorted(seat_map.values())
    return "|".join(str(c) for c in counts)


# ================= DISTRICT LOGIC =================
def fetch_district_data(driver):
    print("\n🚀 STARTING DISTRICT FETCH...")
    results = []

    try:
        driver.get(DISTRICT_FULL_URL)
        html = driver.page_source
        time.sleep(2)
        marker = 'id="__NEXT_DATA__"'
        start = html.find(marker)
        if start == -1:
            return []

        start = html.find('>', start) + 1
        end = html.find('</script>', start)
        data = json.loads(html[start:end])

        sessions = data['props']['pageProps']['data']['serverState']['movieSessions']
        key = list(sessions.keys())[0]
        cinemas = sessions[key].get('arrangedSessions', [])

        for cin in cinemas:
            venue = cin['entityName']

            for s in cin.get('sessions', []):
                sid = str(s.get('sid', ''))

                # ===== NEW: ISOLATED DEDUPE =====
                if sid in processed_district_sids:
                    continue
                processed_district_sids.add(sid)

                b_gross, p_gross, b_tkts, t_tkts = 0, 0, 0, 0
                seat_map = {}
                price_seat_map = defaultdict(int)
                price_seat_list = []

                for a in s.get('areas', []):
                    tot, av, pr = a['sTotal'], a['sAvail'], a['price']
                    bk = tot - av
                    seat_map[a['label']] = tot

                    b_tkts += bk
                    t_tkts += tot
                    b_gross += (bk * pr)
                    p_gross += (tot * pr)
                    price_seat_map[float(pr)] += tot
                    price_seat_list.append((float(pr), tot))

                occ = round((b_tkts / t_tkts) * 100, 2) if t_tkts else 0
                normalized_time = district_gmt_to_ist(s['showTime'])

                print(
                    f"   🎬 {venue[:20]:<20} | {normalized_time} | "
                    f"Occ: {occ:>5}% | Gross: ₹{b_gross:<8,}"
                )

                results.append({
                    "source": "district",
                    "sid": sid,
                    "venue": venue,
                    "showTime": s['showTime'],
                    "normalized_show_time": normalized_time,
                    "seat_category_map": seat_map,
                    "price_seat_map": dict(price_seat_map),
                    "price_seat_signature": sorted(price_seat_list),
                    "seat_signature": build_seat_signature(seat_map),
                    "total_tickets": abs(t_tkts),
                    "booked_tickets": min(abs(b_tkts), abs(t_tkts)),
                    "total_gross": abs(p_gross),
                    "booked_gross": min(abs(int(b_gross)), abs(int(p_gross))),
                    "occupancy": min(100, abs(occ)),
                    "is_fallback": False  # District is always real
                })

        print(f"✅ District: Found {len(results)} shows.")

    except Exception as e:
        print(f"❌ District Error: {e}")

    return results

# ================= BMS LOGIC =================
def decrypt_data(enc):
    try:
        decoded = b64decode(enc)
        cipher = AES.new(BMS_KEY.encode(), AES.MODE_CBC, iv=bytes(16))
        return unpad(cipher.decrypt(decoded), AES.block_size).decode()
    except:
        return None

def calculate_bms_collection(decrypted, price_map):
    header, rows_part = decrypted.split("||")
    rows = rows_part.split("|")

    # Category Map
    cat_map = {}
    for p in header.split("|"):
        parts = p.split(":")
        if len(parts) >= 3:
            cat_map[parts[1]] = parts[2]

    seats, booked = {}, {}
    for row in rows:
        if not row:
            continue
        parts = row.split(":")
        if len(parts) < 3:
            continue
        block = parts[3][0] if len(parts) > 3 else parts[2][0]
        area = cat_map.get(block)
        if not area:
            continue

        for seat in parts:
            if len(seat) < 2:
                continue
            status = seat[1]
            if seat[0] == block and status in ("1", "2"):
                seats[area] = seats.get(area, 0) + 1
            if status in BOOKED_CODES:
                booked[area] = booked.get(area, 0) + 1

    t_tkts, b_tkts, t_gross, b_gross = 0, 0, 0, 0
    for area, total in seats.items():
        bk = booked.get(area, 0)
        pr = price_map.get(area, 0)
        t_tkts += total
        b_tkts += bk
        t_gross += total * pr
        b_gross += bk * pr

    occ = round((b_tkts / t_tkts) * 100, 2) if t_tkts else 0

    # >>> NEW: return seats map also
    return t_tkts, b_tkts, int(t_gross), int(b_gross), occ, seats


def get_seat_layout(driver, venue_code, session_id):
    """ Returns tuple: (EncryptedDataString, ErrorMessage) """
    api = "https://services-in.bookmyshow.com/doTrans.aspx"
    js = f"""
    var cb = arguments[0]; var x = new XMLHttpRequest();
    x.open("POST", "{api}", true);
    x.setRequestHeader("Content-Type", "application/x-www-form-urlencoded");
    x.onload = function() {{ cb(x.responseText); }};
    x.send("strCommand=GETSEATLAYOUT&strAppCode=WEB&strVenueCode={venue_code}&strParam1={session_id}&strParam2=WEB&strParam5=Y&strFormat=json");
    """

    max_retries = 2
    for attempt in range(max_retries + 1):
        try:
            resp = driver.execute_async_script(js)
            j_resp = json.loads(resp).get("BookMyShow", {})

            if j_resp.get("blnSuccess") == "true":
                return j_resp.get("strData"), None

            error_msg = j_resp.get("strException", "")
            if "Rate limit" in error_msg:
                if attempt < max_retries:
                    time.sleep(60)
                    continue
                else:
                    return None, "Rate limit exceeded"

            return None, error_msg

        except Exception as e:
            return None, str(e)

    return None, "Unknown Error"


def process_venue_list(venues):
    results = []
    driver = get_driver()
    
    try:
        for v in venues:
            v_name = v["additionalData"]["venueName"]
            v_code = v["additionalData"]["venueCode"]

            # 1. Init Capacity Map for THIS Venue
            screen_capacity_map = {}

            # 2. Sort Shows: Available First, Sold Out Last
            shows = v.get("showtimes", [])
            shows.sort(
                key=lambda s: s["additionalData"].get("availStatus", "0"),
                reverse=True
            )

            for show in shows:
                sid = str(show["additionalData"]["sessionId"])
                show_time = show["title"]

                # ===== NEW: ISOLATED DEDUPE =====
                if sid in processed_bms_sids:
                    continue
                
                # OPTIMIZATION: Skip if already found in District
                # if sid in processed_district_sids:
                #     print(f"   ⏭️  Skipping {sid} (Found in District)")
                #     continue

                processed_bms_sids.add(sid)

                raw_screen = show.get("screenAttr", "")
                screenName = raw_screen if raw_screen else "Main Screen"

                cats = show["additionalData"].get("categories", [])
                price_map = {
                    c["areaCatCode"]: float(c["curPrice"])
                    for c in cats
                }

                try:
                    enc, error_msg = get_seat_layout(driver, v_code, sid)

                    data = None
                    soldOut = False
                    seat_map = {}
                    is_fallback = False
                    price_seat_map = {}

                    if not enc:
                        if not price_map:
                            continue

                        max_price = max(price_map.values())
                        is_fallback = True

                        # Populate price_seat_map for fallback matching
                        for p in price_map.values():
                            price_seat_map[float(p)] = 0

                        if error_msg and "sold out" in error_msg.lower():
                            # Case 1: Sold Out -> Try Recovery or Fallback
                            recovered_capacity = None
                            recovered_seat_map = None
                            
                            try:
                                base_sid = int(sid)
                                # Try offsets +7 down to +1
                                for offset in range(7, 0, -1):
                                    target_sid = str(base_sid + offset)
                                    time.sleep(1)
                                    n_enc, n_err = get_seat_layout(driver, v_code, target_sid)
                                    if n_enc:
                                        n_dec = decrypt_data(n_enc)
                                        n_res = calculate_bms_collection(n_dec, {})
                                        if n_res[0] > 0:
                                            recovered_capacity = n_res[0]
                                            recovered_seat_map = n_res[5]
                                            print(f"   ✨ Fixed SoldOut {sid} using {target_sid} (Cap: {recovered_capacity})")
                                            break
                            except Exception: pass
                            
                            if recovered_capacity:
                                calc_gross = 0
                                for ac, count in recovered_seat_map.items():
                                    calc_gross += count * price_map.get(ac, 0)
                                
                                if calc_gross > 0:
                                    t_tkts = recovered_capacity; b_tkts = recovered_capacity
                                    t_gross = calc_gross; b_gross = calc_gross
                                    screen_capacity_map[screenName] = t_tkts
                                    
                                    # Update maps for better data
                                    seat_map = recovered_seat_map
                                    ps_map = defaultdict(int)
                                    for ac, count in seat_map.items():
                                        ps_map[float(price_map.get(ac, 0))] += count
                                    price_seat_map = dict(ps_map)
                                else:
                                    recovered_capacity = None

                            if not recovered_capacity:
                                if screenName in screen_capacity_map:
                                    FALLBACK_SEATS = screen_capacity_map[screenName]
                                    print(f"   ⚡ Smart Fallback: Using {FALLBACK_SEATS} seats for {screenName}")
                                else:
                                    FALLBACK_SEATS = 400
                                    print(f"   ⚠️ No history for {screenName}, using default {FALLBACK_SEATS}")
                                t_tkts = FALLBACK_SEATS; b_tkts = FALLBACK_SEATS
                                t_gross = int(t_tkts * max_price); b_gross = t_gross

                            occ = 100.0
                            soldOut = True

                            data = {
                                "total_tickets": t_tkts,
                                "booked_tickets": b_tkts,
                                "total_gross": t_gross,
                                "booked_gross": b_gross,
                                "occupancy": occ
                            }

                        elif error_msg and "Rate limit" in error_msg:
                            print(f"   ⚠️ Skipping {v_name[:15]} (Rate Limit)")
                            continue

                        else:
                            t_tkts = 400
                            b_tkts = 200
                            t_gross = int(t_tkts * max_price)
                            b_gross = int(b_tkts * max_price)
                            occ = 50.0

                            data = {
                                "total_tickets": t_tkts,
                                "booked_tickets": b_tkts,
                                "total_gross": t_gross,
                                "booked_gross": b_gross,
                                "occupancy": occ
                            }

                    else:
                        decrypted = decrypt_data(enc)
                        res = calculate_bms_collection(decrypted, price_map)

                        data = {
                            "total_tickets": res[0],
                            "booked_tickets": res[1],
                            "total_gross": res[2],
                            "booked_gross": res[3],
                            "occupancy": res[4]
                        }

                        seat_map = res[5]

                        if data["total_tickets"] > 0:
                            # Build Price Seat Map for Matching
                            ps_map = defaultdict(int)
                            ps_list = []
                            for ac, count in seat_map.items():
                                pr = float(price_map.get(ac, 0))
                                ps_map[pr] += count
                                ps_list.append((pr, count))
                            price_seat_map = dict(ps_map)
                            data["price_seat_signature"] = sorted(ps_list)
                            screen_capacity_map[screenName] = data["total_tickets"]

                    if data:
                        normalized_time = normalize_bms_time(SHOW_DATE, show_time)
                        tag = "(SOLD OUT)" if soldOut else ""

                        print(
                            f"   🎬 {v_name[:15]:<15} | {normalized_time} | "
                            f"Occ: {data['occupancy']:>5}% | "
                            f"Gross: ₹{data['booked_gross']:<8,} {tag}"
                        )

                        results.append({
                            "source": "bms",
                            "sid": str(sid),
                            "venue": v_name,
                            "showTime": show_time,
                            "normalized_show_time": normalized_time,
                            "seat_category_map": seat_map,
                            "price_seat_map": price_seat_map,
                            "price_seat_signature": data.get("price_seat_signature", []),
                            "seat_signature": build_seat_signature(seat_map),
                            "total_tickets": abs(data["total_tickets"]),
                            "booked_tickets": min(
                                abs(data["booked_tickets"]),
                                abs(data["total_tickets"])
                            ),
                            "total_gross": abs(data["total_gross"]),
                            "booked_gross": min(
                                abs(data["booked_gross"]),
                                abs(data["total_gross"])
                            ),
                            "occupancy": min(100, abs(data["occupancy"])),
                            "is_fallback": is_fallback
                        })

                except Exception:
                    pass

                time.sleep(SLEEP_TIME)
    except Exception as e:
        print(f"❌ Worker Error: {e}")
    finally:
        driver.quit()
    return results

def fetch_bms_data():
    print("\n🚀 STARTING BMS FETCH...")
    results = []
    driver = get_driver()

    try:
        driver.get(BMS_URL)
        time.sleep(2.5)
        html = driver.page_source

        marker = "window.__INITIAL_STATE__"
        start = html.find(marker)
        if start == -1:
            print("   ⚠️ BMS: Could not find initial state (possible bot detection)")
            return []

        start = html.find("{", start)
        brace, end = 0, start
        while end < len(html):
            if html[end] == "{":
                brace += 1
            elif html[end] == "}":
                brace -= 1
            if brace == 0:
                break
            end += 1

        state_data = json.loads(html[start:end + 1])

        venues = []
        try:
            sbe = state_data.get("showtimesByEvent")
            dc = sbe.get("currentDateCode")
            widgets = sbe["showDates"][dc]["dynamic"]["data"]["showtimeWidgets"]
            for w in widgets:
                if w.get("type") == "groupList":
                    for g in w["data"]:
                        if g.get("type") == "venueGroup":
                            venues = g["data"]
        except:
            venues = []

        # Close the initial driver as workers will create their own
        driver.quit()

        if not venues:
            return []

        print(f"   🚀 Launching {MAX_WORKERS} workers for {len(venues)} venues...")
        
        # Split venues into chunks
        chunk_size = math.ceil(len(venues) / MAX_WORKERS)
        venue_chunks = [venues[i:i + chunk_size] for i in range(0, len(venues), chunk_size)]
        
        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = [executor.submit(process_venue_list, chunk) for chunk in venue_chunks]
            
            for future in as_completed(futures):
                results.extend(future.result())
        
        print(f"✅ BMS: Found {len(results)} shows.")

    except Exception as e:
        print(f"❌ BMS Error: {e}")
        # Driver is already quit or will be quit by workers

    return results

# ================= EXCEL GENERATOR =================
def generate_excel(data, filename):
    wb = Workbook()
    reports_dir = "reports"
    os.makedirs(reports_dir, exist_ok=True)

    # 1. THEATRE WISE
    ws_th = wb.active
    ws_th.title = "Theatre Wise"
    ws_th.append([
        "Venue", "Shows", "Total Seats",
        "Booked Seats", "Total Gross ₹", "Booked Gross ₹", "Occ %"
    ])

    th_map = {}
    for r in data:
        k = r["venue"]
        if k not in th_map:
            th_map[k] = {
                "shows": 0,
                "t_seats": 0,
                "b_seats": 0,
                "p_gross": 0,
                "b_gross": 0
            }

        d = th_map[k]
        d["shows"] += 1
        d["t_seats"] += r["total_tickets"]
        d["b_seats"] += r["booked_tickets"]
        d["p_gross"] += r["total_gross"]
        d["b_gross"] += r["booked_gross"]

    for v, d in th_map.items():
        occ = round(
            (d["b_seats"] / d["t_seats"]) * 100, 2
        ) if d["t_seats"] else 0

        ws_th.append([
            v, d["shows"], d["t_seats"],
            d["b_seats"], d["p_gross"], d["b_gross"], occ
        ])

    # 2. SHOW WISE
    ws_show = wb.create_sheet(title="Show Wise")
    ws_show.append([
        "Source", "Venue", "Time", "SID",
        "Total Seats", "Booked Seats",
        "Total Gross ₹", "Booked Gross ₹", "Occ %"
    ])

    for r in data:
        ws_show.append([
            r["source"], r["venue"], r["normalized_show_time"], r["sid"],
            r["total_tickets"], r["booked_tickets"],
            r["total_gross"], r["booked_gross"], r["occupancy"]
        ])

    # 3. SUMMARY
    ws_sum = wb.create_sheet(title="Summary")

    agg_p_gross = sum(r["total_gross"] for r in data)
    agg_b_gross = sum(r["booked_gross"] for r in data)
    agg_t_seats = sum(r["total_tickets"] for r in data)
    agg_b_seats = sum(r["booked_tickets"] for r in data)

    occ = round(
        (agg_b_seats / agg_t_seats) * 100, 2
    ) if agg_t_seats else 0

    ws_sum.append(["Metric", "Value"])
    ws_sum.append(["Total Theatres", len(th_map)])
    ws_sum.append(["Total Shows", len(data)])
    ws_sum.append(["Total Seats", agg_t_seats])
    ws_sum.append(["Booked Tickets", agg_b_seats])
    ws_sum.append(["Total Potential Gross", agg_p_gross])
    ws_sum.append(["Total Booked Gross", agg_b_gross])
    ws_sum.append(["Overall Occupancy %", occ])
    ws_sum.append([
        "Generated At",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ])

    path = os.path.join(reports_dir, filename)
    wb.save(path)
    print(f"📊 Report Saved: {path}")


# ================= EXECUTION =================
if __name__ == "__main__":
    d_driver = get_driver()
    dist_data = fetch_district_data(d_driver)
    d_driver.quit()

    bms_data = fetch_bms_data()

    print("\n🔄 Merging Data (Cross Source Dedupe)...")

    # ================= ROBUST MERGE LOGIC =================
    final_data = []
    SEAT_TOLERANCE = 5
    
    # Index District data by normalized time for fast lookup
    district_by_time = defaultdict(list)
    for r in dist_data:
        district_by_time[r['normalized_show_time']].append(r)

    for bms in bms_data:
        # Get potential matches from District at the same time
        candidates = district_by_time.get(bms['normalized_show_time'], [])
        match_found = None

        # 1. Try EXACT SID Match (Highest Confidence)
        for cand in candidates:
            if cand['sid'] == bms['sid']:
                match_found = cand
                print(f"   🔗 SID Match: {bms['sid']}")
                break

        # 2. Try Price Seat Signature Match (Exact Structure + Tolerance)
        # Only applicable if BMS is not a fallback (since fallback has no seat details)
        if not match_found and not bms.get('is_fallback', False):
            b_sig = bms.get('price_seat_signature', [])
            bms_venue_clean = bms['venue'].lower()
            
            for cand in candidates:
                d_sig = cand.get('price_seat_signature', [])
                if not b_sig or not d_sig: continue
                if len(b_sig) != len(d_sig): continue

                # Check if each pair matches (Price exact, Seats within tolerance)
                # Both lists are sorted, so corresponding indices must match
                if all(bp == dp and abs(bs - ds) <= SEAT_TOLERANCE for (bp, bs), (dp, ds) in zip(b_sig, d_sig)):
                    # Also check fuzzy venue name to avoid false positives with same capacity
                    ratio = difflib.SequenceMatcher(None, bms_venue_clean, cand['venue'].lower()).ratio()
                    if ratio > 0.4:
                        match_found = cand
                        print(f"   🔗 Price/Seat Sig Match: {bms['venue'][:15]}... == {cand['venue'][:15]}... (Tol: {SEAT_TOLERANCE}, Ratio: {int(ratio*100)}%)")
                        break
        
        # 3. Try FUZZY Venue Match (Fallback for Sold Out/Error BMS shows)
        if not match_found and candidates:
            best_ratio = 0
            best_cand = None
            bms_venue_clean = bms['venue'].lower()
            b_prices = set(bms.get('price_seat_map', {}).keys())
            
            for cand in candidates:
                # Check price match (Strict Price Category Match)
                d_prices = set(cand.get('price_seat_map', {}).keys())
                if b_prices != d_prices:
                    continue

                # Check similarity of venue names
                ratio = difflib.SequenceMatcher(None, bms_venue_clean, cand['venue'].lower()).ratio()
                # Threshold 0.5 handles "Sangam Theatre" vs "Sangam Theatre 4K..." well
                if ratio > 0.5 and ratio > best_ratio:
                    best_ratio = ratio
                    best_cand = cand
            
            if best_cand:
                match_found = best_cand
                print(f"   🔗 Fuzzy Match: {bms['venue'][:15]}... == {match_found['venue'][:15]}... ({int(best_ratio*100)}%)")

        if match_found:
            # Remove from pool so we don't match it again
            candidates.remove(match_found)
            
            # DECISION TIME: Which data is better?
            if bms.get('is_fallback', False):
                # BMS is an estimate/fallback. District is real. KEEP DISTRICT.
                final_data.append(match_found)
            else:
                # Both are real. Check who has higher gross.
                if bms['booked_gross'] > match_found['booked_gross']:
                    # BMS wins. Update match_found (District object) with BMS stats.
                    # This preserves the District Venue Name.
                    match_found['total_tickets'] = bms['total_tickets']
                    match_found['booked_tickets'] = bms['booked_tickets']
                    match_found['total_gross'] = bms['total_gross']
                    match_found['booked_gross'] = bms['booked_gross']
                    match_found['occupancy'] = bms['occupancy']
                    match_found['seat_category_map'] = bms['seat_category_map']
                    match_found['price_seat_map'] = bms['price_seat_map']
                    match_found['seat_signature'] = bms['seat_signature']
                    
                    final_data.append(match_found)
                else:
                    # District wins. Keep as is.
                    final_data.append(match_found)
        else:
            # No match found in District, keep BMS unique entry
            final_data.append(bms)

    # Add any remaining unmatched District shows
    for sublist in district_by_time.values():
        final_data.extend(sublist)

    if final_data:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        generate_excel(
            final_data,
            f"Total_City_Report_{ts}.xlsx"
        )
        generate_hybrid_city_image_report(
            final_data,
            DISTRICT_FULL_URL,
            f"reports/Total_City_Report_{ts}.png"
        )
    else:
        print("❌ No data collected.")
